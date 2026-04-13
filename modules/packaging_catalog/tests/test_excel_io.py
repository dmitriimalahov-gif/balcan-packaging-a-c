# -*- coding: utf-8 -*-
from pathlib import Path

import pytest
from openpyxl import Workbook

from modules.packaging_catalog.application.excel_io import (
    load_rows_from_excel,
    save_rows_to_excel,
)
from modules.packaging_catalog.domain.makety_excel_config import HEADERS


def _write_canonical_sheet(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    for c, h in enumerate(HEADERS, start=1):
        ws.cell(1, c, value=h)
    vals = [
        "k",
        "cat",
        "lac",
        "50x50",
        "box",
        "x.pdf",
        "10x10",
        "C01",
        "9",
        "8",
        "7",
        "6",
        "cutii name",
    ]
    for c, v in enumerate(vals, start=1):
        ws.cell(2, c, value=v)
    wb.save(path)


def test_load_rows_canonical(tmp_path: Path) -> None:
    p = tmp_path / "m.xlsx"
    _write_canonical_sheet(p)
    rows = load_rows_from_excel(p)
    assert len(rows) == 1
    r = rows[0]
    assert r["excel_row"] == 2
    assert r["name"] == "cutii name"
    assert r["size"] == "50x50"
    assert r["kind"] == "box"
    assert r["file"] == "x.pdf"


def test_load_rows_strict_raises(tmp_path: Path) -> None:
    p = tmp_path / "bad.xlsx"
    wb = Workbook()
    wb.active.cell(1, 1, value="only")
    wb.save(p)
    with pytest.raises(ValueError, match="эталону"):
        load_rows_from_excel(p, strict_reference_layout=True)


def test_save_rows_roundtrip(tmp_path: Path) -> None:
    p = tmp_path / "out.xlsx"
    _write_canonical_sheet(p)
    item = {
        "excel_row": 2,
        "name": "NewCutii",
        "size": "1x2x3",
        "kind": "Коробка",
        "file": "a.pdf",
        "price": "10",
        "price_new": "11",
        "qty_per_sheet": "4",
        "qty_per_year": "999",
    }
    save_rows_to_excel(p, [item], db_path=None)
    rows = load_rows_from_excel(p)
    assert len(rows) == 1
    assert rows[0]["name"] == "NewCutii"
    assert rows[0]["size"] == "1x2x3"


def test_save_rows_default_enrich_missing_db_clears_cg(tmp_path: Path) -> None:
    """При db_path без файла apply_makety_cg_derived_from_db обнуляет служебные поля CG."""
    p = tmp_path / "out.xlsx"
    _write_canonical_sheet(p)
    item = {
        "excel_row": 2,
        "name": "X",
        "size": "1x1",
        "kind": "Коробка",
        "file": "a.pdf",
        "price": "99",
        "price_new": "",
        "qty_per_sheet": "1",
        "qty_per_year": "1000",
        "_cg_cutit_no": "should_clear",
    }
    missing_db = tmp_path / "no_such.db"
    save_rows_to_excel(p, [item], db_path=missing_db)
    assert item["_cg_cutit_no"] == ""
    rows = load_rows_from_excel(p)
    assert len(rows) == 1
