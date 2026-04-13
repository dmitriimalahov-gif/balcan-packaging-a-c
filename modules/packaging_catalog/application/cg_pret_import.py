# -*- coding: utf-8 -*-
"""
Парсинг Excel прайса Cutting Group (лист «cutii»): ножи, ценовые ступени.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import IO, Any

import openpyxl

CG_TIER_COLS: dict[str, list[tuple[int, int, int | None]]] = {
    "lac_wb": [
        (4, 1000, 4000),
        (5, 5000, 10000),
        (6, 11000, 50000),
        (7, 50000, None),
        (8, 100000, None),
        (9, 200000, None),
    ],
    "uv_no_foil": [
        (10, 1000, 4000),
        (11, 5000, 10000),
        (12, 11000, 50000),
        (13, 50000, None),
    ],
    "uv_foil": [
        (14, 1000, 4000),
        (15, 5000, 10000),
        (16, 11000, 50000),
        (17, 50000, None),
    ],
}


def clean_price_token(tok: str) -> float | None:
    t = re.sub(r"\([^)]*\)", "", tok).strip()
    t = t.split("/")[0].strip()
    t = t.replace("\u00a0", "").replace(" ", "")
    m = re.match(r"^[\d,\.]+", t)
    if not m:
        return None
    t = m.group(0).replace(",", ".").strip(".").strip()
    if not t:
        return None
    try:
        v = float(t)
        return v if v > 0.5 else None
    except (ValueError, TypeError):
        return None


def parse_cg_price(val: Any) -> tuple[float | None, float | None]:
    """(old_price, new_price). Первое число — старая, последнее — новая."""
    if val is None:
        return (None, None)
    if isinstance(val, (int, float)):
        v = float(val) if val > 0 else None
        return (v, v)
    s = str(val).strip()
    if not s:
        return (None, None)
    lines = [ln.strip() for ln in s.replace("\n", "\n").split("\n") if ln.strip()]
    if len(lines) == 1:
        parts = re.split(r"\s{3,}", lines[0])
        if len(parts) > 1:
            lines = [p.strip() for p in parts if p.strip()]
    nums: list[float] = []
    for ln in lines:
        p = clean_price_token(ln)
        if p is not None and p >= 5.0:
            nums.append(p)
    if not nums:
        return (None, None)
    if len(nums) == 1:
        return (nums[0], nums[0])
    return (nums[0], nums[-1])


def parse_cg_pret_workbook(
    source: Path | IO[bytes],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """
    Возвращает ``(knives, prices)`` из Excel CG Preț.

    *knives*: ``[{cutit_no, name, category, cardboard}, ...]``
    *prices*: ``[{cutit_no, finish_type, min_qty, max_qty, price_per_1000, price_old_per_1000}, ...]``
    """
    wb = openpyxl.load_workbook(source, data_only=True, read_only=True)
    ws = None
    for sn in wb.sheetnames:
        if "cutii" in sn.lower():
            ws = wb[sn]
            break
    if ws is None:
        ws = wb.active

    knives: list[dict[str, Any]] = []
    prices: list[dict[str, Any]] = []
    current_cat = ""

    for _ri, row in enumerate(ws.iter_rows(min_row=4, values_only=True)):
        vals = list(row)
        col_a = str(vals[0] or "").strip() if vals else ""
        col_b = str(vals[1] or "").strip() if len(vals) > 1 else ""
        col_c = str(vals[2] or "").strip() if len(vals) > 2 else ""
        col_d = str(vals[3] or "").strip() if len(vals) > 3 else ""

        if col_a and not col_a.replace(".", "").replace(" ", "").isdigit():
            current_cat = col_a
            continue
        if not col_b:
            continue

        cutit = col_b.split("\n")[0].strip()
        if not cutit:
            continue
        knives.append(
            {
                "cutit_no": cutit,
                "name": col_c.replace("\n", " ").strip() if col_c else "",
                "category": current_cat,
                "cardboard": col_d,
            }
        )

        for ft, tier_cols in CG_TIER_COLS.items():
            for col_idx, mn, mx in tier_cols:
                if col_idx < len(vals):
                    pv_old, pv_new = parse_cg_price(vals[col_idx])
                    if pv_new is not None:
                        prices.append(
                            {
                                "cutit_no": cutit,
                                "finish_type": ft,
                                "min_qty": mn,
                                "max_qty": mx,
                                "price_per_1000": pv_new,
                                "price_old_per_1000": pv_old,
                            }
                        )
    wb.close()
    return knives, prices
