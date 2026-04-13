# -*- coding: utf-8 -*-
"""
Чтение и запись листа «Макеты» в Excel (openpyxl), без Streamlit.

При ``db_path`` не ``None``: по умолчанию перед записью вызывается
``apply_makety_cg_derived_from_db`` из ``makety_cg_enrichment``.
Передайте свой ``enrich_from_db`` для другой логики; чтобы не обогащать
(строки уже готовы), передайте, например, ``lambda _p, _r: None``.
"""

from __future__ import annotations

import os
import tempfile
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook

from modules.packaging_catalog.application.excel_headers import (
    build_makety_column_index_map,
    excel_cell_str,
    excel_header_is_makety_v3,
    excel_row_dict_from_column_map,
)
from modules.packaging_catalog.application.makety_cg_enrichment import (
    apply_makety_cg_derived_from_db,
)
from modules.packaging_catalog.domain.makety_excel_config import HEADERS, MAKETY_EXCEL_NCOLS

EnrichRowsFromDb = Callable[[Path, list[dict[str, Any]]], None]


def _enrich_rows_if_db(
    db_path: Path | None,
    rows: list[dict[str, Any]],
    enrich_from_db: EnrichRowsFromDb | None,
) -> None:
    if db_path is None:
        return
    fn = enrich_from_db if enrich_from_db is not None else apply_makety_cg_derived_from_db
    fn(db_path, rows)


def load_rows_from_excel(
    excel_path: Path,
    *,
    strict_reference_layout: bool = False,
) -> list[dict[str, Any]]:
    """
    Читает активный лист. Если в первой строке найдены все 13 заголовков HEADERS
    (порядок столбцов может быть любым), строки разбираются по этим именам.

    strict_reference_layout=True: при отсутствии полного набора заголовков — ValueError.
    """
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    hdr_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    hdr_t = tuple(hdr_row) if hdr_row else tuple()
    col_map = build_makety_column_index_map(hdr_t)
    rows_out: list[dict[str, Any]] = []

    if col_map is not None:
        for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None or all(v is None for v in row):
                continue
            row_t = tuple(row) if row is not None else tuple()
            rows_out.append(excel_row_dict_from_column_map(excel_row, row_t, col_map))
        wb.close()
        return rows_out

    if strict_reference_layout:
        wb.close()
        exp = "\n".join(f"  {i + 1}. {h}" for i, h in enumerate(HEADERS))
        raise ValueError(
            "Файл Excel не соответствует эталону «Макеты»: в первой строке должны быть "
            f"все {len(HEADERS)} заголовков (имена как ниже, порядок столбцов любой).\n{exp}\n\n"
            "Сохраните эталонный шаблон или приведите файл к эталону."
        )

    layout_v3 = excel_header_is_makety_v3(hdr_t)
    f1 = excel_cell_str(hdr_t, 5) if len(hdr_t) > 5 else ""
    f1l = f1.lower()
    layout_v2 = bool(not layout_v3 and f1 and ("нов" in f1l or "new price" in f1l))
    for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None for v in row):
            continue
        row_t = tuple(row) if row is not None else tuple()
        if layout_v3:
            rows_out.append(
                {
                    "excel_row": excel_row,
                    "name": excel_cell_str(row_t, 12),
                    "size": excel_cell_str(row_t, 3),
                    "kind": excel_cell_str(row_t, 4),
                    "file": excel_cell_str(row_t, 5),
                    "price": excel_cell_str(row_t, 8),
                    "price_new": excel_cell_str(row_t, 9),
                    "qty_per_sheet": excel_cell_str(row_t, 10),
                    "qty_per_year": excel_cell_str(row_t, 11),
                }
            )
        elif layout_v2:
            rows_out.append(
                {
                    "excel_row": excel_row,
                    "name": excel_cell_str(row_t, 0),
                    "size": excel_cell_str(row_t, 1),
                    "kind": excel_cell_str(row_t, 2),
                    "file": excel_cell_str(row_t, 3),
                    "price": excel_cell_str(row_t, 4),
                    "price_new": excel_cell_str(row_t, 5),
                    "qty_per_sheet": excel_cell_str(row_t, 6),
                    "qty_per_year": excel_cell_str(row_t, 7),
                }
            )
        else:
            rows_out.append(
                {
                    "excel_row": excel_row,
                    "name": excel_cell_str(row_t, 0),
                    "size": excel_cell_str(row_t, 1),
                    "kind": excel_cell_str(row_t, 2),
                    "file": excel_cell_str(row_t, 3),
                    "price": excel_cell_str(row_t, 4),
                    "price_new": "",
                    "qty_per_sheet": excel_cell_str(row_t, 5),
                    "qty_per_year": excel_cell_str(row_t, 6),
                }
            )
    wb.close()
    return rows_out


def patch_makety_header_row(ws: Any) -> None:
    """Первая строка — заголовки столбцов (13 столбцов)."""
    for col_idx, title in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=title)


def clear_makety_data_below_row(ws: Any, max_row_keep: int) -> None:
    """Удаляет значения ниже последней строки набора (старые «хвосты»)."""
    last = int(ws.max_row or max_row_keep)
    if last <= max_row_keep:
        return
    for r_idx in range(max_row_keep + 1, last + 1):
        for c in range(1, MAKETY_EXCEL_NCOLS + 1):
            ws.cell(row=r_idx, column=c).value = None


def makety_row_to_excel_cells(item: dict[str, Any]) -> dict[int, Any]:
    """13 колонок листа «Макеты» (см. HEADERS)."""
    return {
        1: item.get("_cg_knife_name") or None,
        2: item.get("_cg_category") or None,
        3: item.get("_cg_lacquers") or None,
        4: item.get("size") or None,
        5: item.get("kind") or None,
        6: item.get("file") or None,
        7: item.get("_knife_size_mm") or None,
        8: item.get("_cg_cutit_no") or None,
        9: item.get("price") or None,
        10: item.get("price_new") or None,
        11: item.get("qty_per_sheet") or None,
        12: item.get("qty_per_year") or None,
        13: item.get("name") or None,
    }


def workbook_save_atomic(wb: Any, excel_path: Path) -> None:
    """Сохранение через временный файл и os.replace."""
    target = excel_path.expanduser().resolve()
    target.parent.mkdir(parents=True, exist_ok=True)
    fd, raw = tempfile.mkstemp(suffix=".xlsx", dir=str(target.parent))
    os.close(fd)
    tmp = Path(raw)
    try:
        wb.save(tmp)
        os.replace(tmp, target)
    except Exception:
        if tmp.exists():
            try:
                tmp.unlink()
            except OSError:
                pass
        raise


def save_rows_to_excel(
    excel_path: Path,
    rows: list[dict[str, Any]],
    *,
    db_path: Path | None = None,
    enrich_from_db: EnrichRowsFromDb | None = None,
) -> None:
    _enrich_rows_if_db(db_path, rows, enrich_from_db)
    wb = load_workbook(excel_path)
    ws = wb.active
    patch_makety_header_row(ws)
    for item in rows:
        r = item["excel_row"]
        for c, v in makety_row_to_excel_cells(item).items():
            ws.cell(row=r, column=c, value=v)
    max_er = max((int(r["excel_row"]) for r in rows), default=1)
    clear_makety_data_below_row(ws, max_er)
    workbook_save_atomic(wb, excel_path)


def save_one_row_to_excel(
    excel_path: Path,
    item: dict[str, Any],
    *,
    db_path: Path | None = None,
    enrich_from_db: EnrichRowsFromDb | None = None,
) -> None:
    _enrich_rows_if_db(db_path, [item], enrich_from_db)
    wb = load_workbook(excel_path)
    ws = wb.active
    patch_makety_header_row(ws)
    r = int(item["excel_row"])
    for c, v in makety_row_to_excel_cells(item).items():
        ws.cell(row=r, column=c, value=v)
    workbook_save_atomic(wb, excel_path)
