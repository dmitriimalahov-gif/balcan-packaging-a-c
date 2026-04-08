#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Переносит в БД и Excel все виды из бывшей категории «Прочее» в «Этикетка».
Строки с видом __READ_ERROR__ не трогает.

  python3 migrate_prochee_to_etiketka.py
  python3 migrate_prochee_to_etiketka.py --dry-run
"""

from __future__ import annotations

import argparse
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

import packaging_db as pkg_db

ROOT = Path(__file__).resolve().parent
DEFAULT_EXCEL = ROOT / "Упаковка_макеты.xlsx"


def kind_bucket_raw(kind: str) -> str:
    raw = (kind or "").strip()
    k = raw.lower()
    if "без вторичной" in k or "fara secundar" in k or "fara cutie" in k:
        return "other"
    if raw == "Коробка" or "короб" in k:
        return "box"
    if "блистер" in k or "blister" in k:
        return "blister"
    if raw == "Пакет" or "пакет" in k:
        return "pack"
    if raw == "Этикетка" or "этикетк" in k:
        return "label"
    return "other"


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", type=Path, default=pkg_db.DEFAULT_DB_PATH)
    ap.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    changes_by_row: dict[int, str] = {}

    if args.excel.is_file():
        wb = load_workbook(args.excel)
        ws = wb.active
        for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None or all(v is None for v in row):
                continue
            kind = "" if len(row) < 3 or row[2] is None else str(row[2]).strip()
            if not kind or kind.startswith("__"):
                continue
            if kind_bucket_raw(kind) != "other":
                continue
            changes_by_row[excel_row] = kind
        wb.close()

    if args.db.is_file():
        conn = pkg_db.connect(args.db)
        try:
            pkg_db.init_db(conn)
            for r in pkg_db.load_all(conn):
                kind = (r.get("kind") or "").strip()
                er = int(r["excel_row"])
                if not kind or kind.startswith("__"):
                    continue
                if kind_bucket_raw(kind) != "other":
                    continue
                changes_by_row[er] = kind
        finally:
            conn.close()

    changes = [(er, old, "Этикетка") for er, old in sorted(changes_by_row.items())]

    print(f"Строк к обновлению (Excel/та же логика для БД): {len(changes)}")
    for er, old, new in changes[:30]:
        print(f"  row {er}: {old!r} → {new!r}")
    if len(changes) > 30:
        print(f"  … ещё {len(changes) - 30}")

    if args.dry_run:
        print("--dry-run: запись отключена.")
        return

    if changes and args.excel.is_file():
        wb = load_workbook(args.excel)
        ws = wb.active
        for excel_row, _old, new in changes:
            ws.cell(row=excel_row, column=3, value=new)
        wb.save(args.excel)
        print(f"Excel: обновлено {len(changes)} строк.")

    if args.db.is_file() and changes:
        conn = pkg_db.connect(args.db)
        try:
            pkg_db.init_db(conn)
            now = datetime.now(timezone.utc).isoformat()
            for excel_row, _old, new in changes:
                conn.execute(
                    "UPDATE packaging_items SET kind = ?, updated_at = ? WHERE excel_row = ?",
                    (new, now, excel_row),
                )
            conn.commit()
            print(f"SQLite: обновлено {len(changes)} строк.")
        finally:
            conn.close()
    elif changes and not args.db.is_file():
        print("БД не найдена — только Excel обновлён.")


if __name__ == "__main__":
    main()
