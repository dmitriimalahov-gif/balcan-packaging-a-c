#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Просмотр миниатюр PDF-макетов и правка «Размер (мм)», «Вид» и количеств.
В таблице «Макеты» данные каталога ножей CG (название, категория, лаки), нож и текущая цена CG
подставляются из SQLite; название для сопоставления с cutii хранится в последнем столбце Excel и не показывается в сетке.
PDF открывается в модальном окне (кнопка «PDF» у превью).
Сохранение в Excel и SQLite (packaging_data.db). Размер фиксируется формой (Enter / ↵).

Запуск из папки a-c:
  pip install -r requirements-viewer.txt
  streamlit run packaging_viewer.py
Кнопки «Загрузить обновлённый Excel», «Скачать Excel» и «Профиль Excel» (вся база) — в шапке под переключателем разделов.
Эталон листа «Макеты»: первая строка — все 13 заголовков из HEADERS (порядок столбцов может быть любым).
Опционально `makety_paths_ref.json` — подписи к эталонным путям Excel/SQLite в сайдбаре.
"""

from __future__ import annotations

import base64
from collections import Counter, defaultdict
from datetime import date, datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import hashlib
import html
import io
import math
import os
import re
import sqlite3
import tempfile
from pathlib import Path
from typing import Any

import fitz  # PyMuPDF
import streamlit as st
import streamlit.components.v1 as st_components

import packaging_db as pkg_db
import packaging_pdf_sheet_preview as packaging_pdf_preview
from modules.packaging_catalog.application.makety_display import (
    format_qty_year_caption,
    parse_qty_int_for_cg,
)
from modules.packaging_catalog.application import excel_io
from modules.packaging_catalog.application.excel_headers import excel_cell_str
from modules.packaging_catalog.application.makety_cg_enrichment import (
    CG_FINISH_LABELS_MAKETY,
    apply_makety_cg_derived_from_db,
)
from modules.packaging_catalog.application.cg_auto_mapping import auto_match_cg
from modules.packaging_catalog.application.cg_pret_import import parse_cg_pret_workbook
from modules.packaging_catalog.application.makety_catalog_ref import (
    load_makety_catalog_ref,
    load_makety_paths_ref,
    save_makety_catalog_ref,
    REF_CATALOG_TOTAL_ROWS,
    REF_CATALOG_KIND_STATS,
    MAKETY_CATALOG_REF_PATH,
    MAKETY_PATHS_REF_PATH,
)
from modules.packaging_catalog.application.makety_kind_merge import merge_kind_values_from_sqlite
from modules.packaging_catalog.domain.kind_bucket import (
    build_kind_options,
    kind_bucket,
    kind_stats,
    DEFAULT_KIND_OPTIONS,
)
from modules.packaging_catalog.domain.makety_excel_config import HEADERS, MAKETY_EXCEL_NCOLS
from modules.packaging_catalog.domain.makety_filters import (
    format_size_key_label,
    item_matches_bucket,
    item_matches_size_key,
    item_matches_text_query,
    size_key_str,
)
from modules.packaging_catalog.domain.makety_sort import sort_rows

_excel_cell_str = excel_cell_str
from pdf_outline_to_svg import open_pdf_document
from packaging_sizes import (
    canonicalize_size_mm,
    normalize_size,
    parse_box_dimensions_mm,
    row_size_key,
)

ROOT = Path(__file__).resolve().parent
DEFAULT_EXCEL = ROOT / "Упаковка_макеты.xlsx"

# Один раз за сессию: распространение SVG-ножей по размерам (планировщик и «Печать и заявки»)
_PKG_KNIFE_PROPAGATE_SESSION_KEY = "_pkg_knife_propagate_done"
# Сообщение после «Сохранить нож в БД» (видно после st.rerun() над блоком контура)
_PP_KNIFE_SAVE_FEEDBACK_KEY = "_pp_knife_saved_feedback"

# Относительные веса столбцов таблицы «Макеты» (аргумент st.columns)
MAKETY_COL_WIDTH_DEFAULTS: tuple[float, ...] = (
    0.92,
    1.05,
    0.85,
    1.0,
    1.35,
    1.55,
    1.65,
    0.95,
    0.72,
    0.78,
    0.42,
    0.42,
    0.4,
)
MAKETY_COL_LABELS: tuple[str, ...] = (
    "PDF",
    "Название ножа CG",
    "Категория CG",
    "Лаки CG",
    "Вид",
    "Превью",
    "Размер",
    "Размер ножа",
    "Нож CG",
    "Цена CG",
    "Нов. цена",
    "На листе",
    "За год",
)

MAKETY_COL_WIDTHS_USER_PATH = ROOT / "makety_col_widths_user.json"


def _load_user_makety_col_widths() -> list[float] | None:
    """Сохранённые пользователем доли ширины столбцов (если файл есть и валиден)."""
    import json

    p = MAKETY_COL_WIDTHS_USER_PATH
    if not p.is_file():
        return None
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
        w = data.get("widths")
        n = len(MAKETY_COL_WIDTH_DEFAULTS)
        if not isinstance(w, list) or len(w) != n:
            return None
        out = [float(x) for x in w]
        if any(not math.isfinite(x) or x <= 0 for x in out):
            return None
        return out
    except Exception:
        return None


def save_user_makety_col_widths(widths: list[float]) -> None:
    """Записать доли столбцов в makety_col_widths_user.json (атомарно)."""
    import json

    n = len(MAKETY_COL_WIDTH_DEFAULTS)
    if len(widths) != n:
        raise ValueError("wrong widths length")
    target = MAKETY_COL_WIDTHS_USER_PATH.expanduser().resolve()
    target.parent.mkdir(parents=True, exist_ok=True)
    payload = {"widths": [float(x) for x in widths]}
    fd, raw = tempfile.mkstemp(suffix=".json", dir=str(target.parent))
    os.close(fd)
    tmp = Path(raw)
    try:
        tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        os.replace(tmp, target)
    except Exception:
        if tmp.exists():
            try:
                tmp.unlink()
            except OSError:
                pass
        raise


def _init_makety_col_width_session() -> None:
    user_w = _load_user_makety_col_widths()
    for i, d in enumerate(MAKETY_COL_WIDTH_DEFAULTS):
        k = f"pkg_col_w_{i}"
        if k not in st.session_state:
            st.session_state[k] = float(user_w[i]) if user_w is not None else float(d)


def _makety_col_weights() -> list[float]:
    return [float(st.session_state[f"pkg_col_w_{i}"]) for i in range(len(MAKETY_COL_WIDTH_DEFAULTS))]


def row_snapshot_for_mirror(item: dict[str, Any]) -> tuple[str, str, str, str, str, str]:
    """Снимок полей строки для сравнения с зеркалом (БД + Excel при изменении)."""
    return (
        item.get("kind") or "",
        item.get("size") or "",
        item.get("price") or "",
        item.get("price_new") or "",
        item.get("qty_per_sheet") or "",
        item.get("qty_per_year") or "",
    )










def _widget_key_suffix(filename: str) -> str:
    safe = re.sub(r"[^a-zA-Z0-9]+", "_", filename)[:120]
    return safe.strip("_") or "row"


def _open_fitz_document(path_str: str) -> fitz.Document | None:
    """То же устойчивое открытие, что и для контура/SVG: байты с диска, пустой пароль, проверка 1-й страницы."""
    return open_pdf_document(path_str)


def _pixmap_first_page(
    page: fitz.Page,
    scale: float,
    max_raster_px: int,
) -> fitz.Pixmap | None:
    """Несколько стратегий rasterize: поворот страницы, RGB, аннотации, displaylist."""
    rect = page.rect
    long_side = max(rect.width, rect.height) or 1.0
    cap = max(32, int(max_raster_px)) / long_side
    s = min(max(0.04, float(scale)), cap)
    base = fitz.Matrix(s, s) * page.derotation_matrix

    strategies: list[dict[str, Any]] = [
        {"alpha": False, "colorspace": fitz.csRGB, "annots": False},
        {"alpha": False, "colorspace": fitz.csRGB, "annots": True},
        {"alpha": False, "annots": False},
        {"alpha": True, "colorspace": fitz.csRGB, "annots": False},
    ]
    try:
        cb = page.cropbox
        if not cb.is_empty and cb.width > 0.5 and cb.height > 0.5:
            strategies.append(
                {"alpha": False, "colorspace": fitz.csRGB, "annots": True, "clip": cb}
            )
    except Exception:
        pass

    for factor in (1.0, 0.55, 0.3):
        mat = fitz.Matrix(s * factor, s * factor) * page.derotation_matrix
        for kw in strategies:
            clip = kw.get("clip")
            try:
                opts = {k: v for k, v in kw.items() if k != "clip"}
                pix = page.get_pixmap(
                    matrix=mat,
                    clip=clip if clip is not None and not clip.is_empty else None,
                    **opts,
                )
                if pix.width > 0 and pix.height > 0:
                    return pix
            except Exception:
                continue
        try:
            dl = page.get_displaylist()
            pix = dl.get_pixmap(matrix=mat, alpha=False, colorspace=fitz.csRGB)
            if pix.width > 0 and pix.height > 0:
                return pix
        except Exception:
            pass

    try:
        pix = page.get_pixmap(matrix=base, alpha=False, colorspace=fitz.csRGB)
        if pix.width > 0 and pix.height > 0:
            return pix
    except Exception:
        pass
    return None


@st.cache_data(show_spinner=False)
def render_pdf_thumb(
    path_str: str,
    mtime: float,
    scale: float,
    sharpness: float = 1.0,
    max_raster_px: int = 720,
) -> bytes | None:
    """Рендер первой страницы в PNG (устойчиво к повороту, CMYK, части «битых» макетов)."""
    doc = _open_fitz_document(path_str)
    if doc is None:
        return None
    try:
        if doc.page_count < 1:
            return None
        page = doc.load_page(0)
        s0 = max(0.05, float(scale)) * max(1.0, float(sharpness))
        pix = _pixmap_first_page(page, s0, max_raster_px)
        if pix is None:
            return None
        return pix.tobytes("png")
    except Exception:
        return None
    finally:
        try:
            doc.close()
        except Exception:
            pass


@st.cache_data(show_spinner=False)
def render_pdf_sheet_slot_png(
    path_str: str,
    mtime: float,
    slot_w_mm: float,
    slot_h_mm: float,
    dpi: float,
    transparent_bg: bool,
) -> bytes | None:
    """Первая страница PDF, вписанная в слот листа (мм), для превью в SVG."""
    import packaging_pdf_sheet_preview as ppsp

    return ppsp.render_first_page_fit_to_mm(
        path_str,
        float(slot_w_mm),
        float(slot_h_mm),
        dpi=float(dpi),
        transparent_bg=bool(transparent_bg),
    )


@st.cache_data(show_spinner=False)
def render_pdf_sheet_slot_knife_png(
    path_str: str,
    mtime: float,
    slot_w_mm: float,
    slot_h_mm: float,
    dpi: float,
    transparent_bg: bool,
    layout_kind: str = "",
) -> bytes | None:
    """Область контура ножа (bbox обводок) → PNG, вписанная в слот; без контура — None."""
    import packaging_pdf_sheet_preview as ppsp

    return ppsp.render_knife_bbox_fit_to_mm(
        path_str,
        float(slot_w_mm),
        float(slot_h_mm),
        dpi=float(dpi),
        transparent_bg=bool(transparent_bg),
        layout_kind=(layout_kind or None),
    )


@st.cache_data(show_spinner=False)
def render_pdf_modal_preview_png(
    path_str: str,
    mtime: float,
    max_side_px: int = 1100,
) -> bytes | None:
    """Крупное растровое превью первой страницы для диалога (если встроенный PDF криво отображается)."""
    doc = _open_fitz_document(path_str)
    if doc is None:
        return None
    try:
        if doc.page_count < 1:
            return None
        page = doc.load_page(0)
        scale = max(0.15, min(2.0, max_side_px / max(page.rect.width, page.rect.height, 1.0)))
        pix = _pixmap_first_page(page, scale, max_side_px)
        if pix is None:
            return None
        return pix.tobytes("png")
    except Exception:
        return None
    finally:
        try:
            doc.close()
        except Exception:
            pass


@st.cache_data(show_spinner=False)
def read_small_pdf_for_modal(path_str: str, mtime: float, size_limit: int) -> bytes | None:
    """Читает PDF целиком только если не больше size_limit; иначе None (без чтения тела файла)."""
    p = Path(path_str)
    try:
        st = p.stat()
    except OSError:
        return None
    if st.st_size > size_limit:
        return None
    try:
        return p.read_bytes()
    except OSError:
        return None


def prefetch_thumbs_parallel(
    jobs: list[tuple[str, float]],
    scale: float,
    sharpness: float,
    max_raster_px: int,
    max_workers: int = 12,
) -> None:
    """Параллельно заполняет кэш render_pdf_thumb для уникальных (путь, mtime)."""
    if not jobs:
        return
    n = min(max(1, max_workers), len(jobs))
    with ThreadPoolExecutor(max_workers=n) as ex:
        futs = [
            ex.submit(
                render_pdf_thumb,
                ps,
                mt,
                scale,
                sharpness=sharpness,
                max_raster_px=max_raster_px,
            )
            for ps, mt in jobs
        ]
        for fu in as_completed(futs):
            fu.result()


def load_rows_from_excel(
    excel_path: Path,
    *,
    strict_reference_layout: bool = False,
) -> list[dict[str, Any]]:
    """
    Читает активный лист. Если в первой строке найдены все 13 заголовков из HEADERS
    (порядок столбцов может быть любым), строки разбираются по этим именам — как эталон «Макеты».

    strict_reference_layout=True: при отсутствии полного набора заголовков — ValueError
    (нужно привести файл к эталону или выключить строгий режим в сайдбаре).
    """
    return excel_io.load_rows_from_excel(
        excel_path, strict_reference_layout=strict_reference_layout
    )


def normalize_excel_file_to_makety_reference(
    excel_path: Path,
    db_path: Path | None,
    *,
    overwrite_nonempty_excel: bool = False,
) -> list[dict[str, Any]]:
    """
    Читает Excel (в т.ч. старый формат), подмешивает «Вид» из БД при необходимости,
    перезаписывает файл в каноническом виде: строка 1 = HEADERS, 13 столбцов, данные как в эталоне.
    """
    rows = load_rows_from_excel(excel_path, strict_reference_layout=False)
    if db_path is not None:
        merge_kind_from_db(
            rows,
            db_path,
            excel_path,
            overwrite_nonempty_excel=overwrite_nonempty_excel,
        )
    save_rows_to_excel(excel_path, rows, db_path=db_path)
    return rows


def save_rows_to_excel(
    excel_path: Path,
    rows: list[dict[str, Any]],
    db_path: Path | None = None,
) -> None:
    excel_io.save_rows_to_excel(excel_path, rows, db_path=db_path)


def save_one_row_to_excel(
    excel_path: Path,
    item: dict[str, Any],
    db_path: Path | None = None,
) -> None:
    """Обновляет одну строку листа; файл перезаписывается атомарно."""
    excel_io.save_one_row_to_excel(excel_path, item, db_path=db_path)


def save_rows_to_db(db_path: Path, rows: list[dict[str, Any]]) -> None:
    db_path = db_path.expanduser().resolve()
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = pkg_db.connect(db_path)
    try:
        pkg_db.init_db(conn)
        pkg_db.upsert_all(conn, rows)
    finally:
        conn.close()




def sync_widgets_to_rows(rows: list[dict[str, Any]]) -> None:
    for item in rows:
        r = item["excel_row"]
        suffix = _widget_key_suffix(item["file"] or str(r))
        use_custom_key = f"use_custom_{suffix}_{r}"
        sel_key = f"kind_sel_{suffix}_{r}"
        cust_key = f"kind_cust_{suffix}_{r}"
        price_new_key = f"price_new_{suffix}_{r}"
        qps_key = f"qty_sheet_{suffix}_{r}"
        qpy_key = f"qty_year_{suffix}_{r}"
        if st.session_state.get(use_custom_key, False):
            if cust_key in st.session_state:
                item["kind"] = str(st.session_state[cust_key])
        else:
            if sel_key in st.session_state:
                item["kind"] = str(st.session_state[sel_key])
        if price_new_key in st.session_state:
            item["price_new"] = str(st.session_state[price_new_key])
        if qps_key in st.session_state:
            item["qty_per_sheet"] = str(st.session_state[qps_key])
        if qpy_key in st.session_state:
            item["qty_per_year"] = str(st.session_state[qpy_key])
        # размер задаётся только через форму (Enter / ↵), не из общего session_state
        # «Цена» — из прайса CG (apply_makety_cg_derived_from_db), не из виджета


def _kind_use_custom_off_sync_selectbox(suffix: str, rk: int) -> None:
    """
    Streamlit вызывает on_change до основного скрипта. При снятии «Свой вид» в session_state
    остаётся старое kind_sel_*; sync_widgets_to_rows иначе затрёт item['kind'] этим значением
    вместо текста из kind_cust_* (например Коробка → Blister откатывается).
    """
    use_key = f"use_custom_{suffix}_{rk}"
    if st.session_state.get(use_key, False):
        return
    sel_key = f"kind_sel_{suffix}_{rk}"
    cust_key = f"kind_cust_{suffix}_{rk}"
    v = str(st.session_state.get(cust_key, "")).strip()
    if v:
        st.session_state[sel_key] = v


def reconcile_row_to_storage(
    db_path: Path,
    excel_path: Path,
    rows: list[dict[str, Any]],
) -> None:
    """При изменении вида, размера, цены или количеств — строка в SQLite и Excel."""
    mirror: dict[int, tuple[str, str, str, str, str, str]] = st.session_state.setdefault(
        "_db_row_mirror",
        {},
    )
    row_ids = {int(r["excel_row"]) for r in rows}
    if set(mirror.keys()) != row_ids:
        mirror.clear()
        for r in rows:
            rk = int(r["excel_row"])
            mirror[rk] = row_snapshot_for_mirror(r)
        return
    for item in rows:
        rk = int(item["excel_row"])
        cur = row_snapshot_for_mirror(item)
        if mirror.get(rk) == cur:
            continue
        try:
            # Сначала SQLite: при ошибке Excel старый «Вид» из БД не откатит правку при следующем входе.
            save_rows_to_db(db_path, [item])
            save_one_row_to_excel(excel_path, item, db_path)
        except Exception as e:
            st.toast(f"Строка {rk}: {e}", icon="⚠️")
            continue
        mirror[rk] = cur


@st.dialog("Просмотр PDF", width="large")
def open_real_pdf_dialog(path_str: str, max_iframe_bytes: int, download_key: str) -> None:
    """Просмотр PDF: Streamlit st.pdf (если есть), иначе iframe; растровое превью PyMuPDF при сбоях."""
    p = Path(path_str)
    try:
        mtime = p.stat().st_mtime
    except OSError as e:
        st.error(f"Не удалось прочитать файл: {e}")
        return
    data = read_small_pdf_for_modal(path_str, mtime, max_iframe_bytes)
    if data is None:
        try:
            data = p.read_bytes()
        except OSError as e:
            st.error(f"Не удалось прочитать файл: {e}")
            return
    title = p.name
    st.caption(title if len(title) < 120 else title[:117] + "…")

    preview_png = render_pdf_modal_preview_png(str(p), mtime, 1100)

    if len(data) <= max_iframe_bytes:
        embedded = False
        if hasattr(st, "pdf"):
            try:
                st.pdf(io.BytesIO(data), height=680)
                embedded = True
            except Exception:
                embedded = False
        if not embedded:
            b64 = base64.b64encode(data).decode("ascii")
            st.markdown(
                f'<iframe src="data:application/pdf;base64,{b64}" '
                'style="width:100%;min-height:62vh;border:none;border-radius:4px;"></iframe>',
                unsafe_allow_html=True,
            )
        if preview_png:
            with st.expander(
                "Превью 1-й страницы (растр) — если PDF выше пустой или отображается с ошибкой",
                expanded=False,
            ):
                st.image(preview_png, use_container_width=True)
    else:
        st.info("Файл слишком большой для встроенного просмотра — скачайте и откройте в системе.")
        if preview_png:
            st.caption("Первая страница (растр, PyMuPDF):")
            st.image(preview_png, use_container_width=True)
        st.download_button(
            "Скачать PDF",
            data=data,
            file_name=p.name,
            mime="application/pdf",
            key=download_key,
        )


def _escape_attr(s: str) -> str:
    return (
        s.replace("&", "&amp;")
        .replace('"', "&quot;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def _makety_short_txt(val: str, lim: int = 100) -> str:
    s = (val or "").strip()
    if not s:
        return "—"
    return s if len(s) <= lim else s[: lim - 1] + "…"


def _cutii_row_select_options(
    p: dict[str, Any],
    by_er: dict[int, dict[str, Any]],
    *,
    max_choices: int = 220,
) -> tuple[list[int | None], dict[int | None, str]]:
    """Варианты выпадающего списка: топ совпадений ★, затем остальные коробки из БД по excel_row."""
    tops = p.get("tops") or []
    vals: list[int | None] = [None]
    labels: dict[int | None, str] = {None: "— не выбрано —"}
    seen: set[int] = set()

    for t in tops:
        er = int(t["excel_row"])
        if er in seen:
            continue
        seen.add(er)
        br = by_er.get(er)
        nm = ((br or {}).get("name") or t.get("name") or "").strip()
        score = t.get("score", "")
        vals.append(er)
        labels[er] = f"★ {er} · {nm[:68]} — {score}%"

    for er in sorted(by_er.keys()):
        if er in seen:
            continue
        seen.add(er)
        nm = (by_er[er].get("name") or "").strip()[:68]
        vals.append(er)
        labels[er] = f"{er} · {nm}"
        if len(vals) >= max_choices:
            break

    return vals, labels


def _collect_cutii_picks_from_widgets(pending: list[dict[str, Any]]) -> dict[int, int]:
    out: dict[int, int] = {}
    for p in pending:
        sr = int(p["cutii_sheet_row"])
        key = f"cutii_map_{sr}"
        if key not in st.session_state:
            continue
        v = st.session_state[key]
        if v is None:
            continue
        try:
            er = int(v)
        except (TypeError, ValueError):
            continue
        if er > 0:
            out[sr] = er
    return out


def render_cutii_tab(excel_path: Path, db_path: Path, packaging_rows: list[dict[str, Any]]) -> None:
    """Сопоставление cutii с коробками: анализ, выбор коробки из БД, сохранение в SQLite и импорт."""
    import import_cutii_forecast as ic

    st.title("Cutii → коробки")
    st.caption(
        "Файл cutii (колонка B, buc.) сопоставляется с коробками (вид «Коробка») из SQLite. "
        "Для спорных строк выберите позицию из списка и нажмите «Сохранить сопоставление в БД»."
    )

    def_cutii = ic.DEFAULT_CUTII if ic.DEFAULT_CUTII.is_file() else (ROOT.parent / "Balcan 2025 cutii.xlsx")
    cutii_str = st.text_input("Файл cutii (.xlsx)", value=str(def_cutii), key="cutii_xlsx_path")
    ov_str = st.text_input("Overrides CSV", value=str(ic.DEFAULT_OVERRIDES), key="cutii_ov_path")
    conf_str = st.text_input(
        "Подтверждения CSV (cutii_confirmations)",
        value=str(ic.DEFAULT_CONFIRMATIONS),
        key="cutii_conf_path",
    )

    c1, c2 = st.columns(2)
    with c1:
        min_sc = st.slider("Порог совпадения", 35, 95, 50, 1, key="cutii_min_score")
        amb_gap = st.slider("Разрыв кандидатов", 1, 15, 5, 1, key="cutii_amb_gap")
    with c2:
        no_conf = st.checkbox("Не читать файл подтверждений", value=False, key="cutii_no_conf")
        no_fb = st.checkbox("Без сопоставления по PDF", value=False, key="cutii_no_pdf")

    if st.button("Запустить анализ", type="primary", key="cutii_run_analysis"):
        cp = Path(cutii_str).expanduser().resolve()
        if not cp.is_file():
            st.error(f"Файл не найден: {cp}")
        else:
            with st.spinner("Сопоставление… (~30–60 с)"):
                res = ic.run_cutii_analysis(
                    cp,
                    packaging_rows,
                    Path(ov_str).expanduser().resolve(),
                    Path(conf_str).expanduser().resolve(),
                    no_confirmations=no_conf,
                    min_score=min_sc,
                    ambiguous_gap=amb_gap,
                    fallback_pdf=not no_fb,
                    interactive=False,
                    confirmations_db_path=db_path,
                )
            st.session_state["cutii_analysis"] = res
            st.rerun()

    res = st.session_state.get("cutii_analysis")
    if not res:
        st.info("Нажмите «Запустить анализ» после выбора файла cutii.")
        return

    rep = res["report_rows"]
    n_ok = sum(1 for r in rep if r["status"] == "ok")
    n_amb = sum(1 for r in rep if r["status"] == "ambiguous")
    n_miss = sum(1 for r in rep if r["status"] == "no_match")
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Авто-сопоставлено", n_ok)
    m2.metric("Неоднозначно", n_amb)
    m3.metric("Нет пары", n_miss)
    m4.metric("Нужна ручная проверка", len(res["pending"]))

    pending = res["pending"]

    db_conf: dict[int, int] = {}
    box_from_db: list[dict[str, Any]] = []
    if db_path.is_file():
        conn0 = pkg_db.connect(db_path)
        try:
            pkg_db.init_db(conn0)
            box_from_db = [
                r for r in pkg_db.load_all(conn0) if ic.is_packaging_box(r.get("kind") or "")
            ]
            db_conf, _db_names = pkg_db.load_cutii_confirmations(conn0)
        finally:
            conn0.close()
    else:
        box_from_db = list(res["box_rows"])

    by_er: dict[int, dict[str, Any]] = {int(r["excel_row"]): r for r in box_from_db}

    if pending:
        st.subheader("Ручное подтверждение")
        st.caption(
            "Список строится из таблицы packaging_items (только «Коробка»): сначала кандидаты анализа (★), затем остальные строки."
        )
        for p in pending:
            sr = int(p["cutii_sheet_row"])
            vals, labels = _cutii_row_select_options(p, by_er)
            pref = db_conf.get(sr)
            if pref is not None and pref not in vals:
                br = by_er.get(pref)
                nm = ((br or {}).get("name") or "").strip()[:68]
                vals = [None, pref] + [v for v in vals if v is not None and v != pref]
                labels = dict(labels)
                labels[pref] = f"{pref} · {nm} (сохранено в БД)"

            key = f"cutii_map_{sr}"
            sb_kw: dict[str, Any] = {
                "options": vals,
                "format_func": lambda v, lb=labels: lb.get(v, str(v)),
                "key": key,
                "label_visibility": "collapsed",
            }
            if key not in st.session_state and pref is not None and pref in vals:
                sb_kw["index"] = vals.index(pref)

            c_a, c_b = st.columns([2.2, 5.0], gap="small")
            with c_a:
                st.markdown(
                    f'<p class="pkg-fn" title="{html.escape(p["cutii_name"])}">'
                    f"<strong>Лист {sr}</strong><br/>{html.escape(p['cutii_name'][:140])}"
                    f"{'…' if len(p['cutii_name']) > 140 else ''}</p>",
                    unsafe_allow_html=True,
                )
                st.caption(f"Статус: {p['status']} · год: {p['qty_year']}")
            with c_b:
                st.selectbox(
                    f"Коробка для строки cutii {sr}",
                    **sb_kw,
                )

    picks = _collect_cutii_picks_from_widgets(pending) if pending else {}

    bc1, bc2, bc3, bc4 = st.columns(4)
    with bc1:
        if st.button("Сохранить сопоставление в БД", type="primary", key="cutii_save_db"):
            picks_db = _collect_cutii_picks_from_widgets(pending) if pending else {}
            if not picks_db:
                st.warning("Ни для одной строки не выбрана коробка (не «— не выбрано —»).")
            else:
                entries = [
                    {
                        "cutii_sheet_row": sr,
                        "confirmed_excel_row": er,
                        "cutii_name": next(
                            (x["cutii_name"] for x in pending if int(x["cutii_sheet_row"]) == sr),
                            "",
                        ),
                    }
                    for sr, er in picks_db.items()
                ]
                try:
                    db_path.parent.mkdir(parents=True, exist_ok=True)
                    conn_s = pkg_db.connect(db_path)
                    try:
                        pkg_db.init_db(conn_s)
                        pkg_db.upsert_cutii_confirmations(conn_s, entries)
                    finally:
                        conn_s.close()
                    st.success(f"Сохранено в БД: {len(entries)} сопоставлений ({db_path.name}).")
                except OSError as e:
                    st.error(str(e))
    with bc2:
        if st.button("Сохранить подтверждения в CSV", key="cutii_save_conf"):
            picks_csv = _collect_cutii_picks_from_widgets(pending) if pending else {}
            if not picks_csv:
                st.warning("Нет выбранных коробок для экспорта в CSV.")
            else:
                path_c = Path(conf_str).expanduser().resolve()
                cur, cnames = ic.load_confirmations(path_c)
                name_by_sr = {int(x["cutii_sheet_row"]): x["cutii_name"] for x in pending}
                for sr, er in picks_csv.items():
                    cur[sr] = er
                    if sr in name_by_sr:
                        cnames[sr] = name_by_sr[sr]
                out_list = [
                    {
                        "cutii_sheet_row": k,
                        "confirmed_excel_row": cur[k],
                        "cutii_name": cnames.get(k, ""),
                    }
                    for k in sorted(cur.keys())
                ]
                try:
                    ic.save_cutii_confirmations_csv(path_c, out_list)
                    st.success(f"Записано: {path_c} ({len(out_list)} строк)")
                except OSError as e:
                    st.error(str(e))
    with bc3:
        if st.button("Импорт объёмов (авто + подтверждённые)", type="secondary", key="cutii_apply_import"):
            picks_imp = _collect_cutii_picks_from_widgets(pending) if pending else {}
            merged = ic.build_to_apply_with_ui_picks(
                res["to_apply"],
                res["pending"],
                picks_imp,
                res["box_rows"],
            )
            if not merged:
                st.warning("Нечего импортировать.")
            else:
                try:
                    ic.apply_cutii_import(merged, db_path, excel_path, res["source_tag"])
                except Exception as e:
                    st.error(str(e))
                else:
                    clear_packaging_row_widget_keys()
                    st.session_state.pop("packaging_rows", None)
                    st.session_state.pop("_db_row_mirror", None)
                    st.session_state.pop("cutii_analysis", None)
                    st.success(f"Обновлено позиций: {len(merged)}. Перезагружены данные из Excel.")
                    st.rerun()
    with bc4:
        if st.button("Сбросить результат анализа", key="cutii_clear_analysis"):
            st.session_state.pop("cutii_analysis", None)
            st.rerun()

    st.caption(
        f"Источник cutii: {res.get('source_tag', '')} · Выбрано сопоставлений: {len(picks)} · "
        "Подтверждения в SQLite подхватываются при следующем «Запустить анализ». После импорта объёмов откройте «Макеты»."
    )


def render_print_orders_tab(
    packaging_rows: list[dict[str, Any]],
    pdf_root: Path,
    db_path: Path,
) -> None:
    """Печать: раскладка на лист, превью PDF по слотам, заявки, приоритеты на 1–3 месяца."""
    import packaging_pdf_sheet_preview as ppsp
    import packaging_sheet_export as pse
    import pandas as pd
    import pdf_outline_to_svg as pdf_outline

    import packaging_print_planning as pp

    def _pp_pack_item_dims_mm(w_mm: float, h_mm: float) -> tuple[float, float]:
        """
        Габариты прямоугольника для укладки на лист: при повороте ножа в ячейке 90° или 270°
        меняем ширину и высоту местами (иначе сетка не соответствует занимаемому месту на листе).
        Зеркала габариты не меняют.
        """
        rot = int(st.session_state.get("pp_svg_rot", 0)) % 360
        fw, fh = float(w_mm), float(h_mm)
        if rot in (90, 270):
            return fh, fw
        return fw, fh

    st.title("Печать и заявки")
    st.caption(
        "**Размер коробки (габариты)** для раскладки и геометрии листа — **только из SQLite** (поле `size` в `packaging_items`), без PDF и без подстановки из Excel. "
        "Превью макетов в слотах по-прежнему из файлов PDF. "
        f"Папка PDF: **{pdf_root}**."
    )

    rows_by_er_session = {int(r["excel_row"]): r for r in packaging_rows}
    db_all_by_er: dict[int, dict[str, Any]] = {}
    box_rows_from_db: list[dict[str, Any]] = []
    db_read_ok = False
    db_has_any = False

    if db_path.is_file():
        try:
            conn = pkg_db.connect(db_path)
            try:
                pkg_db.init_db(conn)
                db_has_any = pkg_db.row_count(conn) > 0
                if db_has_any:
                    db_loaded = pkg_db.load_all(conn)
                    for r in db_loaded:
                        db_all_by_er[int(r["excel_row"])] = r
                    box_rows_from_db = pp.sheet_layout_candidate_rows(db_loaded)
                db_read_ok = True
            finally:
                conn.close()
        except Exception as e:
            st.warning(f"Не удалось прочитать базу для раскладки: {e}")

    box_rows: list[dict[str, Any]] = []
    if not db_path.is_file():
        st.warning("Файл SQLite не найден по указанному пути — раскладка по габаритам недоступна (данные только из БД).")
    elif not db_read_ok:
        pass
    elif not db_has_any:
        st.warning("База пуста — загрузите строки в SQLite (сохранение из «Макеты» или «Загрузить из SQLite»).")
    elif not box_rows_from_db:
        st.warning(
            "В базе нет позиций с видом коробка / блистер / пакет / этикетка — добавьте и сохраните в БД."
        )
    else:
        box_rows = box_rows_from_db
        st.caption(
            "Список габаритов и расчёт ячеек — по полю **size** в БД. Имена/файлы для превью подмешиваются из сессии (Excel), если строка там есть."
        )

    rows_by_er = dict(db_all_by_er)
    rows_by_er.update(rows_by_er_session)

    st.subheader("Параметры печатного листа")
    c1, c2, c3, c4, c5 = st.columns(5)
    with c1:
        sheet_w = st.number_input("Ширина листа, мм", min_value=50.0, max_value=2000.0, value=700.0, step=1.0, key="pp_sheet_w")
    with c2:
        sheet_h = st.number_input("Высота листа, мм", min_value=50.0, max_value=2000.0, value=1000.0, step=1.0, key="pp_sheet_h")
    with c3:
        margin_mm = st.number_input("Поля, мм", min_value=0.0, max_value=80.0, value=5.0, step=0.5, key="pp_margin")
    with c4:
        gap_mm = st.number_input(
            "Зазор по X, мм",
            min_value=-100.0,
            max_value=40.0,
            value=2.0,
            step=0.5,
            key="pp_gap",
            help="Расстояние между оттисками в ряду (по горизонтали). Отрицательное — сближение / нахлёст.",
        )
    with c5:
        gap_y_mm = st.number_input(
            "Зазор по Y, мм",
            min_value=-100.0,
            max_value=40.0,
            value=2.0,
            step=0.5,
            key="pp_gap_y",
            help="Расстояние между рядами оттисков (по вертикали). Отрицательное — сближение / нахлёст.",
        )

    sheet_params = pp.SheetParams(
        width_mm=float(sheet_w),
        height_mm=float(sheet_h),
        margin_mm=float(margin_mm),
        gap_mm=float(gap_mm),
        gap_y_mm=float(gap_y_mm),
    )

    st.subheader("Раскладка по размеру коробки")
    st.caption(
        "Без галочек — **коробки, блистеры и пакеты** (этикетки скрыты). Отметьте виды, "
        "чтобы в списке размеров остались только они; **Этикетка** — только явные этикетки в поле «Вид»."
    )
    _pp_lay_bx, _pp_lay_bl, _pp_lay_pk, _pp_lay_lb, _ = st.columns(
        [1.05, 1.05, 1.05, 1.05, 5.2]
    )
    with _pp_lay_bx:
        _pp_filter_box = st.checkbox("Коробка", value=False, key="pp_layout_filter_box")
    with _pp_lay_bl:
        _pp_filter_blister = st.checkbox("Блистер", value=False, key="pp_layout_filter_blister")
    with _pp_lay_pk:
        _pp_filter_pack = st.checkbox("Пакет", value=False, key="pp_layout_filter_pack")
    with _pp_lay_lb:
        _pp_filter_label = st.checkbox("Этикетка", value=False, key="pp_layout_filter_label")

    def _pp_layout_bucket_for_row(r: dict[str, Any]) -> str:
        er = int(r["excel_row"])
        merged = rows_by_er.get(er) or r
        return kind_bucket(merged)

    _pp_any_kind_filter = (
        _pp_filter_box
        or _pp_filter_blister
        or _pp_filter_pack
        or _pp_filter_label
    )
    if _pp_any_kind_filter:
        box_rows_for_layout = [
            r
            for r in box_rows
            if (_pp_filter_box and _pp_layout_bucket_for_row(r) == "box")
            or (_pp_filter_blister and _pp_layout_bucket_for_row(r) == "blister")
            or (_pp_filter_pack and _pp_layout_bucket_for_row(r) == "pack")
            or (_pp_filter_label and _pp_layout_bucket_for_row(r) == "label")
        ]
    else:
        box_rows_for_layout = [
            r for r in box_rows if _pp_layout_bucket_for_row(r) != "label"
        ]

    size_groups = pp.collect_box_size_groups(box_rows_for_layout)

    if (
        box_rows_for_layout
        and size_groups
        and db_path.is_file()
        and not st.session_state.get(_PKG_KNIFE_PROPAGATE_SESSION_KEY)
    ):
        try:
            _conn_pp_prop = pkg_db.connect(db_path)
            try:
                pkg_db.init_db(_conn_pp_prop)
                _knife_meta_pp = pkg_db.load_knives_meta(_conn_pp_prop)
                _n_pp_prop = pkg_db.propagate_knives_in_size_groups(
                    _conn_pp_prop, size_groups, _knife_meta_pp
                )
                if _n_pp_prop > 0:
                    st.toast(
                        f"SVG-ножи распространены на {_n_pp_prop} позиций того же размера (кэш в БД)"
                    )
            finally:
                _conn_pp_prop.close()
        except Exception:
            pass
        st.session_state[_PKG_KNIFE_PROPAGATE_SESSION_KEY] = True

    if not size_groups:
        if box_rows:
            if _pp_any_kind_filter and not box_rows_for_layout:
                st.warning(
                    "С выбранным фильтром вида нет строк в БД. Снимите галочки или проверьте поле «Вид»."
                )
            elif _pp_any_kind_filter and box_rows_for_layout:
                st.warning(
                    "У отфильтрованных позиций нет валидного «Размер (мм)» для группировки — раскладка недоступна."
                )
            elif not box_rows_for_layout:
                st.info(
                    "В выборке только **этикетки** — отметьте галочку **«Этикетка»**, чтобы строить раскладку по размеру."
                )
            else:
                st.warning("В БД среди выбранных позиций нет заполненного «Размер (мм)» — раскладка недоступна.")
    else:
        sk_options = [g["size_key"] for g in size_groups]
        sk_labels = {g["size_key"]: f"{pp.size_key_display(g['size_key'])} — {len(g['rows'])} наимен." for g in size_groups}
        chosen_sk = st.selectbox(
            "Размер коробки (габариты)",
            options=sk_options,
            format_func=lambda sk: sk_labels.get(sk, sk),
            key="pp_size_group_select",
        )
        group = next(g for g in size_groups if g["size_key"] == chosen_sk)
        same_rows = group["rows"]

        er_same = [int(r["excel_row"]) for r in same_rows]
        sk_safe = chosen_sk.replace("|", "_").replace("/", "_")[:80]

        def _pp_row_kind_name(er: int) -> tuple[str, str]:
            """Вид (kind) и наименование из объединённой строки (как у слотов)."""
            rr = rows_by_er.get(int(er))
            if not rr:
                return ("", "")
            k = (rr.get("kind") or "").strip()
            nm = (rr.get("name") or "").strip()
            return (k, nm)

        def _pp_kind_name_sheet_caption(er: int, *, cap: int = 48) -> str:
            """Краткая подпись для схемы листа: вид — наименование."""
            k, nm = _pp_row_kind_name(er)
            if k and nm:
                s = f"{k} — {nm}"
            elif nm:
                s = nm
            elif k:
                s = k
            else:
                s = ""
            return (s[:cap] + "…") if len(s) > cap else s

        def _outline_row_label(er: int) -> str:
            """Слоты/контур: excel_row + вид + имя."""
            k, nm = _pp_row_kind_name(er)
            if k and nm:
                return f"{er} — {k}: {nm[:50]}"
            if nm:
                return f"{er} — {nm[:55]}"
            if k:
                return f"{er} — {k[:55]}"
            return f"{er}"

        _kind_opts = sorted(
            {_pp_row_kind_name(int(e))[0] for e in er_same if _pp_row_kind_name(int(e))[0]},
            key=lambda x: (x.lower(), x),
        )
        if _kind_opts:
            _kind_pick = st.multiselect(
                "Ограничить коробки для слотов листа по виду (пусто — все строки группы размера)",
                options=_kind_opts,
                default=[],
                key=f"pp_kind_filter_{sk_safe}",
                help="Списки **слотов назначения** и позиция для контура — только строки с выбранным «Видом».",
            )
        else:
            _kind_pick = []
        if _kind_pick:
            er_filtered = [e for e in er_same if _pp_row_kind_name(int(e))[0] in set(_kind_pick)]
            if not er_filtered:
                st.warning("Ни одна строка не подходит под выбранные виды — используются все позиции группы.")
                er_for_slots = list(er_same)
            else:
                er_for_slots = er_filtered
        else:
            er_for_slots = list(er_same)

        er_sorted = sorted(
            er_for_slots,
            key=lambda e: (
                _pp_row_kind_name(int(e))[0].lower(),
                _pp_row_kind_name(int(e))[1].lower(),
                int(e),
            ),
        )

        _knives_print_by_er: dict[int, dict[str, Any]] = {}
        if db_path.is_file():
            try:
                _cn_kp2 = pkg_db.connect(db_path)
                try:
                    pkg_db.init_db(_cn_kp2)
                    _knives_print_by_er = pkg_db.load_knives_for_rows(_cn_kp2, er_same)
                finally:
                    _cn_kp2.close()
            except Exception:
                pass

        def _pp_svg_preview_transform_css() -> str:
            """CSS transform для превью SVG (img): поворот и зеркало, без изменения файла."""
            rot = int(st.session_state.get("pp_svg_rot", 0))
            fh = bool(st.session_state.get("pp_svg_flip_h", False))
            fv = bool(st.session_state.get("pp_svg_flip_v", False))
            parts: list[str] = []
            if rot:
                parts.append(f"rotate({rot}deg)")
            if fh:
                parts.append("scaleX(-1)")
            if fv:
                parts.append("scaleY(-1)")
            if not parts:
                return ""
            return f"transform:{' '.join(parts)}; transform-origin:center center;"

        st.caption(
            "Поворот **90° или 270°** пересчитывает **сетку слотов** на листе (ширина и высота ячейки меняются местами под повёрнутый нож). "
            "Зеркала **↔ / ↕** меняют только вид макета внутри уже рассчитанной ячейки (число слотов то же). "
            "Превью одного контура в expander ниже — по-прежнему целиком."
        )
        _pr1, _pr2, _pr3 = st.columns(3)
        with _pr1:
            st.selectbox(
                "Поворот ножа в ячейке",
                [0, 90, 180, 270],
                format_func=lambda x: f"{x}°",
                key="pp_svg_rot",
                help=(
                    "Визуально — вокруг центра ячейки. Для 90° и 270° заново считается укладка на листе "
                    "(обмен габаритов ячейки); для 0° и 180° сетка по базовым ширине/высоте контура."
                ),
            )
        with _pr2:
            st.checkbox(
                "Зеркало ↔",
                key="pp_svg_flip_h",
                help="Отражение макета в ячейке; геометрия сетки слотов не меняется.",
            )
        with _pr3:
            st.checkbox(
                "Зеркало ↕",
                key="pp_svg_flip_v",
                help="Отражение макета в ячейке; геометрия сетки слотов не меняется.",
            )

        pp_png_transparent = st.checkbox(
            "PNG в слотах: прозрачный фон (альфа из PDF)",
            value=False,
            key="pp_png_transparent",
            help=(
                "Рендер PyMuPDF с альфа-каналом: прозрачны незакрашенные области страницы. "
                "Обычная белая подложка в PDF остаётся непрозрачной. "
                "В схеме листа под PNG не рисуется серая подложка — виден белый фон ячейки."
            ),
        )

        _knife_saved_fb = st.session_state.pop(_PP_KNIFE_SAVE_FEEDBACK_KEY, None)
        if _knife_saved_fb:
            st.success(_knife_saved_fb)

        _pp_outline_exp_k = f"pp_outline_exp_{sk_safe}"
        with st.expander(
            "Контур из PDF → SVG (для выбранного размера коробки)",
            key=_pp_outline_exp_k,
        ):
            st.caption(
                f"Габариты группы: **{pp.size_key_display(chosen_sk)}**. "
                "Берётся PDF выбранной позиции (поле файла в БД или сессии Excel). "
                "Обводки в духе эталона Corel (#E61081 и др.); серые вспомогательные линии отсекаются."
            )
            if not er_sorted:
                st.info("В группе нет строк (проверьте фильтр по видам выше).")
            else:
                er_outline = st.selectbox(
                    "Позиция (excel_row) для экспорта контура",
                    options=er_sorted,
                    format_func=_outline_row_label,
                    key=f"pp_outline_er_{sk_safe}",
                )
                row_ol = rows_by_er.get(int(er_outline)) or {}
                rel_ol = (row_ol.get("file") or "").strip()
                p_ol = ppsp.resolve_pdf_path(pdf_root, rel_ol) if rel_ol else None
                if not rel_ol:
                    st.warning("У выбранной строки не указан файл PDF.")
                elif p_ol is None or not p_ol.is_file():
                    st.warning(
                        f"PDF не найден по пути из данных (**{rel_ol}**). Проверьте папку **{pdf_root}**."
                    )
                else:
                    _force_pdf_key = f"pp_outline_force_pdf_once_{sk_safe}_{int(er_outline)}"
                    _fc1, _fc2 = st.columns([1, 2])
                    with _fc1:
                        if st.button(
                            "Перечитать из PDF",
                            key=f"pp_reparse_pdf_{sk_safe}_{er_outline}",
                            help=(
                                "Один раз игнорировать кэш ножа в БД для этой позиции и заново извлечь контур "
                                "и габариты из файла PDF (после обновления макета или смены фильтров обводки)."
                            ),
                        ):
                            st.session_state[_force_pdf_key] = True
                            st.rerun()
                    _force_pdf_once = bool(st.session_state.pop(_force_pdf_key, False))

                    _kn_ol = _knives_print_by_er.get(int(er_outline))
                    _svg_cached = ((_kn_ol or {}).get("svg_full") or "").strip()
                    _w_cached = float((_kn_ol or {}).get("width_mm") or 0)
                    _h_cached = float((_kn_ol or {}).get("height_mm") or 0)
                    outline_from_db = bool(
                        not _force_pdf_once
                        and _svg_cached
                        and _w_cached > 0
                        and _h_cached > 0
                    )
                    if outline_from_db:
                        with _fc2:
                            st.caption(
                                "Сейчас контур из **БД**. «Перечитать из PDF» — снова разобрать макет."
                            )
                    if outline_from_db:
                        svg_ol = _svg_cached
                        pdf_bbox_ol = (_w_cached, _h_cached)
                    else:
                        svg_ol = pdf_outline.extract_outline_svg_from_pdf(
                            str(p_ol), page_index=0
                        )
                        pdf_bbox_ol = pdf_outline.knife_bbox_mm_from_pdf(
                            str(p_ol), page_index=0
                        )
                    pl_imp: list[pp.PlacedRect] = []
                    if pdf_bbox_ol is not None:
                        _pw, _ph = _pp_pack_item_dims_mm(
                            pdf_bbox_ol[0], pdf_bbox_ol[1]
                        )
                        _, pl_imp, _ = pp.pack_shelf_single_item(
                            sheet_params, _pw, _ph
                        )
                    else:
                        fp_ol = pp.footprint_mm_from_size(group["sample_size_str"])
                        if fp_ol is not None:
                            fw_o, fh_o = fp_ol
                            _pw, _ph = _pp_pack_item_dims_mm(fw_o, fh_o)
                            _, pl_imp, _ = pp.pack_shelf_single_item(
                                sheet_params, _pw, _ph
                            )

                    dpi_outline = float(st.session_state.get("pp_preview_dpi", 144))
                    slot_b64_imp: str | None = None
                    if pl_imp and pdf_bbox_ol is not None:
                        png_i = None
                        if p_ol is not None and p_ol.is_file():
                            try:
                                mt_o = float(p_ol.stat().st_mtime)
                            except OSError:
                                mt_o = 0.0
                            png_i = render_pdf_sheet_slot_knife_png(
                                str(p_ol),
                                mt_o,
                                float(pl_imp[0].w),
                                float(pl_imp[0].h),
                                dpi_outline,
                                pp_png_transparent,
                                kind_bucket(row_ol),
                            )
                        if png_i is None and outline_from_db and svg_ol:
                            png_i = packaging_pdf_preview.render_cached_svg_knife_fit_to_mm(
                                svg_ol,
                                float(pl_imp[0].w),
                                float(pl_imp[0].h),
                                dpi=dpi_outline,
                                transparent_bg=pp_png_transparent,
                            )
                        if png_i:
                            slot_b64_imp = base64.b64encode(png_i).decode("ascii")

                    col_ol, col_imp = st.columns(2, gap="medium")
                    with col_ol:
                        st.markdown("**Один нож (контур SVG)**")
                        if svg_ol:
                            b64_ol = base64.b64encode(svg_ol.encode("utf-8")).decode("ascii")
                            row_dl, row_zoom = st.columns([1, 2])
                            with row_dl:
                                st.download_button(
                                    "Скачать SVG контура",
                                    data=svg_ol.encode("utf-8"),
                                    file_name=f"outline_{sk_safe}_{er_outline}.svg",
                                    mime="image/svg+xml",
                                    key=f"pp_outline_dl_{sk_safe}_{er_outline}",
                                )
                                if st.button(
                                    "Сохранить нож в БД",
                                    key=f"pp_save_knife_db_{sk_safe}_{er_outline}",
                                    help=(
                                        "Запись в таблицу knife_cache для выбранной строки Excel — "
                                        "планировщик и «Распространить ножи» подхватят этот SVG."
                                    ),
                                    disabled=not (db_path.is_file() and p_ol is not None and p_ol.is_file()),
                                ):
                                    try:
                                        _cn_sv = pkg_db.connect(db_path)
                                        try:
                                            pkg_db.init_db(_cn_sv)
                                            _payload_sv = pdf_outline.try_extract_knife_from_pdf(
                                                str(p_ol), page_index=0
                                            )
                                            if _payload_sv is None and svg_ol and pdf_bbox_ol:
                                                _payload_sv = (
                                                    svg_ol,
                                                    float(pdf_bbox_ol[0]),
                                                    float(pdf_bbox_ol[1]),
                                                )
                                            if _payload_sv is None and svg_ol:
                                                _bbox_fb = pdf_outline.knife_bbox_mm_from_pdf(
                                                    str(p_ol), page_index=0
                                                )
                                                if _bbox_fb:
                                                    _payload_sv = (
                                                        svg_ol,
                                                        float(_bbox_fb[0]),
                                                        float(_bbox_fb[1]),
                                                    )
                                            if _payload_sv is None:
                                                st.error(
                                                    "Не удалось сформировать SVG и габариты для сохранения "
                                                    "(ни расширенное извлечение, ни превью контура с bbox)."
                                                )
                                            else:
                                                _svg_sv, _w_sv, _h_sv = _payload_sv
                                                if _w_sv <= 0 or _h_sv <= 0:
                                                    st.error("Габариты ножа должны быть больше нуля.")
                                                else:
                                                    pkg_db.save_knife(
                                                        _cn_sv,
                                                        int(er_outline),
                                                        _svg_sv,
                                                        _w_sv,
                                                        _h_sv,
                                                        rel_ol or "",
                                                    )
                                                    _n_prop_grp = pkg_db.propagate_knife_from_donor_to_group_rows(
                                                        _cn_sv,
                                                        int(er_outline),
                                                        same_rows,
                                                    )
                                                    _fb_parts = [
                                                        f"**Нож сохранён в базе данных.** Строка Excel: **{er_outline}**, "
                                                        f"габариты: **{_w_sv:.1f} × {_h_sv:.1f} мм** (`knife_cache`)."
                                                    ]
                                                    if _n_prop_grp:
                                                        _fb_parts.append(
                                                            f" Тот же SVG записан ещё на **{_n_prop_grp}** позиций в этой размерной группе."
                                                        )
                                                    _fb_parts.append(
                                                        " Планировщик и кнопки «Распространить ножи» используют этот SVG."
                                                    )
                                                    st.session_state[_PP_KNIFE_SAVE_FEEDBACK_KEY] = "".join(
                                                        _fb_parts
                                                    )
                                                    st.toast(
                                                        f"Нож: строка {er_outline}, {_w_sv:.1f}×{_h_sv:.1f} мм"
                                                        + (
                                                            f"; +{_n_prop_grp} в группе"
                                                            if _n_prop_grp
                                                            else ""
                                                        ),
                                                        icon="✅",
                                                    )
                                                    st.session_state.pop(
                                                        f"_pl_knives_{chosen_sk}_v2", None
                                                    )
                                                    st.session_state[_pp_outline_exp_k] = True
                                                    st.rerun()
                                        finally:
                                            _cn_sv.close()
                                    except Exception as _e_sv:
                                        st.error(f"Ошибка сохранения в БД: {_e_sv}")
                            with row_zoom:
                                z_preview = st.slider(
                                    "Масштаб просмотра",
                                    min_value=25,
                                    max_value=400,
                                    value=100,
                                    step=5,
                                    format="%d%%",
                                    key=f"pp_outline_zoom_{sk_safe}_{er_outline}",
                                    help="Ширина относительно области предпросмотра; при большом значении появится прокрутка.",
                                )
                            st.markdown(
                                "<div style=\"font-size:0.8rem;opacity:0.85;margin:0 0 0.35rem 0;\">Предпросмотр</div>",
                                unsafe_allow_html=True,
                            )
                            _tr_ol = _pp_svg_preview_transform_css()
                            _sx_ol = (
                                f"width:{int(z_preview)}%;min-width:80px;height:auto;display:block;margin:0 auto;"
                                + (f" {_tr_ol}" if _tr_ol else "")
                            )
                            st.markdown(
                                f'<div style="overflow:auto;max-height:72vh;border:1px solid #e0e0e0;border-radius:6px;'
                                f'padding:10px;background:#fafafa;">'
                                f'<img src="data:image/svg+xml;base64,{b64_ol}" '
                                f'style="{_sx_ol}" '
                                f'alt="контур SVG"/>'
                                f"</div>",
                                unsafe_allow_html=True,
                            )
                        else:
                            st.warning(
                                "Не удалось извлечь подходящие обводки из этого PDF. "
                                "Попробуйте другой цвет/толщину через CLI `extract_pdf_outline_svg.py`."
                            )
                    with col_imp:
                        st.markdown("**Раскладка ножей на лист**")
                        st.caption(
                            "Те же параметры листа и bbox ножа, что и в блоке «Сетка листа» ниже; превью PNG — DPI из слайдера ниже (по умолчанию 144)."
                        )
                        if pl_imp:
                            svg_imp = pp.imposition_preview_svg_mm(
                                sheet_params,
                                pl_imp,
                                slot_image_b64=slot_b64_imp,
                                knife_rotate_deg=int(st.session_state.get("pp_svg_rot", 0)),
                                knife_flip_h=bool(st.session_state.get("pp_svg_flip_h", False)),
                                knife_flip_v=bool(st.session_state.get("pp_svg_flip_v", False)),
                            )
                            b64_imp = base64.b64encode(svg_imp.encode("utf-8")).decode("ascii")
                            st.markdown(
                                '<div style="border:1px solid #e0e0e0;border-radius:6px;padding:8px;background:#fff;">'
                                '<img src="data:image/svg+xml;base64,'
                                f'{b64_imp}" style="width:100%;height:auto;display:block;" alt="раскладка ножей"/>'
                                "</div>",
                                unsafe_allow_html=True,
                            )
                        else:
                            st.caption(
                                "Сетку не построить: нет bbox контура в PDF и нет двух габаритов в «Размер» для группы."
                            )

        st.markdown("#### Сетка листа (нож)")
        inner_w, inner_h = pp.inner_sheet_size(sheet_params)
        pl_active: list[pp.PlacedRect] = []
        n_slots = 0
        fill_active = 0.0
        footprint_ok = False
        grid_fail_note = ""

        _forced = st.session_state.get("_pl_forced_layout")
        _used_forced = False
        if _forced and _forced.get("size_key") == chosen_sk:
            _fn = _forced["n_slots"]
            _fp = _forced["placements"]
            _fkw = _forced.get("knife_w", 0)
            _fkh = _forced.get("knife_h", 0)
            if _fn > 0 and _fp:
                pl_active = [pp.PlacedRect(x=t[0], y=t[1], w=t[2], h=t[3], rotated=t[4]) for t in _fp]
                n_slots = _fn
                sheet_area = inner_w * inner_h
                item_area = sum(p.w * p.h for p in pl_active)
                fill_active = (item_area / sheet_area * 100.0) if sheet_area > 0 else 0.0
                footprint_ok = True
                _used_forced = True
                st.success(
                    f"Раскладка из планировщика: **{n_slots}** слотов "
                    f"({_fkw:.1f}×{_fkh:.1f} мм, зазоры X={sheet_params.gap_mm:g} Y={sheet_params.gap_y_mm:g})"
                )
                st.session_state.pop("_pl_forced_layout", None)

        if not _used_forced:
            if not er_sorted:
                st.info("Нет строк для сетки (проверьте фильтр по видам).")
                st.caption(
                    "Задайте строки группы и позицию **«Позиция для экспорта контура»** в блоке выше. "
                    "Размер ячейки сетки: **сначала кэш ножа в БД** по группе размера, при отсутствии — по PDF опорной строки."
                )
            else:
                er_impos = st.session_state.get(f"pp_outline_er_{sk_safe}")
                if er_impos is None:
                    er_impos = er_sorted[0]
                else:
                    er_impos = int(er_impos)
                    if er_impos not in er_sorted:
                        er_impos = er_sorted[0]
                row_im = rows_by_er.get(int(er_impos)) or {}
                rel_im = (row_im.get("file") or "").strip()
                p_im = ppsp.resolve_pdf_path(pdf_root, rel_im) if rel_im else None
                pdf_bbox = None
                _kr0 = _knives_print_by_er.get(int(er_impos))
                if _kr0 and float(_kr0.get("width_mm", 0)) > 0:
                    pdf_bbox = (
                        float(_kr0["width_mm"]),
                        float(_kr0["height_mm"]),
                    )
                if pdf_bbox is None:
                    for _e_alt in er_same:
                        _ka = _knives_print_by_er.get(int(_e_alt))
                        if _ka and float(_ka.get("width_mm", 0)) > 0:
                            pdf_bbox = (
                                float(_ka["width_mm"]),
                                float(_ka["height_mm"]),
                            )
                            break
                if pdf_bbox is None and rel_im and p_im is not None and p_im.is_file():
                    pdf_bbox = pdf_outline.knife_bbox_mm_from_pdf(str(p_im), page_index=0)
                if pdf_bbox is not None:
                    k_w, k_h = pdf_bbox
                    pk_w, pk_h = _pp_pack_item_dims_mm(k_w, k_h)
                    n_slots, pl_active, fill_active = pp.pack_shelf_single_item(
                        sheet_params, pk_w, pk_h
                    )
                    footprint_ok = n_slots > 0 and bool(pl_active)
                    st.caption(
                        "Размер ячейки и число слотов — **по кэшу SVG-ножа в БД** (вся размерная группа), "
                        "а если в группе ещё нет сохранённого ножа — **по контуру из PDF** опорной строки. "
                        "При повороте ножа **90° или 270°** сетка пересчитывается с учётом обмена ширины и высоты ячейки. "
                        "Ниже вы назначаете коробки в ячейки; превью листа — SVG с растром, контуром и подписью."
                    )
                    st.markdown(
                        f"**Ячейка (нож, для укладки):** {pk_w:.2f}×{pk_h:.2f} мм"
                        + (
                            f" (база контура {k_w:.2f}×{k_h:.2f} мм)"
                            if abs(pk_w - k_w) > 1e-6 or abs(pk_h - k_h) > 1e-6
                            else ""
                        )
                        + f" · **Поле листа:** {inner_w:.1f}×{inner_h:.1f} мм · "
                        f"**Слотов:** {n_slots} · **Заполнение:** {fill_active:.1f}%"
                    )
                else:
                    fp = pp.footprint_mm_from_size(group["sample_size_str"])
                    if fp is None:
                        grid_fail_note = (
                            "Контур из PDF не посчитан и в «Размер» нет двух габаритов для оценки по БД."
                        )
                        st.warning(
                            "Контур из PDF не посчитан (нет векторных обводок по фильтру или нет PDF). "
                            "Оценка по размеру коробки из БД недоступна (нужны два ненулевых габарита в «Размер»)."
                        )
                    else:
                        fw, fh = fp
                        pfw, pfh = _pp_pack_item_dims_mm(fw, fh)
                        n_slots, pl_active, fill_active = pp.pack_shelf_single_item(
                            sheet_params, pfw, pfh
                        )
                        footprint_ok = n_slots > 0 and bool(pl_active)
                        st.warning(
                            "Контур из PDF не посчитан — сетка **по габаритам из БД** для группы размера (не по PDF)."
                        )
                        st.caption(
                            "После появления валидного контура в PDF сетка пересчитается по ножу; число слотов может измениться."
                        )
                        st.markdown(
                            f"**Ячейка (БД):** {fw:.1f}×{fh:.1f} мм · **Поле листа:** {inner_w:.1f}×{inner_h:.1f} мм · "
                            f"**Слотов:** {n_slots} · **Заполнение:** {fill_active:.1f}%"
                        )

        pp_prev_dpi = st.slider("DPI превью PDF", 72, 220, 144, 8, key="pp_preview_dpi")
        pp_show_outline_slots = st.checkbox(
            "Показывать контур вырубки (SVG) в слотах поверх превью",
            value=True,
            key="pp_show_outline_slots",
            help="Тот же отбор обводок, что и в блоке «Контур из PDF → SVG» (#E61081 и др.), вписанный в ячейку.",
        )
        pp_knife_raster_slots = st.checkbox(
            "Растр по bbox контура ножа (обрезка), иначе вся страница",
            value=False,
            key="pp_knife_raster_slots",
            help="PNG строится по ограничивающему прямоугольнику отфильтрованных обводок; если контура нет — как раньше, целая страница.",
        )

        EMPTY_SLOT = -999999
        while EMPTY_SLOT in er_sorted:
            EMPTY_SLOT -= 1
        slot_options = [EMPTY_SLOT] + er_sorted

        def _pp_box_name_db_first(er: int) -> str:
            """Наименование коробки: сначала из SQLite, иначе из сессии (Excel)."""
            br = db_all_by_er.get(int(er))
            if br is not None:
                return (br.get("name") or "").strip()
            rr = rows_by_er.get(int(er))
            return (rr.get("name") or "").strip() if rr else ""

        def _slot_lbl(v: int) -> str:
            if v == EMPTY_SLOT:
                return "— пусто —"
            e = int(v)
            k, nm = _pp_row_kind_name(e)
            if k and nm:
                return f"{e} — {k}: {nm[:58]}"
            if nm:
                return f"{e} — {nm[:65]}"
            if k:
                return f"{e} — {k[:65]}"
            return f"{e} — (нет названия в БД)"

        er_set_slots = set(er_sorted)
        same_rows_for_counts = [r for r in same_rows if int(r["excel_row"]) in er_set_slots]
        if not same_rows_for_counts:
            same_rows_for_counts = list(same_rows)

        layout_mode = st.radio(
            "Как задать раскладку на листе (одна SVG-схема внизу)",
            [
                "По слотам: в каждой ячейке своя коробка этого размера (вид + наименование)",
                "По количеству: сколько ячеек подряд на каждую коробку группы",
            ],
            key=f"pp_layout_mode_{sk_safe}",
            horizontal=True,
        )

        st.markdown("**Назначение коробок по слотам**")
        st.caption(
            "Это **единственное превью печатного листа**: в каждой ячейке сетки ножа — растр PDF, контур SVG и подпись выбранной строки каталога (тот же размер группы). "
            "Режим **«По слотам»**: у слота 1…N свой список. Режим **«По количеству»**: блоки ячеек по строкам таблицы."
        )
        if n_slots == 0:
            st.info("При текущих параметрах листа ни одного слота не помещается.")
            slot_er_list: list[int | None] = []
            slot_labels: list[str] = []
            slot_b64: list[str | None] = []
        elif layout_mode.startswith("По количеству"):
            st.caption(
                "Ячейки заполняются **по порядку строк таблицы** ниже: сначала все ячейки первой позиции (вид + наименование), затем следующей и т.д."
            )
            total_assigned = 0
            counts_by_er: dict[int, int] = {}
            _pp_qty_thumb_px = 56
            _pp_qty_thumb_scale = 0.32
            for r in same_rows_for_counts:
                er = int(r["excel_row"])
                k, nm = _pp_row_kind_name(er)
                if k and nm:
                    line = f"«{k}» — {nm[:42]}"
                elif nm:
                    line = nm[:55]
                elif k:
                    line = f"«{k}»"
                else:
                    line = f"er {er}"
                thumb_col, num_col = st.columns([1, 5], gap="small")
                with thumb_col:
                    row_pdf = rows_by_er.get(er) or {}
                    rel_pdf = (row_pdf.get("file") or "").strip()
                    p_thumb = ppsp.resolve_pdf_path(pdf_root, rel_pdf) if rel_pdf else None
                    if p_thumb is not None and p_thumb.is_file():
                        try:
                            mt_th = float(p_thumb.stat().st_mtime)
                        except OSError:
                            mt_th = 0.0
                        png_th = render_pdf_thumb(
                            str(p_thumb),
                            mt_th,
                            _pp_qty_thumb_scale,
                            sharpness=1.0,
                            max_raster_px=220,
                        )
                        if png_th:
                            st.image(png_th, width=_pp_qty_thumb_px)
                        else:
                            st.caption("—")
                    else:
                        st.caption("—")
                    _qpy_full = (row_pdf.get("qty_per_year") or r.get("qty_per_year") or "").strip()
                    st.caption(format_qty_year_caption(_qpy_full or None))
                with num_col:
                    cnt = st.number_input(
                        f"Ячеек для er {er}: {line}",
                        min_value=0,
                        max_value=n_slots,
                        value=0,
                        step=1,
                        key=f"pp_cellcnt_{sk_safe}_{er}",
                    )
                counts_by_er[er] = int(cnt)
                total_assigned += int(cnt)
            if total_assigned > n_slots:
                st.error(
                    f"Сумма ячеек ({total_assigned}) больше, чем помещается на лист ({n_slots}). Уменьшите значения."
                )
            elif total_assigned < n_slots:
                st.warning(f"На листе останется пустых ячеек: {n_slots - total_assigned} (будут без макета).")
            slot_er_list = []
            for r in same_rows_for_counts:
                er = int(r["excel_row"])
                for _ in range(counts_by_er.get(er, 0)):
                    if len(slot_er_list) < n_slots:
                        slot_er_list.append(er)
            while len(slot_er_list) < n_slots:
                slot_er_list.append(None)
            slot_labels = []
            for idx in range(n_slots):
                er = slot_er_list[idx] if idx < len(slot_er_list) else None
                if er is not None:
                    slot_labels.append(_pp_kind_name_sheet_caption(int(er)))
                else:
                    slot_labels.append("")
        else:
            n_cols = min(4, max(1, n_slots))
            cols_rows = (n_slots + n_cols - 1) // n_cols
            for idx in range(n_slots):
                sk_new = f"pp_slot_{sk_safe}_{idx}"
                if sk_new not in st.session_state:
                    sk_old = f"pp_slot_{chosen_sk}_{idx}"
                    if sk_old in st.session_state:
                        st.session_state[sk_new] = st.session_state[sk_old]
                    else:
                        st.session_state[sk_new] = EMPTY_SLOT

            st.caption(
                "В каждом слоте — **своя коробка** из списка позиций с **этим размером**; на схеме листа ниже — её PDF, контур и подпись."
            )
            for row_i in range(cols_rows):
                cols_slot = st.columns(n_cols)
                for col_j in range(n_cols):
                    idx = row_i * n_cols + col_j
                    if idx >= n_slots:
                        break
                    with cols_slot[col_j]:
                        st.selectbox(
                            f"Слот {idx + 1}: коробка (этот размер)",
                            options=slot_options,
                            format_func=_slot_lbl,
                            key=f"pp_slot_{sk_safe}_{idx}",
                            help="Любая строка группы размера; в ячейке листа — её PDF и контур.",
                        )

            slot_er_list = []
            for idx in range(n_slots):
                v = st.session_state.get(f"pp_slot_{sk_safe}_{idx}", EMPTY_SLOT)
                if v == EMPTY_SLOT:
                    slot_er_list.append(None)
                else:
                    slot_er_list.append(int(v))

            slot_labels = []
            for idx in range(n_slots):
                er = slot_er_list[idx] if idx < len(slot_er_list) else None
                if er is not None:
                    slot_labels.append(_pp_kind_name_sheet_caption(int(er)))
                else:
                    slot_labels.append("")

        n_print_sheets = st.number_input(
            "Сколько печатных листов в партии",
            min_value=1,
            max_value=500_000,
            value=1,
            step=1,
            key=f"pp_n_sheets_{sk_safe}",
            help="Итоговые оттиски = число ячеек этого наименования на одном листе × это количество листов.",
        )

        _pp_pdf_finish_label = (st.session_state.get("pl_finish_type") or "").strip()
        _pp_pdf_finish_code = {
            "Lac WB (водный лак)": "lac_wb",
            "UV без фольги": "uv_no_foil",
            "UV с фольгой": "uv_foil",
        }.get(_pp_pdf_finish_label, "lac_wb")

        sum_rows: list[dict[str, Any]] = []
        if n_slots > 0 and slot_er_list:
            ctr = Counter(er for er in slot_er_list if er is not None)
            empty_cells = sum(1 for er in slot_er_list if er is None)
            for er in sorted(ctr.keys()):
                cells = int(ctr[er])
                row = rows_by_er.get(int(er)) or {}
                qps = pp.parse_qty_per_sheet(row.get("qty_per_sheet"))
                total_impr = cells * int(n_print_sheets)
                sum_rows.append(
                    {
                        "excel_row": er,
                        "Название": _pp_kind_name_sheet_caption(int(er), cap=55) or (_pp_box_name_db_first(int(er)) or "")[:55],
                        "Ячеек на 1 листе": cells,
                        "Листов в партии": int(n_print_sheets),
                        "Всего оттисков (ячейки×листы)": total_impr,
                        "Кол-во на листе (БД)": qps if qps is not None else "—",
                    }
                )
            if empty_cells:
                sum_rows.append(
                    {
                        "excel_row": "—",
                        "Название": f"Пустые ячейки на листе ({empty_cells} шт.)",
                        "Ячеек на 1 листе": empty_cells,
                        "Листов в партии": int(n_print_sheets),
                        "Всего оттисков (ячейки×листы)": empty_cells * int(n_print_sheets),
                        "Кол-во на листе (БД)": "—",
                    }
                )
            st.subheader("Итого по партии")
            st.caption(
                "«Всего оттисков» — сколько раз на печать попадёт макет данной позиции "
                "(при трактовке «одна ячейка = один оттиск этикетки/стороны коробки»). "
                "Столбец **€/1000 (CG)** — оценка по прайсу типографии для выбранного в планировщике типа лакирования "
                "(если открывали планировщик) или **Lac WB** по умолчанию."
            )
            if db_path.is_file() and sum_rows:
                try:
                    _cn_cg = pkg_db.connect(db_path)
                    try:
                        pkg_db.init_db(_cn_cg)
                        _cmap_pdf = pkg_db.load_cg_mapping(_cn_cg)
                        for _ix, _sr in enumerate(sum_rows):
                            _er0 = _sr.get("excel_row")
                            if _er0 in (None, "—"):
                                sum_rows[_ix] = {
                                    **_sr,
                                    "€/1000 (CG)": "—",
                                    "_cutit": "",
                                    "_finish": _pp_pdf_finish_code,
                                }
                                continue
                            _m = _cmap_pdf.get(int(_er0))
                            _ct = (_m or {}).get("cutit_no") if _m else None
                            if not _ct:
                                sum_rows[_ix] = {
                                    **_sr,
                                    "€/1000 (CG)": "—",
                                    "_cutit": "",
                                    "_finish": _pp_pdf_finish_code,
                                }
                                continue
                            _pq = pkg_db.load_cg_prices(_cn_cg, cutit_no=str(_ct))
                            _tq = int(_sr.get("Всего оттисков (ячейки×листы)", 0) or 0)
                            _pv = pkg_db.cg_price_for_qty(
                                _pq, _pp_pdf_finish_code, max(1, _tq)
                            )
                            sum_rows[_ix] = {
                                **_sr,
                                "€/1000 (CG)": f"{_pv:.2f}" if _pv is not None else "—",
                                "_cutit": str(_ct),
                                "_finish": _pp_pdf_finish_code,
                            }
                    finally:
                        _cn_cg.close()
                except Exception:
                    for _ix, _sr in enumerate(sum_rows):
                        if "€/1000 (CG)" not in _sr:
                            sum_rows[_ix] = {**_sr, "€/1000 (CG)": "—"}
            elif sum_rows:
                for _ix, _sr in enumerate(sum_rows):
                    sum_rows[_ix] = {**_sr, "€/1000 (CG)": "—"}
            _df_sum_party = pd.DataFrame(sum_rows)
            _hide_pdf_cols = [c for c in _df_sum_party.columns if str(c).startswith("_")]
            st.dataframe(
                _df_sum_party.drop(columns=_hide_pdf_cols, errors="ignore"),
                use_container_width=True,
                hide_index=True,
            )

        slot_b64_full: list[str | None] = []
        slot_outline_full: list[str | None] = []
        slot_png_full: list[bytes | None] = []
        if pl_active and n_slots:
            slot_b64_full, slot_outline_full, slot_png_full = pse.build_slot_png_and_outline(
                pl_active=pl_active,
                slot_er_list=slot_er_list,
                rows_by_er=rows_by_er,
                pdf_root=pdf_root,
                dpi=float(pp_prev_dpi),
                knife_raster=bool(st.session_state.get("pp_knife_raster_slots", False)),
                transparent_png=pp_png_transparent,
                show_outline=bool(pp_show_outline_slots),
                knives_by_er=_knives_print_by_er,
            )
        n_pl = len(pl_active)
        slot_b64 = [
            slot_b64_full[i] if i < len(slot_b64_full) else None for i in range(n_pl)
        ]
        slot_outline = [
            slot_outline_full[i] if i < len(slot_outline_full) else None for i in range(n_pl)
        ]

        pp_highlight_slot: int | None = None

        st.markdown("#### Схема листа")
        if footprint_ok and pl_active:
            title_svg = f"{pp.size_key_display(chosen_sk)} · слотов: {n_slots}"
            svg = pp.sheet_layout_svg(
                sheet_params,
                pl_active,
                title=title_svg,
                slot_labels=slot_labels or None,
                slot_images_b64=slot_b64 if slot_b64 else None,
                slot_outline_svg_inner=slot_outline
                if (pp_show_outline_slots and slot_outline and any(slot_outline))
                else None,
                highlight_slot_index=pp_highlight_slot,
                slot_image_gray_matte=not pp_png_transparent,
                knife_rotate_deg=int(st.session_state.get("pp_svg_rot", 0)),
                knife_flip_h=bool(st.session_state.get("pp_svg_flip_h", False)),
                knife_flip_v=bool(st.session_state.get("pp_svg_flip_v", False)),
            )
            b64 = base64.b64encode(svg.encode("utf-8")).decode("ascii")
            st.markdown(
                '<img src="data:image/svg+xml;base64,'
                f'{b64}" style="width:25%;max-width:25%;height:auto;display:block;" alt="лист"/>',
                unsafe_allow_html=True,
            )
            svg_export = pp.sheet_layout_svg(
                sheet_params,
                pl_active,
                title=title_svg,
                slot_labels=slot_labels or None,
                slot_images_b64=slot_b64_full if slot_b64_full else None,
                slot_outline_svg_inner=slot_outline_full
                if (pp_show_outline_slots and slot_outline_full and any(slot_outline_full))
                else None,
                highlight_slot_index=pp_highlight_slot,
                slot_image_gray_matte=not pp_png_transparent,
                knife_rotate_deg=int(st.session_state.get("pp_svg_rot", 0)),
                knife_flip_h=bool(st.session_state.get("pp_svg_flip_h", False)),
                knife_flip_v=bool(st.session_state.get("pp_svg_flip_v", False)),
            )
            _ol_vec = (
                slot_outline_full
                if (pp_show_outline_slots and slot_outline_full and any(slot_outline_full))
                else None
            )
            svg_export_vector = pp.sheet_layout_svg(
                sheet_params,
                pl_active,
                title=f"{title_svg} · лист 2: без изображений (слоты и контуры)",
                slot_labels=slot_labels or None,
                slot_images_b64=None,
                slot_outline_svg_inner=_ol_vec,
                highlight_slot_index=pp_highlight_slot,
                slot_image_gray_matte=True,
                knife_rotate_deg=int(st.session_state.get("pp_svg_rot", 0)),
                knife_flip_h=bool(st.session_state.get("pp_svg_flip_h", False)),
                knife_flip_v=bool(st.session_state.get("pp_svg_flip_v", False)),
            )
            _base_fn = f"sheet_{sk_safe}"

            def _export_stats_lines_pdf() -> list[str]:
                lines: list[str] = []
                lines.append(f"Дата экспорта: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
                lines.append(f"Группа размера: {pp.size_key_display(chosen_sk)} (ключ {chosen_sk})")
                lines.append(
                    f"Лист {sheet_params.width_mm:g}×{sheet_params.height_mm:g} мм; поле {sheet_params.margin_mm:g} мм; "
                    f"зазор X {sheet_params.gap_mm:g} мм, Y {sheet_params.gap_y_mm:g} мм"
                )
                lines.append(f"Слотов на листе: {n_slots}; листов в партии: {int(n_print_sheets)}")
                ec = sum(1 for er in slot_er_list if er is None) if slot_er_list else 0
                if ec:
                    lines.append(f"Пустых ячеек на листе: {ec}")
                lines.append("")
                lines.append("Итого по партии:")
                if sum_rows:
                    for row in sum_rows:
                        lines.append(
                            f"  {row.get('excel_row')}: {row.get('Название', '')} | "
                            f"ячеек/лист={row.get('Ячеек на 1 листе')} | листов={row.get('Листов в партии')} | "
                            f"оттисков={row.get('Всего оттисков (ячейки×листы)')} | БД={row.get('Кол-во на листе (БД)')} "
                            f"| €/1000(CG)={row.get('€/1000 (CG)', '—')}"
                        )
                else:
                    lines.append("  (нет строк итогов — задайте назначение слотов)")
                lines.append("")
                lines.append("Назначение по слотам:")
                for idx in range(n_slots):
                    er = slot_er_list[idx] if idx < len(slot_er_list) else None
                    lbl = slot_labels[idx] if idx < len(slot_labels) else ""
                    er_s = str(er) if er is not None else "—"
                    tail = f" — {lbl}" if lbl else ""
                    lines.append(f"  слот {idx + 1}: excel_row={er_s}{tail}")
                lines.append("")
                lines.append(
                    f"Превью/экспорт: DPI {float(pp_prev_dpi):g}; "
                    f"растр по контуру ножа={'да' if st.session_state.get('pp_knife_raster_slots') else 'нет'}; "
                    f"контур в ячейках={'да' if pp_show_outline_slots else 'нет'}; "
                    f"PNG с альфой={'да' if pp_png_transparent else 'нет'}"
                )
                lines.append(
                    f"Нож в ячейке (в SVG): {int(st.session_state.get('pp_svg_rot', 0))}°; "
                    f"зеркало ↔={st.session_state.get('pp_svg_flip_h')}; ↕={st.session_state.get('pp_svg_flip_v')}"
                )
                lines.append(
                    "PDF стр. 1: PNG по слотам + мини-схема раскладки (SVG→PNG) в углу. "
                    "Далее — сводка A4 с таблицами; растровый PDF без поворота/зеркала внутри ячеек (как в SVG-превью)."
                )
                return lines

            def _pp_extras_note_for_pdf() -> str:
                _parts: list[str] = []
                if not db_path.is_file():
                    return "БД не найдена — доплаты планировщика недоступны."
                try:
                    _cx = pkg_db.connect(db_path)
                    try:
                        pkg_db.init_db(_cx)
                        _exrows = pkg_db.load_print_finish_extras(_cx)
                    finally:
                        _cx.close()
                    _byc = {e["code"]: e for e in _exrows}
                    for _sk, _cd in (
                        ("pl_fin_lac_uv", "lac_uv"),
                        ("pl_fin_lac_wb", "lac_wb"),
                        ("pl_fin_foil", "foil"),
                        ("pl_fin_pantone_plus1", "pantone_plus1"),
                    ):
                        if st.session_state.get(_sk) and _cd in _byc:
                            _r = _byc[_cd]
                            _parts.append(
                                f"{_r['label']} (+{float(_r['extra_per_sheet']):.2f}/лист)"
                            )
                except Exception:
                    return ""
                if not _parts:
                    return "Доплаты (лак UV/WB, фольга, Pantone+1): в планировщике не отмечены."
                return "Отмечено для расчёта: " + "; ".join(_parts)

            _econ_raw = st.session_state.get("pl_last_print_economics")
            _econ_block: dict[str, Any]
            if isinstance(_econ_raw, dict) and _econ_raw.get("size_key") == chosen_sk:
                _u = " €" if _econ_raw.get("use_cg") else ""
                _elns = [
                    f"Период потребности: {_econ_raw.get('period', '—')}",
                    f"Сборный тираж: {int(_econ_raw.get('n_sheets', 0)):,} листов · "
                    f"стоимость {float(_econ_raw.get('total_cost', 0)):,.2f}{_u}",
                    f"Раздельная печать: {int(_econ_raw.get('sep_sheets', 0)):,} листов · "
                    f"стоимость {float(_econ_raw.get('sep_cost', 0)):,.2f}{_u}",
                    f"Экономия сборного тиража: {float(_econ_raw.get('savings', 0)):,.2f}{_u} "
                    f"({float(_econ_raw.get('savings_pct', 0)):.1f}% к раздельной)",
                    f"Цена за оттиск: сборно {float(_econ_raw.get('cost_per_imprint_c', 0)):.4f}{_u} · "
                    f"раздельно {float(_econ_raw.get('cost_per_imprint_s', 0)):.4f}{_u}",
                ]
                if _econ_raw.get("finish_label"):
                    _elns.insert(1, f"Лакирование CG в расчёте: {_econ_raw.get('finish_label')}")
                if _econ_raw.get("extras_note"):
                    _elns.append(f"Доплаты (суммарно по отмеченным): {_econ_raw.get('extras_note')}")
                _econ_block = {"has_data": True, "lines": _elns}
            else:
                _econ_block = {"has_data": False, "lines": []}

            _knife_pdf_note = ""
            if pl_active:
                _knife_pdf_note = (
                    f"Ячейка слота 1 на листе: {pl_active[0].w:g} × {pl_active[0].h:g} мм"
                )
                _er_f = next((e for e in slot_er_list if e is not None), None) if slot_er_list else None
                if _er_f is not None:
                    _knf = _knives_print_by_er.get(int(_er_f))
                    if _knf and float(_knf.get("width_mm") or 0) > 0:
                        _knife_pdf_note += (
                            f" · нож в кэше БД: {float(_knf['width_mm']):.1f}×{float(_knf['height_mm']):.1f} мм"
                        )

            _party_pdf = [
                {
                    "name": r.get("Название", ""),
                    "cells": r.get("Ячеек на 1 листе", ""),
                    "sheets": r.get("Листов в партии", ""),
                    "imprints": r.get("Всего оттисков (ячейки×листы)", ""),
                    "db_qps": r.get("Кол-во на листе (БД)", ""),
                    "eur_per_1000": r.get("€/1000 (CG)", "—"),
                }
                for r in sum_rows
            ]
            _slot_pdf = []
            for _si in range(n_slots):
                _er_s = slot_er_list[_si] if _si < len(slot_er_list) else None
                _lb_s = slot_labels[_si] if _si < len(slot_labels) else ""
                _slot_pdf.append({
                    "slot": _si + 1,
                    "er": _er_s if _er_s is not None else "—",
                    "label": _lb_s,
                })
            _cg_finish_disp = {
                "lac_wb": "Lac WB",
                "uv_no_foil": "UV б/ф",
                "uv_foil": "UV+фольга",
            }
            _cg_pdf = []
            for r in sum_rows:
                _er_c = r.get("excel_row")
                if _er_c in (None, "—"):
                    continue
                _fc = str(r.get("_finish", _pp_pdf_finish_code) or "")
                _cg_pdf.append({
                    "er": _er_c,
                    "name": str(r.get("Название", ""))[:48],
                    "cutit": r.get("_cutit", "—"),
                    "finish": _cg_finish_disp.get(_fc, _fc),
                    "qty": r.get("Всего оттисков (ячейки×листы)", ""),
                    "eur_per_1000": r.get("€/1000 (CG)", "—"),
                })

            _pdf_tech_lines = [
                f"Превью/экспорт: DPI {float(pp_prev_dpi):g}; "
                f"растр по контуру ножа={'да' if st.session_state.get('pp_knife_raster_slots') else 'нет'}; "
                f"контур в ячейках={'да' if pp_show_outline_slots else 'нет'}; "
                f"PNG с альфой={'да' if pp_png_transparent else 'нет'}",
                f"Нож в ячейке (в SVG): {int(st.session_state.get('pp_svg_rot', 0))}°; "
                f"зеркало ↔={st.session_state.get('pp_svg_flip_h')}; ↕={st.session_state.get('pp_svg_flip_v')}",
                "Сводка PDF: таблицы + мини-схема (SVG раскладки без растров в ячейках).",
            ]
            _pdf_summary: dict[str, Any] = {
                "layout_svg_bytes": svg_export_vector.encode("utf-8"),
                "sheet_meta": {
                    "w": sheet_params.width_mm,
                    "h": sheet_params.height_mm,
                    "m": sheet_params.margin_mm,
                    "gx": sheet_params.gap_mm,
                    "gy": sheet_params.gap_y_mm,
                    "n_slots": n_slots,
                    "n_sheets": int(n_print_sheets),
                },
                "knife_note": _knife_pdf_note,
                "party_rows": _party_pdf,
                "slot_rows": _slot_pdf,
                "cg_rows": _cg_pdf,
                "tech_lines": _pdf_tech_lines,
                "extras_note": _pp_extras_note_for_pdf(),
                "economics": _econ_block,
                "legacy_stats_lines": _export_stats_lines_pdf(),
            }

            _pdf_bytes = pse.sheet_layout_to_pdf_bytes(
                sheet_params,
                pl_active,
                slot_png_full,
                None,
                title_line=title_svg,
                summary=_pdf_summary,
            )
            # Streamlit: при новом объекте bytes/str на каждом rerun download_button часто не отдаёт файл
            # (см. streamlit#7308). Держим стабильные ссылки в session_state, пока не изменилась раскладка.
            _pp_export_sig = (
                chosen_sk,
                tuple(slot_er_list) if slot_er_list else (),
                int(n_print_sheets),
                float(pp_prev_dpi),
                bool(st.session_state.get("pp_knife_raster_slots", False)),
                bool(pp_png_transparent),
                bool(pp_show_outline_slots),
                int(st.session_state.get("pp_svg_rot", 0) or 0),
                bool(st.session_state.get("pp_svg_flip_h", False)),
                bool(st.session_state.get("pp_svg_flip_v", False)),
                title_svg,
                hashlib.sha256(svg_export.encode("utf-8")).hexdigest()[:24],
            )
            _sig_key = f"_pp_export_sig_{sk_safe}"
            _pdf_key = f"_pp_export_pdf_{sk_safe}"
            _svg_key = f"_pp_export_svg_{sk_safe}"
            if st.session_state.get(_sig_key) != _pp_export_sig:
                st.session_state[_sig_key] = _pp_export_sig
                st.session_state[_pdf_key] = bytes(_pdf_bytes)
                st.session_state[_svg_key] = svg_export.encode("utf-8")
            _pdf_dl = st.session_state.get(_pdf_key) or bytes(_pdf_bytes)
            _svg_dl = st.session_state.get(_svg_key) or svg_export.encode("utf-8")

            dlc1, dlc2 = st.columns(2)
            with dlc1:
                st.download_button(
                    "Скачать схему листа (SVG, все слоты)",
                    data=_svg_dl,
                    file_name=f"{_base_fn}.svg",
                    mime="application/octet-stream",
                    key=f"pp_dl_sheet_svg_{sk_safe}",
                    help="SVG с встроенными растрами; при отказе браузера откройте в новой вкладке или сохраните через «Сохранить как».",
                )
            with dlc2:
                st.download_button(
                    "Скачать лист + сводка (PDF)",
                    data=_pdf_dl,
                    file_name=f"{_base_fn}.pdf",
                    mime="application/pdf",
                    key=f"pp_dl_sheet_pdf_{sk_safe}",
                )
            st.caption(
                "Печать из браузера: **лист 1** — полная раскладка с макетами в слотах; **лист 2** — сетка без растров (контуры). "
                "**Скачать PDF**: стр. 1 — те же PNG + мини-схема раскладки (SVG→PNG) в углу; далее **сводка A4** — таблица партии, "
                "слоты, **€/1000 по CG**, отмеченные в планировщике **лак/foil/Pantone**, экономия сборного тиража (если планировщик "
                "считал этот размер). В диалоге печати выберите ориентацию и «По размеру страницы»."
            )
            _print_html = (
                "<!DOCTYPE html><html><head><meta charset=\"utf-8\"/><style>"
                "body{margin:0;font-family:system-ui,sans-serif;}"
                "#pp-print-btn{margin:8px 0 12px 0;padding:10px 18px;font-size:15px;cursor:pointer;"
                "border-radius:6px;border:1px solid #ccc;background:#f5f5f5;}"
                "#pp-print-btn:hover{background:#eaeaea;}"
                ".pp-svg-wrap{overflow:auto;max-height:520px;border:1px solid #ddd;background:#fafafa;margin-bottom:16px;}"
                ".pp-svg-wrap svg{display:block;width:100%;height:auto;}"
                ".pp-print-h2{font-size:14px;margin:10px 0 6px;color:#222;font-weight:600;}"
                ".pp-page2{page-break-before:always;padding-top:8px;}"
                "@media print{#pp-print-btn{display:none!important;}"
                ".pp-svg-wrap{max-height:none!important;border:none!important;background:#fff!important;margin-bottom:0;}"
                ".pp-print-h2{font-size:11pt;margin:6pt 0 4pt;}"
                ".pp-page2{padding-top:0;}"
                "body{-webkit-print-color-adjust:exact;print-color-adjust:exact;}}"
                "</style></head><body>"
                '<button type="button" id="pp-print-btn" onclick="window.print()">'
                "Печать (лист 1 + лист 2)</button>"
                '<p class="pp-print-h2">Лист 1 — макеты в слотах (растровая раскладка ножей)</p>'
                f"<div class=\"pp-svg-wrap\">{svg_export}</div>"
                '<div class="pp-page2">'
                '<p class="pp-print-h2">Лист 2 — без изображений (схема слотов и контуры, как в SVG)</p>'
                f"<div class=\"pp-svg-wrap\">{svg_export_vector}</div>"
                "</div></body></html>"
            )
            st_components.html(_print_html, height=1180, scrolling=True)
        else:
            st.info(grid_fail_note or "Нет данных для схемы листа (проверьте PDF контура или поле «Размер» в БД).")

    st.divider()
    st.subheader("Заявки")
    up = st.file_uploader("CSV или XLSX", type=["csv", "xlsx", "xls"], key="pp_orders_upload")
    if up is not None:
        try:
            raw = up.read()
            df = pp.read_orders_file(raw, up.name)
            st.session_state["pkg_orders_df"] = df
            st.session_state["pkg_orders_fname"] = up.name
        except Exception as e:
            st.error(f"Не удалось прочитать файл: {e}")

    df_o = st.session_state.get("pkg_orders_df")
    if df_o is not None and not df_o.empty:
        st.caption(f"Файл: {st.session_state.get('pkg_orders_fname', '')} · строк: {len(df_o)}")
        cols = [str(c) for c in df_o.columns.tolist()]
        oc1, oc2, oc3, oc4 = st.columns(4)
        with oc1:
            col_name = st.selectbox("Колонка наименования", cols, key="pp_col_name")
        with oc2:
            col_qty = st.selectbox("Колонка количества", cols, key="pp_col_qty")
        with oc3:
            col_month = st.selectbox("Колонка месяца / даты", cols, key="pp_col_month")
        with oc4:
            default_year = st.number_input("Год, если в файле только месяц 1–12", 2020, 2035, 2026, 1, key="pp_def_year")

        if st.button("Построить записи заявок", key="pp_build_records"):
            rec = pp.build_order_records(df_o, col_name, col_qty, col_month, int(default_year))
            st.session_state["pkg_orders_records"] = rec
            st.session_state.pop("pkg_orders_annotations", None)
            for _k in list(st.session_state.keys()):
                if _k.startswith("pkg_ord_pick_"):
                    st.session_state.pop(_k, None)
            st.session_state.pop("pp_ann_editor", None)
            st.rerun()

    records = st.session_state.get("pkg_orders_records") or []
    if records:
        st.success(f"Заявок с валидными полями: {len(records)} (показаны первые 15)")
        st.dataframe(pd.DataFrame(records[:15]), use_container_width=True, hide_index=True)

        st.subheader("Сопоставление с коробками")
        m1, m2 = st.columns(2)
        with m1:
            om_score = st.slider("Порог совпадения", 35, 95, 50, 1, key="pp_ord_min_score")
        with m2:
            om_gap = st.slider("Разрыв кандидатов", 1, 15, 5, 1, key="pp_ord_gap")
        om_pdf = st.checkbox("Учитывать имя PDF", value=True, key="pp_ord_pdf")

        if st.button("Автосопоставление", type="primary", key="pp_run_match"):
            if not box_rows:
                st.error("Нет коробок в таблице — нельзя сопоставить заявки.")
            else:
                ann, _ = pp.auto_match_all_orders(
                    records,
                    box_rows,
                    min_score=int(om_score),
                    ambiguous_gap=int(om_gap),
                    fallback_pdf=om_pdf,
                )
                st.session_state["pkg_orders_annotations"] = ann
                st.rerun()

        annotations = st.session_state.get("pkg_orders_annotations")
        if annotations:
            n_ok = sum(1 for a in annotations if a["status"] == "ok" and a.get("excel_row"))
            n_bad = len(annotations) - n_ok
            mcol1, mcol2 = st.columns(2)
            with mcol1:
                st.metric("Сопоставлено автоматически", n_ok)
            with mcol2:
                st.metric("Без однозначной пары", n_bad)

            st.caption(
                "Колонка «Вручную excel_row» перекрывает авто. Оставьте пустым, чтобы взять значение из «Авто excel_row»."
            )
            ed_rows: list[dict[str, Any]] = []
            for a in annotations:
                aer = a.get("excel_row")
                ed_rows.append(
                    {
                        "idx": a["idx"],
                        "Название": a["raw_name"][:100],
                        "Кол-во": a["qty"],
                        "Год": a["year"],
                        "Мес": a["month"],
                        "Статус": a["status"],
                        "Балл": a.get("score", ""),
                        "Авто excel_row": int(aer) if aer is not None else None,
                        "Вручную excel_row": None,
                    }
                )
            df_ann = pd.DataFrame(ed_rows)
            edited_ann = st.data_editor(
                df_ann,
                column_config={
                    "idx": st.column_config.NumberColumn("№", disabled=True, format="%d"),
                    "Авто excel_row": st.column_config.NumberColumn("Авто er", disabled=True, format="%d"),
                    "Вручную excel_row": st.column_config.NumberColumn(
                        "Вручную er",
                        min_value=0,
                        step=1,
                        format="%d",
                    ),
                },
                disabled=["Название", "Кол-во", "Год", "Мес", "Статус", "Балл"],
                use_container_width=True,
                height=min(520, 38 * (len(df_ann) + 2)),
                key="pp_ann_editor",
                hide_index=True,
            )

            st.subheader("Приоритеты печати")
            today = date.today()
            pc1, pc2, pc3 = st.columns(3)
            with pc1:
                start_y = st.number_input("Стартовый год", 2020, 2035, today.year, key="pp_pri_y")
            with pc2:
                start_m = st.number_input("Стартовый месяц", 1, 12, today.month, key="pp_pri_m")
            with pc3:
                horizon = st.radio("Горизонт, мес.", [1, 2, 3], horizontal=True, key="pp_horizon")

            sort_by = st.radio(
                "Сортировка",
                ["Спрос (убыв.)", "Листы (убыв.)", "Заполнение листа % (убыв.)"],
                horizontal=True,
                key="pp_sort",
            )

            line_to_er: dict[int, int] = {}
            for _, row in edited_ann.iterrows():
                i = int(row["idx"])
                manual = row["Вручную excel_row"]
                auto = row["Авто excel_row"]
                picked: int | None = None
                if manual is not None and not (isinstance(manual, float) and pd.isna(manual)):
                    try:
                        mi = int(manual)
                        if mi > 0:
                            picked = mi
                    except (TypeError, ValueError):
                        pass
                if picked is None and auto is not None and not (isinstance(auto, float) and pd.isna(auto)):
                    try:
                        ai = int(auto)
                        if ai > 0:
                            picked = ai
                    except (TypeError, ValueError):
                        pass
                if picked is not None:
                    line_to_er[i] = picked

            filtered = pp.filter_orders_in_horizon(records, int(start_y), int(start_m), int(horizon))
            demand = pp.aggregate_demand_by_excel_row(filtered, line_to_er)
            pri_rows = pp.build_priority_rows(demand, rows_by_er, sheet_params, db_all_by_er)
            if not pri_rows:
                st.warning("Нет спроса в выбранном окне месяцев для сопоставленных заявок.")
            else:
                if sort_by.startswith("Спрос"):
                    pri_rows.sort(key=lambda r: -r["demand_qty"])
                elif sort_by.startswith("Листы"):
                    pri_rows.sort(
                        key=lambda r: (
                            1 if r["sheets_estimate"] is None else 0,
                            -(r["sheets_estimate"] or 0),
                            -r["demand_qty"],
                        ),
                    )
                else:
                    pri_rows.sort(
                        key=lambda r: (
                            1 if r["fill_pct_sheet"] is None else 0,
                            -(r["fill_pct_sheet"] or 0),
                            -r["demand_qty"],
                        ),
                    )
                st.dataframe(pd.DataFrame(pri_rows), use_container_width=True, hide_index=True)

    st.divider()
    if st.button("Сбросить заявки в сессии", key="pp_clear_orders"):
        for k in (
            "pkg_orders_df",
            "pkg_orders_fname",
            "pkg_orders_records",
            "pkg_orders_annotations",
            "pp_ann_editor",
        ):
            st.session_state.pop(k, None)
        st.rerun()


def _pl_float_cell_to_qty_str(val: Any) -> str:
    """Значение из data_editor (число/NA) → строка для qty_per_year / qty_per_sheet."""
    import pandas as pd

    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    try:
        f = float(val)
    except (TypeError, ValueError):
        return ""
    if f < 0:
        return ""
    if abs(f - round(f)) < 1e-9:
        return str(int(round(f)))
    s = f"{f:.6f}".rstrip("0").rstrip(".")
    return s


def _pl_apply_volume_sheet_stock_edits(
    edited: Any,
    original: Any,
    *,
    excel_path: Path | None,
    db_path: Path | None,
    conn: sqlite3.Connection | None,
    packaging_rows: list[dict[str, Any]],
) -> tuple[int, int, list[str]]:
    """
    Сравнивает отредактированную таблицу с исходной; пишет объёмы в SQLite + Excel,
    склад — в stock_on_hand по GMP. Возвращает (число строк с правкой объёмов, число правок склада, сообщения).
    """
    import pandas as pd

    msgs: list[str] = []
    if conn is None:
        return 0, 0, ["Нет подключения к базе."]
    er_to_idx = {int(r["excel_row"]): i for i, r in enumerate(packaging_rows)}

    edited_df = edited if isinstance(edited, pd.DataFrame) else pd.DataFrame(edited)
    orig_df = original if isinstance(original, pd.DataFrame) else pd.DataFrame(original)
    if len(edited_df) != len(orig_df):
        return 0, 0, ["Число строк таблицы изменилось; обновите страницу."]

    n_vol = 0
    stock_batch: list[dict[str, Any]] = []

    def _cell_float(series: Any, col: str, i: int) -> float:
        try:
            v = series.iloc[i][col]
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return 0.0
            return float(v)
        except (TypeError, ValueError, KeyError, IndexError):
            return 0.0

    for i in range(len(edited_df)):
        er = int(edited_df.iloc[i]["er"])
        if er not in er_to_idx:
            msgs.append(f"Excel-строка {er}: нет в текущей сессии «Макеты».")
            continue

        o_g = _cell_float(orig_df, "Год (шт)", i)
        e_g = _cell_float(edited_df, "Год (шт)", i)
        o_s = _cell_float(orig_df, "На листе", i)
        e_s = _cell_float(edited_df, "На листе", i)
        o_st = _cell_float(orig_df, "Склад (шт)", i)
        e_st = max(0.0, _cell_float(edited_df, "Склад (шт)", i))

        ch_vol = abs(e_g - o_g) > 1e-6 or abs(e_s - o_s) > 1e-6
        ch_st = abs(e_st - o_st) > 1e-6

        if not ch_vol and not ch_st:
            continue

        idx = er_to_idx[er]
        row = dict(packaging_rows[idx])
        gmp = (row.get("gmp_code") or "").strip().upper() or (
            pkg_db.extract_gmp_code(row.get("name") or "", row.get("file") or "").upper()
        )

        if ch_vol:
            row["qty_per_year"] = _pl_float_cell_to_qty_str(e_g)
            row["qty_per_sheet"] = _pl_float_cell_to_qty_str(e_s)
            packaging_rows[idx] = row
            try:
                pkg_db.upsert_all(conn, [row])
            except Exception as ex:
                msgs.append(f"Строка {er}: БД — {ex}")
                continue
            if excel_path is not None and excel_path.is_file():
                try:
                    save_one_row_to_excel(excel_path, row, db_path)
                except Exception as ex:
                    msgs.append(f"Строка {er}: Excel — {ex}")
            n_vol += 1

        if ch_st:
            if gmp:
                stock_batch.append({"gmp_code": gmp, "qty": e_st})
            else:
                msgs.append(f"Строка {er}: склад не сохранён — нет GMP-кода в названии или файле.")

    n_stock = 0
    if stock_batch:
        try:
            n_stock = pkg_db.upsert_stock_batch(conn, stock_batch, source="planner_manual")
        except Exception as ex:
            msgs.append(f"Склад: {ex}")

    return n_vol, n_stock, msgs


# Чекбоксы доплат к печати в планировщике (коды = print_finish_extras.code)
_PL_PRINT_FIN_CHECK_KEYS: dict[str, str] = {
    "lac_uv": "pl_fin_lac_uv",
    "lac_wb": "pl_fin_lac_wb",
    "foil": "pl_fin_foil",
    "pantone_plus1": "pl_fin_pantone_plus1",
}


def _pl_editable_section_title(text: str) -> None:
    """Заголовок блока с редактируемыми полями (зелёная акцентная полоса)."""
    st.markdown(
        '<p style="color:#1b5e20;font-weight:600;font-size:1.05rem;margin:0.35rem 0 0.55rem 0;'
        'border-left:4px solid #43a047;padding-left:10px;">'
        + html.escape(text)
        + "</p>",
        unsafe_allow_html=True,
    )


def _pl_sum_enabled_print_extras(
    extras: list[dict[str, Any]],
) -> tuple[float, list[dict[str, Any]]]:
    """Сумма доплат за 1 лист по включённым чекбоксам + строки для разбора в анализе цен."""
    total = 0.0
    breakdown: list[dict[str, Any]] = []
    for row in extras:
        code = str(row.get("code") or "")
        key = _PL_PRINT_FIN_CHECK_KEYS.get(code)
        if not key:
            continue
        amt = float(row.get("extra_per_sheet") or 0.0)
        on = bool(st.session_state.get(key, False))
        if on:
            total += amt
        breakdown.append({
            "code": code,
            "label": row.get("label") or code,
            "per_sheet": amt,
            "enabled": on,
        })
    return total, breakdown


def _render_volume_analysis(
    box_rows: list[dict[str, Any]],
    monthly_db: list[dict[str, Any]],
    rows_by_er: dict[int, dict[str, Any]],
    size_groups: list[dict[str, Any]] | None = None,
    *,
    planner_conn: sqlite3.Connection | None = None,
    excel_path: Path | None = None,
    db_path: Path | None = None,
    packaging_rows: list[dict[str, Any]] | None = None,
    stock_by_gmp: dict[str, float] | None = None,
) -> None:
    """Блок «Анализ существующих объёмов» внутри вкладки Планировщик."""
    import pandas as pd
    import packaging_print_planning as pp
    from collections import defaultdict

    if not box_rows:
        st.info("В базе нет печатной продукции с заполненным полем «Вид». Заполните данные на вкладке «Макеты».")
        return

    if size_groups is None:
        size_groups = pp.collect_box_size_groups(box_rows)
    if not size_groups:
        st.info("Нет продукции с заполненным размером.")
        return

    _can_edit_volumes = (
        planner_conn is not None
        and packaging_rows is not None
        and len(packaging_rows) > 0
    )
    _raw_st = stock_by_gmp if stock_by_gmp is not None else {}
    _stock_lookup = {(str(k).strip().upper()): float(v) for k, v in _raw_st.items() if k}

    with st.expander(
        f"Анализ существующих объёмов ({len(size_groups)} размеров, {len(box_rows)} позиций)",
        expanded=False,
    ):
        if _can_edit_volumes:
            st.caption(
                "В детализации по размеру можно вручную изменить **годовой объём**, **кол-во на листе** и **склад (шт)**; "
                "кнопка «Записать» сохраняет объёмы в SQLite и Excel, склад — в таблицу остатков по GMP."
            )
        monthly_by_er: dict[int, list[dict[str, Any]]] = defaultdict(list)
        for m in monthly_db:
            monthly_by_er[int(m["excel_row"])].append(m)

        def _parse_qty(val: str) -> float:
            if not val:
                return 0.0
            cleaned = val.replace(" ", "").replace("\u00a0", "").replace(",", ".")
            try:
                return float(cleaned)
            except (ValueError, TypeError):
                return 0.0

        summary_rows: list[dict[str, Any]] = []
        for sg in size_groups:
            sk = sg["size_key"]
            sk_disp = pp.size_key_display(sk)
            group_items = sg["rows"]
            total_annual = 0.0
            total_monthly_sum = 0.0
            for r in group_items:
                er = int(r["excel_row"])
                full = rows_by_er.get(er) or r
                total_annual += _parse_qty(full.get("qty_per_year") or "")
                for m in monthly_by_er.get(er, []):
                    total_monthly_sum += m["qty"]
            summary_rows.append({
                "Размер": sk_disp,
                "Видов": len(group_items),
                "Годовой объём (из макетов)": int(total_annual) if total_annual else "—",
                "Сумма помесячных (БД)": int(total_monthly_sum) if total_monthly_sum else "—",
                "size_key": sk,
            })

        summary_rows.sort(key=lambda x: (
            0 if isinstance(x["Годовой объём (из макетов)"], int) else 1,
            -(x["Годовой объём (из макетов)"] if isinstance(x["Годовой объём (из макетов)"], int) else 0),
        ))

        summary_df = pd.DataFrame(summary_rows)
        display_df = summary_df.drop(columns=["size_key"])
        st.dataframe(display_df, use_container_width=True, hide_index=True)

        with st.expander("Детализация по каждому размеру", expanded=False):
            import hashlib as _hashlib_vol

            if _can_edit_volumes:
                _pl_editable_section_title("Объёмы и склад по размеру (редактируемые столбцы)")

            for sg in size_groups:
                sk = sg["size_key"]
                sk_disp = pp.size_key_display(sk)
                group_items = sg["rows"]

                st.markdown(f"##### {sk_disp}  ({len(group_items)} видов)")

                detail_rows: list[dict[str, Any]] = []
                for r in group_items:
                    er = int(r["excel_row"])
                    full = rows_by_er.get(er) or r
                    name = (full.get("name") or "")[:60]
                    kind = (full.get("kind") or "")[:30]
                    annual = _parse_qty(full.get("qty_per_year") or "")
                    per_sheet = _parse_qty(full.get("qty_per_sheet") or "")

                    er_monthly = monthly_by_er.get(er, [])
                    monthly_total = sum(m["qty"] for m in er_monthly)

                    months_with_data: list[str] = []
                    for m in sorted(er_monthly, key=lambda x: (x["year"], x["month"])):
                        months_with_data.append(f"{m['month']:02d}/{m['year']}: {int(m['qty'])}")

                    gmp_raw = (full.get("gmp_code") or "").strip() or pkg_db.extract_gmp_code(
                        full.get("name") or "", full.get("file") or ""
                    )
                    gmp_u = gmp_raw.upper() if gmp_raw else ""
                    stock_val = float(_stock_lookup.get(gmp_u, 0.0)) if gmp_u else 0.0

                    detail_rows.append({
                        "er": er,
                        "GMP": gmp_u or "—",
                        "Название": name,
                        "Вид": kind,
                        "Год (шт)": float(annual) if annual else 0.0,
                        "На листе": float(per_sheet) if per_sheet else 0.0,
                        "Склад (шт)": stock_val,
                        "Помесячно (сумма)": int(monthly_total) if monthly_total else 0,
                        "Помесячные данные": "; ".join(months_with_data) if months_with_data else "—",
                    })

                df_detail = pd.DataFrame(detail_rows)
                _h = _hashlib_vol.md5(str(sk).encode("utf-8")).hexdigest()[:12]

                if _can_edit_volumes:
                    edited = st.data_editor(
                        df_detail,
                        column_config={
                            "er": st.column_config.NumberColumn("Excel-стр.", disabled=True, format="%d"),
                            "GMP": st.column_config.TextColumn("GMP", disabled=True, width="small"),
                            "Название": st.column_config.TextColumn("Название", disabled=True, width="large"),
                            "Вид": st.column_config.TextColumn("Вид", disabled=True, width="small"),
                            "Год (шт)": st.column_config.NumberColumn("Год (шт)", min_value=0.0, step=1.0, format="%d"),
                            "На листе": st.column_config.NumberColumn("На листе", min_value=0.0, step=1.0, format="%d"),
                            "Склад (шт)": st.column_config.NumberColumn("Склад (шт)", min_value=0.0, step=1.0, format="%d"),
                            "Помесячно (сумма)": st.column_config.NumberColumn(
                                "Помес. сумма", disabled=True, format="%d"
                            ),
                            "Помесячные данные": st.column_config.TextColumn("Помесячные данные", disabled=True),
                        },
                        hide_index=True,
                        num_rows="fixed",
                        use_container_width=True,
                        key=f"pl_vol_ed_{_h}",
                    )
                    if st.button(
                        f"Записать правки для {sk_disp}",
                        key=f"pl_vol_save_{_h}",
                        type="primary",
                    ):
                        n_v, n_st, mlist = _pl_apply_volume_sheet_stock_edits(
                            edited,
                            df_detail,
                            excel_path=excel_path,
                            db_path=db_path,
                            conn=planner_conn,
                            packaging_rows=packaging_rows,
                        )
                        for m in mlist:
                            st.warning(m)
                        if n_v or n_st:
                            st.success(
                                f"Сохранено: объёмы — {n_v} поз., склад — {n_st} GMP."
                            )
                            st.rerun()
                        elif not mlist:
                            st.info("Нет изменений для записи.")
                else:
                    disp_rows: list[dict[str, Any]] = []
                    for row in detail_rows:
                        g = float(row.get("Год (шт)") or 0)
                        s = float(row.get("На листе") or 0)
                        mt = int(row.get("Помесячно (сумма)") or 0)
                        disp_rows.append({
                            "Название": row["Название"],
                            "Вид": row["Вид"],
                            "Год. объём": int(g) if g else "—",
                            "На листе": int(s) if s else "—",
                            "Склад (шт)": int(float(row.get("Склад (шт)") or 0)),
                            "Помесячно (сумма)": mt if mt else "—",
                            "Помесячные данные": row["Помесячные данные"],
                        })
                    st.dataframe(pd.DataFrame(disp_rows), use_container_width=True, hide_index=True)

                st.divider()

    st.divider()


_auto_match_cg = auto_match_cg


def _parse_packaging_price(val: Any) -> float | None:
    """Парсинг цены из полей price / price_new (запятая как десятичный разделитель)."""
    if val is None:
        return None
    s = str(val).strip().replace("\u00a0", "").replace(" ", "")
    if not s:
        return None
    s = s.replace(",", ".")
    try:
        v = float(s)
        return v if v > 0 else None
    except (ValueError, TypeError):
        return None


def _group_cutit_no(
    sg_rows: list[dict[str, Any]],
    cg_mapping: dict[int, dict[str, Any]],
) -> str | None:
    for r in sg_rows:
        er = int(r["excel_row"])
        m = cg_mapping.get(er)
        if m and m.get("cutit_no"):
            return str(m["cutit_no"])
    return None


def _group_knife_dims_display(
    knife_meta: dict[int, dict[str, Any]],
    ers: list[int],
) -> tuple[str, float | None, float | None]:
    """Текст для таблицы и эталонные w,h (первый нож из самого частого кластера 5 мм)."""
    from collections import Counter

    cluster: Counter[tuple[float, float]] = Counter()
    er_by_c: dict[tuple[float, float], list[int]] = {}
    for er in ers:
        m = knife_meta.get(er)
        if not m or m.get("width_mm", 0) <= 0:
            continue
        w5 = round(float(m["width_mm"]) / 5) * 5
        h5 = round(float(m["height_mm"]) / 5) * 5
        key = (w5, h5)
        cluster[key] += 1
        er_by_c.setdefault(key, []).append(er)
    if not cluster:
        return ("—", None, None)
    best_k, _ = cluster.most_common(1)[0]
    for er in er_by_c[best_k]:
        m = knife_meta[er]
        return (
            f"{m['width_mm']:.1f}×{m['height_mm']:.1f}",
            float(m["width_mm"]),
            float(m["height_mm"]),
        )
    return ("—", None, None)


def render_planner_tab(
    packaging_rows: list[dict[str, Any]],
    db_path: Path,
    pdf_dir: Path | None = None,
    excel_path: Path | None = None,
) -> None:
    """Планировщик оптимизации печати (четвёртая вкладка)."""
    import hashlib as _hashlib_pl
    import packaging_sheet_export as pse
    import pandas as pd
    import packaging_print_planning as pp
    import pdf_outline_to_svg as posv
    from packaging_pdf_sheet_preview import resolve_pdf_path

    st.title("Планировщик оптимизации печати")
    st.caption(
        "Анализирует существующие объёмы, извлекает ножи из PDF-макетов, "
        "рассчитывает оптимальную раскладку разных видов упаковки одного размера на печатный лист. "
        "Заголовки с **зелёной полосой слева** отмечают блоки с редактируемыми полями."
    )

    rows_by_er_session = {int(r["excel_row"]): r for r in packaging_rows}
    db_all_by_er: dict[int, dict[str, Any]] = {}
    box_rows: list[dict[str, Any]] = []
    monthly_db: list[dict[str, Any]] = []
    knife_meta: dict[int, dict[str, Any]] = {}
    _stock_db: dict[str, float] = {}
    _cg_knives: list[dict[str, Any]] = []
    _cg_prices: list[dict[str, Any]] = []
    _cg_mapping: dict[int, dict[str, Any]] = {}
    _planner_conn: sqlite3.Connection | None = None
    if db_path.is_file():
        try:
            _planner_conn = pkg_db.connect(db_path)
            pkg_db.init_db(_planner_conn)
            if pkg_db.row_count(_planner_conn) > 0:
                db_loaded = pkg_db.load_all(_planner_conn)
                for r in db_loaded:
                    db_all_by_er[int(r["excel_row"])] = r
                box_rows = pp.sheet_layout_candidate_rows(db_loaded)
            monthly_db = pkg_db.load_monthly_for_rows(_planner_conn, [])
            knife_meta = pkg_db.load_knives_meta(_planner_conn)
            _stock_db = pkg_db.load_stock(_planner_conn)
            _cg_knives = pkg_db.load_cg_knives(_planner_conn)
            _cg_prices = pkg_db.load_cg_prices(_planner_conn)
            _cg_mapping = pkg_db.load_cg_mapping(_planner_conn)
        except Exception as e:
            st.warning(f"Не удалось прочитать базу: {e}")
            if _planner_conn:
                try:
                    _planner_conn.close()
                except Exception:
                    pass
                _planner_conn = None
    rows_by_er = dict(db_all_by_er)
    rows_by_er.update(rows_by_er_session)

    st.markdown("**Фильтр по виду упаковки**")
    st.caption(
        "Без галочек — **коробки, блистеры и пакеты** (этикетки скрыты). Отметьте виды для узкого фильтра; "
        "**Этикетка** — только строки с явным видом «Этикетка» (или слово «этикет» в «Вид»), как на вкладке «Печать»."
    )
    _pl_c1, _pl_c2, _pl_c3, _pl_c4, _ = st.columns([1.0, 1.0, 1.0, 1.0, 4.5])
    with _pl_c1:
        _pl_filter_box = st.checkbox("Коробка", value=False, key="pl_layout_filter_box")
    with _pl_c2:
        _pl_filter_blister = st.checkbox("Блистер", value=False, key="pl_layout_filter_blister")
    with _pl_c3:
        _pl_filter_pack = st.checkbox("Пакет", value=False, key="pl_layout_filter_pack")
    with _pl_c4:
        _pl_filter_label = st.checkbox("Этикетка", value=False, key="pl_layout_filter_label")

    def _pl_layout_bucket_for_row(r: dict[str, Any]) -> str:
        er = int(r["excel_row"])
        merged = rows_by_er.get(er) or r
        return kind_bucket(merged)

    _pl_any_kind_filter = (
        _pl_filter_box
        or _pl_filter_blister
        or _pl_filter_pack
        or _pl_filter_label
    )
    if _pl_any_kind_filter:
        box_rows_for_planner = [
            r
            for r in box_rows
            if (_pl_filter_box and _pl_layout_bucket_for_row(r) == "box")
            or (_pl_filter_blister and _pl_layout_bucket_for_row(r) == "blister")
            or (_pl_filter_pack and _pl_layout_bucket_for_row(r) == "pack")
            or (_pl_filter_label and _pl_layout_bucket_for_row(r) == "label")
        ]
    else:
        box_rows_for_planner = [
            r for r in box_rows if _pl_layout_bucket_for_row(r) != "label"
        ]

    # ── Анализ существующих объёмов ──
    size_groups = pp.collect_box_size_groups(box_rows_for_planner)

    if not size_groups and box_rows:
        if _pl_any_kind_filter and not box_rows_for_planner:
            st.warning(
                "С выбранным фильтром вида нет строк в БД. Снимите галочки или проверьте поле «Вид»."
            )
        elif _pl_any_kind_filter and box_rows_for_planner:
            st.warning(
                "У отфильтрованных позиций нет валидного «Размер (мм)» для групп — заполните размер на «Макеты»."
            )
        elif not box_rows_for_planner:
            st.info(
                "В базе для планировщика только **этикетки** — отметьте галочку **«Этикетка»**, чтобы учитывать их в анализе и оптимизации."
            )

    # ── Автораспространение ножей по размерным группам ──
    if _planner_conn is not None and size_groups and not st.session_state.get(
        _PKG_KNIFE_PROPAGATE_SESSION_KEY
    ):
        _n_propagated = pkg_db.propagate_knives_in_size_groups(
            _planner_conn, size_groups, knife_meta
        )
        if _n_propagated > 0:
            st.toast(f"Ножи распространены на {_n_propagated} позиций по размерным группам")
        st.session_state[_PKG_KNIFE_PROPAGATE_SESSION_KEY] = True

    _render_volume_analysis(
        box_rows_for_planner,
        monthly_db,
        rows_by_er,
        size_groups=size_groups,
        planner_conn=_planner_conn,
        excel_path=excel_path,
        db_path=db_path if db_path.is_file() else None,
        packaging_rows=packaging_rows,
        stock_by_gmp=_stock_db,
    )

    # ── Складские остатки ──
    with st.expander(
        f"Остатки на складе ({len(_stock_db)} позиций загружено)",
        expanded=False,
    ):
        st.caption(
            "Загрузите Excel-файл с остатками на складе. "
            "Файл должен содержать колонку с GMP-кодом (например, ВУМ-169-01) "
            "и колонку с количеством. Система найдёт их автоматически. "
            "Скачайте шаблон с предзаполненными GMP-кодами из каталога."
        )

        # ── Шаблон для загрузки остатков ──
        def _build_stock_template() -> bytes:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.worksheet.datavalidation import DataValidation
            from io import BytesIO

            _STOCK_UNITS = ("шт.", "кг", "г", "л", "мл")

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Остатки на складе"

            hdr_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
            hdr_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin", color="B4C6E7"),
                right=Side(style="thin", color="B4C6E7"),
                top=Side(style="thin", color="B4C6E7"),
                bottom=Side(style="thin", color="B4C6E7"),
            )

            headers_tpl = ["GMP-код", "Название", "Остаток", "Ед. изм."]
            for ci, h in enumerate(headers_tpl, start=1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = hdr_align
                cell.border = thin_border

            gmp_set: dict[str, str] = {}
            for r in packaging_rows:
                g = (r.get("gmp_code") or "").strip().upper()
                if not g:
                    g = pkg_db.extract_gmp_code(
                        r.get("name") or "", r.get("file") or ""
                    ).strip().upper()
                if g and g not in gmp_set:
                    gmp_set[g] = (r.get("name") or "").strip()[:80]

            unit_list = ",".join(_STOCK_UNITS)
            dv_unit = DataValidation(
                type="list",
                formula1=f'"{unit_list}"',
                allow_blank=True,
                showDropDown=False,
            )
            dv_unit.error = "Выберите единицу из списка"
            dv_unit.errorTitle = "Недопустимая единица"
            dv_unit.prompt = "Выберите единицу измерения"
            dv_unit.promptTitle = "Ед. изм."
            ws.add_data_validation(dv_unit)

            data_font = Font(name="Calibri", size=11)
            qty_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            for ri, (gmp, name) in enumerate(sorted(gmp_set.items()), start=2):
                c1 = ws.cell(row=ri, column=1, value=gmp)
                c1.font = data_font
                c1.border = thin_border
                c2 = ws.cell(row=ri, column=2, value=name)
                c2.font = data_font
                c2.border = thin_border
                c3 = ws.cell(row=ri, column=3, value=_stock_db.get(gmp, 0))
                c3.font = data_font
                c3.fill = qty_fill
                c3.border = thin_border
                c3.number_format = "#,##0.##"
                c4 = ws.cell(row=ri, column=4, value="шт.")
                c4.font = data_font
                c4.border = thin_border
                c4.alignment = Alignment(horizontal="center")
                dv_unit.add(c4)

            last_row = max(len(gmp_set) + 1, 2)
            for extra_ri in range(last_row + 1, last_row + 51):
                c4e = ws.cell(row=extra_ri, column=4)
                dv_unit.add(c4e)

            ws.column_dimensions["A"].width = 18
            ws.column_dimensions["B"].width = 50
            ws.column_dimensions["C"].width = 14
            ws.column_dimensions["D"].width = 12
            ws.auto_filter.ref = f"A1:D{last_row}"

            buf = BytesIO()
            wb.save(buf)
            return buf.getvalue()

        _tpl_col, _upl_col = st.columns([1, 2], gap="small")
        with _tpl_col:
            st.download_button(
                "Скачать шаблон",
                data=_build_stock_template(),
                file_name="stock_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="pl_stock_template_dl",
                use_container_width=True,
                help="Excel-шаблон с GMP-кодами из каталога. Заполните колонку «Остаток» и загрузите обратно.",
            )
        with _upl_col:
            _stock_file = st.file_uploader(
                "Excel-файл с остатками",
                type=["xlsx", "xls"],
                key="pl_stock_upload",
            )
        if _stock_file is not None:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(_stock_file, data_only=True, read_only=True)
                ws = wb.active
                headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]

                import re as _re_stock
                gmp_col_idx: int | None = None
                qty_col_idx: int | None = None
                unit_col_idx: int | None = None
                name_col_idx: int | None = None
                for ci, h in enumerate(headers):
                    hl = h.lower()
                    if gmp_col_idx is None and any(
                        kw in hl for kw in ("код", "code", "gmp", "шифр", "cod", "articol")
                    ):
                        gmp_col_idx = ci
                    if gmp_col_idx is None and _re_stock.search(
                        r"[а-яa-z]{2,4}-\d{2,4}-\d{2}", hl
                    ):
                        gmp_col_idx = ci
                    if qty_col_idx is None and any(
                        kw in hl
                        for kw in ("кол", "qty", "остат", "stock", "stoc", "cantitat", "buc")
                    ):
                        qty_col_idx = ci
                    if unit_col_idx is None and any(
                        kw in hl
                        for kw in ("ед", "unit", "изм", "um", "măsur")
                    ):
                        unit_col_idx = ci
                    if name_col_idx is None and any(
                        kw in hl
                        for kw in ("назван", "name", "наименован", "denumir", "produs")
                    ):
                        name_col_idx = ci

                if gmp_col_idx is None or qty_col_idx is None:
                    for ri, row_vals in enumerate(ws.iter_rows(min_row=2, max_row=5, values_only=True)):
                        if ri > 3:
                            break
                        for ci, val in enumerate(row_vals):
                            s = str(val or "")
                            if gmp_col_idx is None and _re_stock.search(
                                r"[A-ZА-Яa-zа-я]{2,4}-\d{2,4}-\d{2}", s
                            ):
                                gmp_col_idx = ci
                            if qty_col_idx is None and ci != gmp_col_idx:
                                try:
                                    float(str(val).replace(" ", "").replace(",", "."))
                                    qty_col_idx = ci
                                except (ValueError, TypeError):
                                    pass

                wb.close()

                if gmp_col_idx is None:
                    st.error(
                        "Не найдена колонка с GMP-кодом. "
                        "Убедитесь, что в файле есть колонка с кодами вида ВУМ-169-01."
                    )
                elif qty_col_idx is None:
                    st.error("Не найдена колонка с количеством.")
                else:
                    _valid_units = {"шт.", "кг", "г", "л", "мл"}
                    wb2 = openpyxl.load_workbook(_stock_file, data_only=True, read_only=True)
                    ws2 = wb2.active
                    _stock_entries: list[dict[str, Any]] = []
                    _preview_rows: list[dict[str, Any]] = []
                    for row_vals in ws2.iter_rows(min_row=2, values_only=True):
                        vals = list(row_vals)
                        if gmp_col_idx >= len(vals) or qty_col_idx >= len(vals):
                            continue
                        raw_code = str(vals[gmp_col_idx] or "").strip()
                        m_code = _re_stock.search(
                            r"([A-ZА-Яa-zа-я]{2,4}-\d{2,4}-\d{2})", raw_code
                        )
                        code = m_code.group(1).upper() if m_code else raw_code.upper()
                        if not code:
                            continue
                        try:
                            qty = float(
                                str(vals[qty_col_idx] or "0")
                                .replace(" ", "")
                                .replace("\u00a0", "")
                                .replace(",", ".")
                            )
                        except (ValueError, TypeError):
                            continue
                        if qty < 0:
                            continue
                        unit_val = "шт."
                        if unit_col_idx is not None and unit_col_idx < len(vals):
                            raw_unit = str(vals[unit_col_idx] or "").strip().lower().rstrip(".")
                            for vu in _valid_units:
                                if raw_unit == vu.rstrip(".").lower():
                                    unit_val = vu
                                    break
                        name_val = ""
                        if name_col_idx is not None and name_col_idx < len(vals):
                            name_val = str(vals[name_col_idx] or "").strip()[:80]
                        _stock_entries.append({
                            "gmp_code": code, "qty": qty, "unit": unit_val, "name": name_val,
                        })
                        if len(_preview_rows) < 20:
                            _preview_rows.append({
                                "GMP-код": code,
                                "Название": name_val or "—",
                                "Остаток": qty if unit_val != "шт." else int(qty),
                                "Ед. изм.": unit_val,
                            })
                    wb2.close()

                    _cols_found = [f"«{headers[gmp_col_idx]}» → код", f"«{headers[qty_col_idx]}» → кол-во"]
                    if unit_col_idx is not None:
                        _cols_found.append(f"«{headers[unit_col_idx]}» → ед. изм.")
                    if name_col_idx is not None:
                        _cols_found.append(f"«{headers[name_col_idx]}» → название")
                    st.info(
                        f"Найдено **{len(_stock_entries)}** позиций "
                        f"(колонки: {', '.join(_cols_found)})."
                    )
                    if _preview_rows:
                        st.dataframe(
                            pd.DataFrame(_preview_rows),
                            use_container_width=True,
                            hide_index=True,
                        )
                    if _stock_entries and st.button(
                        f"Загрузить {len(_stock_entries)} позиций в базу",
                        key="pl_stock_import_btn",
                        type="primary",
                    ):
                        if _planner_conn is not None:
                            n_saved = pkg_db.upsert_stock_batch(
                                _planner_conn, _stock_entries, source=_stock_file.name
                            )
                            _stock_db = pkg_db.load_stock(_planner_conn)
                            st.success(f"Загружено {n_saved} позиций складских остатков.")
                            st.rerun()
                        else:
                            st.error("Нет подключения к БД.")
            except Exception as e:
                st.error(f"Ошибка чтения файла: {e}")

        if _stock_db:
            st.markdown(
                f"**В базе:** {len(_stock_db)} позиций, "
                f"суммарный остаток: {int(sum(_stock_db.values())):,} шт."
            )
            if st.button("Очистить все остатки", key="pl_stock_clear"):
                if _planner_conn is not None:
                    pkg_db.clear_stock(_planner_conn)
                    _stock_db = {}
                    st.success("Складские остатки очищены.")
                    st.rerun()

    # ── Прайс типографии (CG) ──
    _cg_label = f"Прайс типографии ({len(_cg_knives)} ножей загружено)" if _cg_knives else "Прайс типографии"
    with st.expander(_cg_label, expanded=False):
        st.caption(
            "Загрузите Excel-файл с ценами типографии (CG Preț). "
            "Лист «Cutii» будет разобран автоматически: ножи, тиражные ступени, цены."
        )
        _cg_file = st.file_uploader("Excel CG Preț", type=["xlsx", "xls"], key="pl_cg_upload")
        if _cg_file is not None:
            try:
                _parsed_knives, _parsed_prices = parse_cg_pret_workbook(_cg_file)

                st.info(f"Разобрано **{len(_parsed_knives)}** ножей, **{len(_parsed_prices)}** ценовых ступеней.")

                _has_diff_prices = any(
                    p.get("price_old_per_1000") and p["price_old_per_1000"] != p["price_per_1000"]
                    for p in _parsed_prices
                )
                if _has_diff_prices:
                    st.caption("Ячейки с двумя ценами: первая → старая, последняя → новая (актуальная)")

                if _parsed_knives:
                    _cg_preview: list[dict[str, Any]] = []
                    for k in _parsed_knives[:15]:
                        pr = [p for p in _parsed_prices if p["cutit_no"] == k["cutit_no"]]
                        _old_pr = [p["price_old_per_1000"] for p in pr if p.get("price_old_per_1000")]
                        _new_pr = [p["price_per_1000"] for p in pr if p["price_per_1000"]]
                        row_d: dict[str, Any] = {
                            "Нож": k["cutit_no"],
                            "Название": k["name"][:50],
                            "Категория": k["category"][:30],
                            "Ступеней": len(pr),
                        }
                        if _old_pr:
                            row_d["Стар."] = f"{min(_old_pr):.1f}–{max(_old_pr):.1f}"
                        if _new_pr:
                            row_d["Нов."] = f"{min(_new_pr):.1f}–{max(_new_pr):.1f}"
                        _cg_preview.append(row_d)
                    st.dataframe(pd.DataFrame(_cg_preview), use_container_width=True, hide_index=True)

                if _parsed_knives and st.button(
                    f"Загрузить {len(_parsed_knives)} ножей и {len(_parsed_prices)} цен в базу",
                    key="pl_cg_import_btn",
                    type="primary",
                ):
                    if _planner_conn is not None:
                        pkg_db.clear_cg_data(_planner_conn)
                        n_k = pkg_db.upsert_cg_knives(_planner_conn, _parsed_knives)
                        n_p = pkg_db.upsert_cg_prices(_planner_conn, _parsed_prices)

                        _auto_map = _auto_match_cg(
                            _parsed_knives, rows_by_er, box_rows
                        )
                        if _auto_map:
                            pkg_db.upsert_cg_mapping(_planner_conn, _auto_map)

                        _cg_knives = pkg_db.load_cg_knives(_planner_conn)
                        _cg_prices = pkg_db.load_cg_prices(_planner_conn)
                        _cg_mapping = pkg_db.load_cg_mapping(_planner_conn)
                        st.success(
                            f"Загружено {n_k} ножей, {n_p} цен. "
                            f"Автосопоставлено: {len(_auto_map)} продуктов."
                        )
                        st.rerun()
            except Exception as e:
                st.error(f"Ошибка чтения CG файла: {e}")

        if _cg_knives:
            st.markdown(f"**В базе:** {len(_cg_knives)} ножей, {len(_cg_prices)} ценовых ступеней, {len(_cg_mapping)} сопоставлений")

            st.markdown("### Сравнение прайса: наша база vs типография CG")
            st.caption(
                "Наша цена — поля «Цена» / «Цена новая» в базе (за 1 шт.). CG — € за 1000 шт. по выбранной ступени. "
                "Эквивалент за 1 шт. CG = цена÷1000; столбец «Разница %» = (CG за шт. − наша новая ср.) / наша новая ср."
            )
            _pl_finish_labels = {
                "lac_wb": "Lac WB",
                "uv_no_foil": "UV без фольги",
                "uv_foil": "UV с фольги",
            }
            _ft_list_cmp = sorted(set(p["finish_type"] for p in _cg_prices))
            if not _ft_list_cmp:
                st.info("Нет строк цен CG — перезагрузите Excel прайса.")
            if _ft_list_cmp:
                _sel_pl_finish = st.selectbox(
                    "Тип лакирования (для сводной таблицы)",
                    options=_ft_list_cmp,
                    format_func=lambda x: _pl_finish_labels.get(x, x),
                    key="pl_cg_cmp_finish",
                )
                _tier_seen_cmp: set[int] = set()
                _tier_opts_cmp: list[tuple[int, str]] = []
                for p in sorted(_cg_prices, key=lambda x: (x["finish_type"], x["min_qty"])):
                    if p["finish_type"] != _sel_pl_finish:
                        continue
                    mn = int(p["min_qty"])
                    if mn in _tier_seen_cmp:
                        continue
                    _tier_seen_cmp.add(mn)
                    mx = p.get("max_qty")
                    lab = f"{mn:,} шт — " + (
                        f"до {int(mx):,}" if mx is not None else "от порога и выше"
                    )
                    _tier_opts_cmp.append((mn, lab))
                if _tier_opts_cmp:
                    _tier_labels_cmp = [t[1] for t in _tier_opts_cmp]
                    _ti_cmp = st.selectbox(
                        "Тиражная ступень CG",
                        options=list(range(len(_tier_opts_cmp))),
                        format_func=lambda i: _tier_labels_cmp[int(i)],
                        key="pl_cg_cmp_tier",
                    )
                    _sel_pl_min_qty = _tier_opts_cmp[int(_ti_cmp)][0]
                else:
                    _sel_pl_min_qty = 1000

                _cmp_rows: list[dict[str, Any]] = []
                _n_with_cg_cmp = 0
                _n_no_cg_cmp = 0
                for sg in size_groups:
                    sk = sg["size_key"]
                    sk_disp = pp.size_key_display(sk)
                    gitems = sg["rows"]
                    cutit_cmp = _group_cutit_no(gitems, _cg_mapping)
                    new_list_cmp: list[float] = []
                    for r in gitems:
                        er = int(r["excel_row"])
                        full = rows_by_er.get(er) or r
                        pn = _parse_packaging_price(full.get("price_new"))
                        if pn is not None:
                            new_list_cmp.append(pn)
                    our_new_avg = (
                        sum(new_list_cmp) / len(new_list_cmp) if new_list_cmp else None
                    )
                    cg_o_cmp, cg_n_cmp = (None, None)
                    if cutit_cmp:
                        cg_o_cmp, cg_n_cmp = pkg_db.cg_price_pair_at_tier(
                            _cg_prices, cutit_cmp, _sel_pl_finish, _sel_pl_min_qty
                        )
                    if cutit_cmp and cg_n_cmp is not None:
                        _n_with_cg_cmp += 1
                    else:
                        _n_no_cg_cmp += 1
                    diff_s_cmp = "—"
                    if our_new_avg and cg_n_cmp is not None:
                        cg_unit = cg_n_cmp / 1000.0
                        diff_s_cmp = (
                            f"{((cg_unit - our_new_avg) / our_new_avg * 100):+.1f}%"
                        )
                    _cmp_rows.append({
                        "size_key": sk,
                        "Размер": sk_disp,
                        "Видов": len(gitems),
                        "Нож CG": cutit_cmp or "—",
                        "Наша (нов.) ср.": f"{our_new_avg:.4f}" if our_new_avg else "—",
                        "CG нов. €/1000": f"{cg_n_cmp:.2f}" if cg_n_cmp is not None else "—",
                        "Разница %": diff_s_cmp,
                    })
                st.markdown(
                    f"**Размеров с привязкой CG и ценой на ступени:** {_n_with_cg_cmp} · "
                    f"**без CG / без цены:** {_n_no_cg_cmp}"
                )
                _cmp_df = pd.DataFrame(
                    [{k: v for k, v in r.items() if k != "size_key"} for r in _cmp_rows]
                )
                st.dataframe(_cmp_df, use_container_width=True, hide_index=True, height=380)

                _detail_choices_cmp = ["— не выбрано —"] + [r["Размер"] for r in _cmp_rows]
                _det_pick_cmp = st.selectbox(
                    "Детализация по размеру (все виды в группе)",
                    options=_detail_choices_cmp,
                    key="pl_cg_cmp_detail",
                )
                if _det_pick_cmp != "— не выбрано —":
                    _pick_sk_cmp = None
                    for r in _cmp_rows:
                        if r["Размер"] == _det_pick_cmp:
                            _pick_sk_cmp = r["size_key"]
                            break
                    if _pick_sk_cmp:
                        _sg_pick_cmp = next(
                            (g for g in size_groups if g["size_key"] == _pick_sk_cmp),
                            None,
                        )
                        if _sg_pick_cmp:
                            cutit_d = _group_cutit_no(_sg_pick_cmp["rows"], _cg_mapping)
                            cg_o_d, cg_n_d = (None, None)
                            if cutit_d:
                                cg_o_d, cg_n_d = pkg_db.cg_price_pair_at_tier(
                                    _cg_prices,
                                    cutit_d,
                                    _sel_pl_finish,
                                    _sel_pl_min_qty,
                                )
                            _det_rows_cmp: list[dict[str, Any]] = []
                            for r in _sg_pick_cmp["rows"]:
                                er = int(r["excel_row"])
                                full = rows_by_er.get(er) or r
                                gmp_d = (full.get("gmp_code") or "").strip()
                                if not gmp_d:
                                    gmp_d = pkg_db.extract_gmp_code(
                                        full.get("name") or "",
                                        full.get("file") or "",
                                    )
                                kn_d = knife_meta.get(er)
                                if kn_d and kn_d.get("width_mm", 0) > 0:
                                    _prop_d = (
                                        kn_d.get("pdf_file") or ""
                                    ).startswith("propagated_from")
                                    ks_d = (
                                        f"{kn_d['width_mm']:.1f}×{kn_d['height_mm']:.1f}"
                                    )
                                    if _prop_d:
                                        ks_d += " (от размера)"
                                elif cutit_d:
                                    ks_d = f"нет SVG · CG {cutit_d}"
                                else:
                                    ks_d = "—"
                                po_d = _parse_packaging_price(full.get("price"))
                                pn_d = _parse_packaging_price(full.get("price_new"))
                                _det_rows_cmp.append({
                                    "Название": (full.get("name") or "")[:52],
                                    "GMP": gmp_d or "—",
                                    "Вид": (full.get("kind") or "")[:22],
                                    "Наша стар.": f"{po_d:.4f}" if po_d else "—",
                                    "Наша нов.": f"{pn_d:.4f}" if pn_d else "—",
                                    "CG стар. €/1000": f"{cg_o_d:.2f}" if cg_o_d else "—",
                                    "CG нов. €/1000": f"{cg_n_d:.2f}" if cg_n_d else "—",
                                    "Нож / CG": ks_d,
                                })
                            st.dataframe(
                                pd.DataFrame(_det_rows_cmp),
                                use_container_width=True,
                                hide_index=True,
                            )

            st.markdown("#### Каталог ножей CG (типография)")
            # Каталог ножей и ценовые ступени
            _cg_cat_rows = []
            for k in _cg_knives:
                pr = [p for p in _cg_prices if p["cutit_no"] == k["cutit_no"]]
                _f_types = sorted(set(p["finish_type"] for p in pr))
                _new_prices = [p["price_per_1000"] for p in pr if p["price_per_1000"]]
                _old_prices = [p["price_old_per_1000"] for p in pr if p.get("price_old_per_1000")]
                _cg_cat_rows.append({
                    "Нож": k["cutit_no"],
                    "Название": k["name"][:50],
                    "Категория": k["category"][:30],
                    "Картон": k["cardboard"][:20],
                    "Лаки": ", ".join(_f_types),
                    "Стар. мин.": f"{min(_old_prices):.1f}" if _old_prices else "—",
                    "Стар. макс.": f"{max(_old_prices):.1f}" if _old_prices else "—",
                    "Нов. мин.": f"{min(_new_prices):.1f}" if _new_prices else "—",
                    "Нов. макс.": f"{max(_new_prices):.1f}" if _new_prices else "—",
                })
            st.dataframe(pd.DataFrame(_cg_cat_rows), use_container_width=True, hide_index=True, height=250)

            # Ручное сопоставление
            st.markdown("**Сопоставление ножей CG → наши продукты**")
            _mapped_ers = set(_cg_mapping.keys())
            _unmapped_box = [r for r in box_rows if int(r["excel_row"]) not in _mapped_ers]
            _cutit_options = [""] + [k["cutit_no"] for k in _cg_knives]
            _cutit_names = {k["cutit_no"]: k["name"][:40] for k in _cg_knives}

            _manual_changes: list[dict[str, Any]] = []
            if _unmapped_box:
                st.caption(f"{len(_unmapped_box)} продуктов без сопоставления с CG-ножом:")
                for ub in _unmapped_box[:20]:
                    ub_er = int(ub["excel_row"])
                    ub_full = rows_by_er.get(ub_er) or ub
                    ub_name = (ub_full.get("name") or "")[:50]
                    sel = st.selectbox(
                        f"{ub_name}",
                        options=_cutit_options,
                        format_func=lambda c: f"{c} — {_cutit_names.get(c, '')}" if c else "— не привязан —",
                        key=f"pl_cg_map_{ub_er}",
                    )
                    if sel:
                        _manual_changes.append({"excel_row": ub_er, "cutit_no": sel, "confirmed": 1})

            if _manual_changes and st.button(
                f"Сохранить {len(_manual_changes)} сопоставлений",
                key="pl_cg_save_map",
            ):
                if _planner_conn is not None:
                    pkg_db.upsert_cg_mapping(_planner_conn, _manual_changes)
                    _cg_mapping = pkg_db.load_cg_mapping(_planner_conn)
                    st.success(f"Сохранено {len(_manual_changes)} сопоставлений.")
                    st.rerun()

            # Список существующих маппингов
            if _cg_mapping:
                _map_rows = []
                for mer, mv in sorted(_cg_mapping.items()):
                    _mi = rows_by_er.get(mer) or {}
                    _map_rows.append({
                        "Excel Row": mer,
                        "Продукт": (_mi.get("name") or "")[:40],
                        "Нож CG": mv["cutit_no"],
                        "Статус": "✓ ручное" if mv["confirmed"] else "авто",
                    })
                st.dataframe(pd.DataFrame(_map_rows), use_container_width=True, hide_index=True, height=200)

            if st.button("Очистить данные CG", key="pl_cg_clear"):
                if _planner_conn is not None:
                    pkg_db.clear_cg_data(_planner_conn)
                    _cg_knives = []
                    _cg_prices = []
                    _cg_mapping = {}
                    st.success("Данные прайса типографии очищены.")
                    st.rerun()

    with st.expander("Ножи и штанцформы", expanded=False):
        st.caption(
            "Наши ножи (SVG из PDF в кэше) и номера ножей CG (Cutit). "
            "«Распространить» копирует эталонный SVG на позиции того же размера без ножа. "
            "«Применить размеры» меняет только width/height в БД для записей, где нож уже есть."
        )
        if _planner_conn is not None and size_groups:
            if st.button("Распространить ножи по всем размерам", key="pl_knife_prop_all"):
                _n_prop = pkg_db.propagate_knives_in_size_groups(
                    _planner_conn, size_groups, knife_meta
                )
                st.success(f"Заполнено новых позиций: {_n_prop}")
                st.rerun()

            _kf_rows: list[dict[str, Any]] = []
            for sg in size_groups:
                sk_k = sg["size_key"]
                sk_disp_k = pp.size_key_display(sk_k)
                ers_k = [int(r["excel_row"]) for r in sg["rows"]]
                n_tot_k = len(ers_k)
                n_k_k = sum(
                    1
                    for er in ers_k
                    if er in knife_meta and knife_meta[er].get("width_mm", 0) > 0
                )
                dim_s_k, ref_w_k, ref_h_k = _group_knife_dims_display(knife_meta, ers_k)
                cutit_k = _group_cutit_no(sg["rows"], _cg_mapping)
                _kf_rows.append({
                    "size_key": sk_k,
                    "Размер": sk_disp_k,
                    "Видов": n_tot_k,
                    "Наш нож (мм)": dim_s_k,
                    "Нож CG": cutit_k or "—",
                    "Покрытие": f"{n_k_k}/{n_tot_k}",
                })
            st.dataframe(
                pd.DataFrame([{k: v for k, v in r.items() if k != "size_key"} for r in _kf_rows]),
                use_container_width=True,
                hide_index=True,
                height=320,
            )

            _edit_labels_k = [r["Размер"] for r in _kf_rows]
            _edit_pick_k = st.selectbox(
                "Группа: распространение / размеры / удаление",
                ["—"] + _edit_labels_k,
                key="pl_knife_edit_pick",
            )
            if _edit_pick_k != "—":
                _esk = next(r["size_key"] for r in _kf_rows if r["Размер"] == _edit_pick_k)
                _esg = next(g for g in size_groups if g["size_key"] == _esk)
                _eers = [int(r["excel_row"]) for r in _esg["rows"]]
                _, ew_d, eh_d = _group_knife_dims_display(knife_meta, _eers)
                if ew_d is None or eh_d is None:
                    ew_d, eh_d = 100.0, 100.0
                _kh = _hashlib_pl.md5(_esk.encode()).hexdigest()[:12]
                _w_in = st.number_input(
                    "Ширина ножа, мм",
                    min_value=1.0,
                    max_value=2000.0,
                    value=float(ew_d),
                    step=0.5,
                    key=f"pl_kw_{_kh}",
                )
                _h_in = st.number_input(
                    "Высота ножа, мм",
                    min_value=1.0,
                    max_value=2000.0,
                    value=float(eh_d),
                    step=0.5,
                    key=f"pl_kh_{_kh}",
                )
                kb1, kb2, kb3 = st.columns(3)
                with kb1:
                    if st.button("Применить размеры к ножам группы", key=f"pl_kapply_{_kh}"):
                        _nu = pkg_db.update_knife_dimensions(
                            _planner_conn, _eers, _w_in, _h_in
                        )
                        st.success(f"Обновлено записей в кэше: {_nu}")
                        st.rerun()
                with kb2:
                    if st.button("Удалить ножи группы из кэша", key=f"pl_kdel_{_kh}"):
                        _nd = pkg_db.delete_knives_for_rows(_planner_conn, _eers)
                        st.success(f"Удалено записей: {_nd}")
                        st.rerun()
                with kb3:
                    if st.button("Распространить в этой группе", key=f"pl_kprop1_{_kh}"):
                        _np1 = pkg_db.propagate_knives_in_size_groups(
                            _planner_conn,
                            size_groups,
                            knife_meta,
                            size_key_filter=_esk,
                        )
                        st.success(f"Заполнено позиций: {_np1}")
                        st.rerun()
        elif not size_groups:
            st.info("Нет групп размеров — загрузите данные в базу.")

    _pl_editable_section_title("Параметры печатного листа")
    c1, c2, c3, c4, c5 = st.columns(5)
    with c1:
        sheet_w = st.number_input(
            "Ширина, мм",
            min_value=50.0,
            max_value=2000.0,
            value=700.0,
            step=1.0,
            key="pl_sheet_w",
            help="Редактируемо · формат печатного листа",
        )
    with c2:
        sheet_h = st.number_input(
            "Высота, мм",
            min_value=50.0,
            max_value=2000.0,
            value=1000.0,
            step=1.0,
            key="pl_sheet_h",
            help="Редактируемо · формат печатного листа",
        )
    with c3:
        margin_mm = st.number_input(
            "Поле, мм",
            min_value=0.0,
            max_value=50.0,
            value=5.0,
            step=0.5,
            key="pl_margin",
            help="Редактируемо · отступ от края листа",
        )
    with c4:
        gap_mm = st.number_input(
            "Зазор X, мм",
            min_value=-100.0,
            max_value=40.0,
            value=2.0,
            step=0.5,
            key="pl_gap_x",
            help="Редактируемо · зазор между оттисками по X",
        )
    with c5:
        gap_y_mm = st.number_input(
            "Зазор Y, мм",
            min_value=-100.0,
            max_value=40.0,
            value=2.0,
            step=0.5,
            key="pl_gap_y",
            help="Редактируемо · зазор между оттисками по Y",
        )
    sheet_params = pp.SheetParams(
        width_mm=float(sheet_w),
        height_mm=float(sheet_h),
        margin_mm=float(margin_mm),
        gap_mm=float(gap_mm),
        gap_y_mm=float(gap_y_mm),
    )

    st.divider()
    with st.expander("Тарифы печати (ступенчатые и доплаты за отделку)", expanded=False):
        _pl_editable_section_title("Ступени: тираж листов → цена за лист")
        _tariffs_db: list[dict[str, Any]] = []
        if _planner_conn is not None:
            try:
                _tariffs_db = pkg_db.load_tariffs(_planner_conn)
            except Exception:
                pass
        if not _tariffs_db:
            _tariffs_db = [
                {"min_sheets": 1, "max_sheets": 100, "price_per_sheet": 5.0},
                {"min_sheets": 101, "max_sheets": 500, "price_per_sheet": 3.5},
                {"min_sheets": 501, "max_sheets": None, "price_per_sheet": 2.0},
            ]
        _te_rows = [
            {
                "От (листов)": int(t["min_sheets"]),
                "До (листов)": int(t["max_sheets"]) if t.get("max_sheets") is not None else 999999,
                "Цена за лист": float(t["price_per_sheet"]),
            }
            for t in _tariffs_db
        ]
        _te_df = pd.DataFrame(_te_rows)
        _te_edited = st.data_editor(
            _te_df,
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True,
            key="pl_tariff_editor",
            column_config={
                "От (листов)": st.column_config.NumberColumn(
                    "От (листов)",
                    min_value=1,
                    step=1,
                    format="%d",
                    help="Редактируемо · нижняя граница тиража (листов)",
                ),
                "До (листов)": st.column_config.NumberColumn(
                    "До (листов)",
                    min_value=1,
                    step=1,
                    format="%d",
                    help="Редактируемо · 999999 = без верхней границы",
                ),
                "Цена за лист": st.column_config.NumberColumn(
                    "Цена за лист",
                    min_value=0.0,
                    step=0.1,
                    format="%.2f",
                    help="Редактируемо · базовая цена листа на этой ступени",
                ),
            },
        )
        if st.button("Сохранить ступени", key="pl_save_tariffs"):
            _new_tariffs: list[dict[str, Any]] = []
            for _, r in _te_edited.iterrows():
                mn = int(r["От (листов)"])
                mx_v = int(r["До (листов)"])
                mx = None if mx_v >= 999999 else mx_v
                _new_tariffs.append({"min_sheets": mn, "max_sheets": mx, "price_per_sheet": float(r["Цена за лист"])})
            _new_tariffs.sort(key=lambda x: x["min_sheets"])
            if _planner_conn is not None:
                try:
                    pkg_db.save_tariffs(_planner_conn, _new_tariffs)
                    st.success(f"Ступени сохранены ({len(_new_tariffs)}).")
                except Exception as e:
                    st.error(f"Ошибка сохранения тарифов: {e}")
            else:
                st.warning("Файл БД не найден — тарифы не сохранены.")

        st.divider()
        _pl_editable_section_title("Доплаты за отделку (на 1 печатный лист)")
        _xfe_source: list[dict[str, Any]] = []
        if _planner_conn is not None:
            try:
                _xfe_source = pkg_db.load_print_finish_extras(_planner_conn)
            except Exception:
                _xfe_source = []
        if not _xfe_source:
            _xfe_source = [dict(x) for x in pkg_db.DEFAULT_PRINT_FINISH_EXTRAS]
        _xfe_df = pd.DataFrame(
            [
                {
                    "Код": r["code"],
                    "Опция": r["label"],
                    "Доплата за лист": float(r["extra_per_sheet"]),
                }
                for r in _xfe_source
            ]
        )
        _xfe_edited = st.data_editor(
            _xfe_df,
            num_rows="fixed",
            use_container_width=True,
            hide_index=True,
            key="pl_finish_extras_editor",
            column_config={
                "Код": st.column_config.TextColumn("Код", disabled=True, width="small"),
                "Опция": st.column_config.TextColumn(
                    "Опция",
                    help="Редактируемо · подпись в отчётах",
                ),
                "Доплата за лист": st.column_config.NumberColumn(
                    "Доплата / лист",
                    min_value=0.0,
                    step=0.05,
                    format="%.2f",
                    help="Редактируемо · добавляется к каждому отпечатанному листу при включённой галочке",
                ),
            },
        )
        if st.button("Сохранить доплаты за отделку", key="pl_save_finish_extras"):
            if _planner_conn is None:
                st.warning("Файл БД не найден — доплаты не сохранены.")
            else:
                try:
                    _xfe_out: list[dict[str, Any]] = []
                    for _, r in _xfe_edited.iterrows():
                        _xfe_out.append({
                            "code": str(r["Код"]),
                            "label": str(r["Опция"]),
                            "extra_per_sheet": float(r["Доплата за лист"]),
                        })
                    pkg_db.save_print_finish_extras(_planner_conn, _xfe_out)
                    st.success(f"Доплаты сохранены ({len(_xfe_out)} опций).")
                except Exception as e:
                    st.error(f"Ошибка сохранения доплат: {e}")

    _tariffs_for_plan: list[dict[str, Any]] = []
    if _planner_conn is not None:
        try:
            _tariffs_for_plan = pkg_db.load_tariffs(_planner_conn)
        except Exception:
            pass

    _extras_for_plan: list[dict[str, Any]] = [dict(x) for x in pkg_db.DEFAULT_PRINT_FINISH_EXTRAS]
    if _planner_conn is not None:
        try:
            _extras_for_plan = pkg_db.load_print_finish_extras(_planner_conn)
        except Exception:
            pass

    st.divider()
    _pl_editable_section_title("Учитывать доплаты в расчёте стоимости")
    st.caption(
        "Отметьте лак UV, лак WB, фольгу, Pantone +1 — суммы из таблицы «Доплата / лист» умножаются на число листов "
        "и показываются в анализе цен ниже (сборный и раздельный тираж, CG и ступени)."
    )
    _xcols = st.columns(4)
    _sorted_extras = sorted(_extras_for_plan, key=lambda x: str(x.get("code") or ""))
    _xi = 0
    for row in _sorted_extras:
        _ck = _PL_PRINT_FIN_CHECK_KEYS.get(str(row.get("code") or ""))
        if _ck is None:
            continue
        _amt = float(row.get("extra_per_sheet") or 0.0)
        _lab = str(row.get("label") or row.get("code") or "")
        with _xcols[_xi % 4]:
            st.checkbox(
                f"{_lab} (+{_amt:.2f}/лист)",
                key=_ck,
                help="Включает эту доплату ко всем расчётам листов в этом размере.",
            )
        _xi += 1

    # ── Выбор размера для оптимизации ──
    st.subheader("Оптимизация печати по размеру")

    def _close_planner():
        if _planner_conn is not None:
            try:
                _planner_conn.close()
            except Exception:
                pass

    if not size_groups:
        if not box_rows:
            st.info("В базе нет печатной продукции или данные не загружены.")
        elif _pl_any_kind_filter:
            st.info(
                "Нет групп размеров для текущего фильтра вида — см. предупреждение выше или снимите галочки."
            )
        else:
            st.info("Нет продукции с заполненным «Размер (мм)». Заполните данные на вкладке «Макеты».")
        _close_planner()
        return

    from collections import defaultdict
    _monthly_by_er: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for m in monthly_db:
        _monthly_by_er[int(m["excel_row"])].append(m)

    def _parse_qty_opt(val: str) -> float:
        if not val:
            return 0.0
        cleaned = val.replace(" ", "").replace("\u00a0", "").replace(",", ".")
        try:
            return float(cleaned)
        except (ValueError, TypeError):
            return 0.0

    sg_options: list[str] = []
    sg_map: dict[str, dict[str, Any]] = {}
    for sg in size_groups:
        sk = sg["size_key"]
        disp = pp.size_key_display(sk)
        n_items = len(sg["rows"])
        total_annual = sum(_parse_qty_opt((rows_by_er.get(int(r["excel_row"])) or r).get("qty_per_year") or "") for r in sg["rows"])
        label = f"{disp}  —  {n_items} видов"
        if total_annual > 0:
            label += f",  ~{int(total_annual)} шт/год"
        sg_options.append(label)
        sg_map[label] = sg

    _sel_sg_label = st.selectbox(
        "Выберите размер (габариты)",
        options=sg_options,
        key="pl_sel_size_group",
    )
    if not _sel_sg_label:
        _close_planner()
        return
    _sel_sg = sg_map[_sel_sg_label]
    _sel_sk = _sel_sg["size_key"]
    _sel_sk_disp = pp.size_key_display(_sel_sk)
    _sel_items = _sel_sg["rows"]

    st.markdown(f"#### Размер {_sel_sk_disp}: {len(_sel_items)} видов продукции")

    if db_path.is_file():
        import importlib

        import packaging_profile_excel as pprof

        # Streamlit не перезагружает импортированные модули при правках .py — иначе остаётся старая сборка Excel.
        importlib.reload(pprof)

        if "pl_profile_doc_code" not in st.session_state:
            st.session_state["pl_profile_doc_code"] = "OM/ПУМ-192-01-373"
        if "pl_profile_report_year" not in st.session_state:
            st.session_state["pl_profile_report_year"] = int(date.today().year)
        st.markdown("##### Экспорт Excel «профиль» этой группы")
        st.caption(
            "Шапка Balkan, **широкая таблица до столбца Z** (№, GMP, вид, размер, нож CG, кол-во на листе из макетов, €/1000 CG, "
            "подрядчики 1–5, годовой объём, Янв–Дек), затем анализ и график. "
            "Файл **packaging_profile_….xlsx** только для **этой размерной группы**. **Вся база** — кнопка «Профиль Excel» в шапке рядом с «Скачать Excel»; исходный макет — по «Скачать Excel»."
        )
        _pr_c1, _pr_c2, _pr_c3 = st.columns([2, 2, 1])
        with _pr_c1:
            st.text_input(
                "Codul documentului (шапка Excel)",
                key="pl_profile_doc_code",
                help="Серый блок справа в шапке, как на фирменном бланке.",
            )
        _pl_prof_fin_labels = {
            "lac_wb": "Lac WB (водный лак)",
            "uv_no_foil": "UV без фольги",
            "uv_foil": "UV с фольгой",
        }
        with _pr_c2:
            st.selectbox(
                "Отделка CG для столбца €/1000",
                options=list(_pl_prof_fin_labels.keys()),
                format_func=lambda k: _pl_prof_fin_labels[k],
                key="pl_profile_finish_pick",
            )
        with _pr_c3:
            st.number_input(
                "Год (колонки месяцев)",
                min_value=2000,
                max_value=2100,
                step=1,
                key="pl_profile_report_year",
                help="Помесячные объёмы из БД за этот год в столбцах Янв–Дек.",
            )
        try:
            _finish_profile = str(st.session_state.get("pl_profile_finish_pick") or "lac_wb")
            _prof_bytes = pprof.build_profile_workbook_bytes(
                db_path=db_path,
                size_key=_sel_sk,
                size_key_display_override=_sel_sk_disp,
                group_rows=list(_sel_items),
                rows_by_er=rows_by_er,
                sheet_params=sheet_params,
                document_code=str(st.session_state.get("pl_profile_doc_code") or ""),
                finish_code=_finish_profile,
                logo_path=None,
                report_year=int(st.session_state.get("pl_profile_report_year") or date.today().year),
            )
            _safe_fn = re.sub(r"[^\w\-.]+", "_", _sel_sk)[:80]
            st.download_button(
                label="Скачать Excel профиль (эта группа)",
                data=_prof_bytes,
                file_name=f"packaging_profile_{_safe_fn}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="pl_download_profile_xlsx",
            )
        except Exception as _e_prof:
            st.warning(f"Не удалось сформировать профиль Excel: {_e_prof}")

    # ── Извлечение ножей из PDF (с кэшированием в session_state) ──
    _sel_ers = [int(r["excel_row"]) for r in _sel_items]
    # v2: более мягкие допуски цвета/толщины + второй проход извлечения (инвалидирует старый кэш)
    _knife_ss_key = f"_pl_knives_{_sel_sk}_v2"
    _knives_found: dict[int, dict[str, Any]] = {}
    _knives_missing_pdf: list[int] = []
    _knives_no_outline: list[int] = []

    _cached_result = st.session_state.get(_knife_ss_key)
    if isinstance(_cached_result, dict) and set(_cached_result.get("ers", [])) == set(_sel_ers):
        _knives_found = _cached_result.get("found", {})
        _knives_missing_pdf = _cached_result.get("missing_pdf", [])
        _knives_no_outline = _cached_result.get("no_outline", [])
    else:
        _new_knives_to_save: list[dict[str, Any]] = []
        for r in _sel_items:
            er = int(r["excel_row"])
            full = rows_by_er.get(er) or r
            meta = knife_meta.get(er)
            if meta and meta["width_mm"] > 0:
                _knives_found[er] = meta
                continue

            pdf_file_val = (full.get("file") or "").strip()
            if not pdf_file_val:
                _knives_missing_pdf.append(er)
                continue

            pdf_path = resolve_pdf_path(pdf_dir, pdf_file_val) if pdf_dir else None
            if pdf_path is None or not pdf_path.is_file():
                _knives_missing_pdf.append(er)
                continue

            result = posv.try_extract_knife_from_pdf(str(pdf_path))
            if result is None:
                _knives_no_outline.append(er)
                continue

            svg_full, w_mm, h_mm = result
            knife_entry = {
                "excel_row": er,
                "svg_full": svg_full,
                "width_mm": w_mm,
                "height_mm": h_mm,
                "pdf_file": pdf_file_val,
            }
            _knives_found[er] = knife_entry
            _new_knives_to_save.append(knife_entry)

        if _new_knives_to_save and _planner_conn is not None:
            try:
                pkg_db.save_knives_batch(_planner_conn, _new_knives_to_save)
            except Exception:
                pass

        st.session_state[_knife_ss_key] = {
            "ers": _sel_ers,
            "found": _knives_found,
            "missing_pdf": _knives_missing_pdf,
            "no_outline": _knives_no_outline,
        }

    n_with_knife = len(_knives_found)
    n_total_items = len(_sel_items)
    if n_with_knife > 0:
        sample_knife = next(iter(_knives_found.values()))
        knife_w = sample_knife["width_mm"]
        knife_h = sample_knife["height_mm"]
        st.caption(
            f"Нож из PDF: **{knife_w:.1f} × {knife_h:.1f} мм** "
            f"(найден у {n_with_knife} из {n_total_items} видов)"
        )
    else:
        knife_w = knife_h = 0.0

    _item_detail: list[dict[str, Any]] = []
    for r in _sel_items:
        er = int(r["excel_row"])
        full = rows_by_er.get(er) or r
        annual = _parse_qty_opt(full.get("qty_per_year") or "")
        per_sheet = _parse_qty_opt(full.get("qty_per_sheet") or "")
        er_monthly = _monthly_by_er.get(er, [])
        monthly_total = sum(m["qty"] for m in er_monthly)
        kn = _knives_found.get(er)
        if kn:
            _is_propagated = (kn.get("pdf_file") or "").startswith("propagated_from")
            dims = f"{kn['width_mm']:.1f}×{kn['height_mm']:.1f}"
            knife_status = f"{dims} (от размера)" if _is_propagated else dims
        elif er in _knives_missing_pdf:
            knife_status = "нет PDF"
        elif er in _knives_no_outline:
            knife_status = "нет контура"
        else:
            knife_status = "—"
        gmp = (full.get("gmp_code") or "").strip()
        if not gmp:
            gmp = pkg_db.extract_gmp_code(full.get("name") or "", full.get("file") or "")
        stock_qty = _stock_db.get(gmp, 0) if gmp else 0
        _item_detail.append({
            "er": er,
            "GMP-код": gmp or "—",
            "Название": (full.get("name") or "")[:55],
            "Вид": (full.get("kind") or "")[:25],
            "Нож (мм)": knife_status,
            "Год. объём": int(annual) if annual else "—",
            "На листе": int(per_sheet) if per_sheet else "—",
            "Склад": int(stock_qty) if stock_qty > 0 else "—",
            "Помесячно": int(monthly_total) if monthly_total else "—",
        })
    _df_kinds = pd.DataFrame(_item_detail)
    _kind_pick_key = f"pl_kind_pick_{_sel_sk.replace('|', '_').replace('/', '_')[:72]}"
    _n_kinds_rows = len(_item_detail)
    st.caption(
        "**Какие виды включить в расчёт печати:** выделите строки в таблице (клик; **Ctrl/⌘** или **Shift** — несколько). "
        "Учитываются только выбранные позиции (потребность, слоты, стоимость, переход в «Печать»)."
    )
    _sel_default = None
    if _n_kinds_rows > 0:
        _sel_default = {
            "selection": {
                "rows": list(range(_n_kinds_rows)),
                "columns": [],
                "cells": [],
            },
        }
    _vis_cols = [c for c in _df_kinds.columns if c != "er"]
    st.dataframe(
        _df_kinds,
        use_container_width=True,
        hide_index=True,
        column_order=_vis_cols if _vis_cols else None,
        key=_kind_pick_key,
        on_select="rerun",
        selection_mode="multi-row",
        selection_default=_sel_default,
    )
    _pick_raw = st.session_state.get(_kind_pick_key)
    _picked_rows: list[int] = []
    if isinstance(_pick_raw, dict):
        _sel_d = _pick_raw.get("selection") or {}
        if isinstance(_sel_d, dict):
            _picked_rows = [int(x) for x in (_sel_d.get("rows") or [])]
    elif _pick_raw is not None and hasattr(_pick_raw, "selection"):
        _ps = getattr(_pick_raw, "selection", None)
        if _ps is not None:
            _picked_rows = [int(x) for x in (getattr(_ps, "rows", None) or [])]
    if _n_kinds_rows > 0 and not _picked_rows:
        st.warning(
            "Не выбран ни один вид — отметьте строки в таблице выше или снимите лишнее выделение так, "
            "чтобы осталась хотя бы одна строка."
        )
        _close_planner()
        return
    _selected_ers: set[int] = (
        {
            int(_item_detail[i]["er"])
            for i in _picked_rows
            if 0 <= i < len(_item_detail)
        }
        if _picked_rows
        else set()
    )
    _active_items = [r for r in _sel_items if int(r["excel_row"]) in _selected_ers]
    if not _active_items and _sel_items:
        st.warning("Внутренняя ошибка отбора видов — обновите страницу.")
        _close_planner()
        return

    if knife_w > 0 and knife_h > 0:
        n_fit, _pl_preview, fill_pct = pp.pack_shelf_single_item(sheet_params, knife_w, knife_h)
        st.caption(
            f"По ножу из PDF: на листе {sheet_params.width_mm:g}×{sheet_params.height_mm:g} мм помещается "
            f"**{n_fit}** шт. ({knife_w:.1f}×{knife_h:.1f} мм), заполнение **{fill_pct:.1f}%**"
        )
    else:
        fp_r = pp.footprint_mm_from_size(_sel_sg["sample_size_str"])
        if fp_r:
            knife_w, knife_h = fp_r
            n_fit, _pl_preview, fill_pct = pp.pack_shelf_single_item(sheet_params, fp_r[0], fp_r[1])
            st.caption(
                f"Нож не найден — по габаритам из текста: на листе помещается "
                f"**{n_fit}** шт. ({fp_r[0]:g}×{fp_r[1]:g} мм), заполнение **{fill_pct:.1f}%**"
            )
        else:
            n_fit = 0
            _pl_preview = []
            st.warning("Не удалось определить размер оттиска ни из PDF, ни из строки размера.")

    # ── Оптимизация: распределение слотов по историческим объёмам ──
    st.divider()
    _pl_editable_section_title("Оптимизация раскладки (период, тип лакирования CG)")

    # Определим cutit_no для текущего размера через cg_mapping
    _sel_cutit: str | None = None
    _sel_cutit_prices: list[dict[str, Any]] = []
    for r in _active_items:
        er = int(r["excel_row"])
        cm = _cg_mapping.get(er)
        if cm:
            _sel_cutit = cm["cutit_no"]
            break
    if _sel_cutit:
        _sel_cutit_prices = [p for p in _cg_prices if p["cutit_no"] == _sel_cutit]

    _FINISH_LABELS = {
        "lac_wb": "Lac WB (водный лак)",
        "uv_no_foil": "UV без фольги",
        "uv_foil": "UV с фольгой",
    }
    _available_finishes = sorted(set(p["finish_type"] for p in _sel_cutit_prices)) if _sel_cutit_prices else []

    pc1, pc2, pc3 = st.columns(3)
    with pc1:
        _period_label = st.selectbox(
            "Период потребности",
            options=["1 месяц", "3 месяца (квартал)", "6 месяцев (полгода)", "12 месяцев (год)"],
            index=3,
            key="pl_demand_period",
        )
    with pc2:
        if _available_finishes:
            _finish_options = [_FINISH_LABELS.get(f, f) for f in _available_finishes]
            _sel_finish_label = st.selectbox(
                "Тип лакирования",
                options=_finish_options,
                key="pl_finish_type",
            )
            _sel_finish = _available_finishes[_finish_options.index(_sel_finish_label)]
        else:
            _sel_finish = ""
            if _cg_knives:
                st.caption("Нож CG не сопоставлен")
            else:
                st.caption("Прайс CG не загружен")
    with pc3:
        if _sel_cutit:
            _cg_knife_info = next((k for k in _cg_knives if k["cutit_no"] == _sel_cutit), None)
            if _cg_knife_info:
                st.caption(f"Нож CG: **{_sel_cutit}** ({_cg_knife_info['name'][:40]})")
            else:
                st.caption(f"Нож CG: **{_sel_cutit}**")

    _period_map = {"1 месяц": 1, "3 месяца (квартал)": 3, "6 месяцев (полгода)": 6, "12 месяцев (год)": 12}
    _period_months = _period_map.get(_period_label, 12)

    annual_demand_by_er: dict[int, float] = {}
    _gmp_by_er: dict[int, str] = {}
    _stock_by_er: dict[int, float] = {}
    for r in _active_items:
        er = int(r["excel_row"])
        full = rows_by_er.get(er) or r
        annual = _parse_qty_opt(full.get("qty_per_year") or "")
        if annual > 0:
            annual_demand_by_er[er] = annual
        else:
            er_monthly = _monthly_by_er.get(er, [])
            monthly_total = sum(m["qty"] for m in er_monthly)
            if monthly_total > 0:
                annual_demand_by_er[er] = monthly_total
        gmp = (full.get("gmp_code") or "").strip()
        if not gmp:
            gmp = pkg_db.extract_gmp_code(full.get("name") or "", full.get("file") or "")
        _gmp_by_er[er] = gmp
        if gmp and gmp in _stock_db:
            _stock_by_er[er] = _stock_db[gmp]

    raw_demand_by_er: dict[int, float] = {}
    demand_by_er: dict[int, float] = {}
    _total_stock_deducted = 0.0
    for er, annual in annual_demand_by_er.items():
        period_demand = annual * _period_months / 12.0
        raw_demand_by_er[er] = period_demand
        stock = _stock_by_er.get(er, 0.0)
        net = max(0.0, period_demand - stock)
        demand_by_er[er] = net
        if stock > 0:
            _total_stock_deducted += min(stock, period_demand)

    _pl_demand_is_synthetic = False
    if not demand_by_er or all(v == 0 for v in demand_by_er.values()):
        if _total_stock_deducted > 0:
            st.success(
                f"Потребность за период полностью покрыта складскими остатками "
                f"({int(_total_stock_deducted):,} шт.). Печать не требуется."
            )
            _close_planner()
            return
        st.warning(
            "Нет данных об объёмах ни для одного вида этого размера. "
            "Заполните «Кол-во в год» на вкладке «Макеты» или загрузите помесячные данные через Cutii."
        )
        if not st.checkbox(
            "Продолжить без объёмов: равномерная потребность по выбранным видам (только для раскладки и макета)",
            key="pl_continue_no_volume",
        ):
            _close_planner()
            return
        _pl_demand_is_synthetic = True
        annual_demand_by_er = {}
        raw_demand_by_er = {}
        demand_by_er = {}
        _total_stock_deducted = 0.0
        for r in _active_items:
            er = int(r["excel_row"])
            annual_demand_by_er[er] = 1.0
        for er, annual in annual_demand_by_er.items():
            period_demand = annual * _period_months / 12.0
            raw_demand_by_er[er] = period_demand
            demand_by_er[er] = period_demand
    if n_fit <= 0:
        st.warning("Не удалось рассчитать количество ячеек на листе для этого размера.")
        _close_planner()
        return

    demand_by_er = {er: d for er, d in demand_by_er.items() if d > 0}
    total_demand = sum(demand_by_er.values())
    total_raw_demand = sum(raw_demand_by_er.get(er, 0) for er in demand_by_er)

    _demand_caption = (
        f"Потребность за **{_period_label}**: **{int(total_demand):,}** шт. · "
        f"Ячеек на листе: **{n_fit}** · "
        f"Видов с данными: **{len(demand_by_er)}** из **{len(_active_items)}** выбранных "
        f"(по размеру всего **{len(_sel_items)}** видов)"
    )
    if _total_stock_deducted > 0:
        _demand_caption += f" · Вычтено со склада: **{int(_total_stock_deducted):,}** шт."
    if _pl_demand_is_synthetic:
        _demand_caption += " · **Условная** потребность (одинаковая по видам, без учёта склада)."
    st.caption(_demand_caption)

    ers_sorted = sorted(demand_by_er.items(), key=lambda x: x[1], reverse=True)
    n_kinds = min(len(ers_sorted), n_fit)
    if n_kinds < len(ers_sorted):
        st.info(f"На листе {n_fit} ячеек, но видов {len(ers_sorted)} — берём {n_kinds} самых востребованных.")
        ers_sorted = ers_sorted[:n_kinds]
        total_demand = sum(d for _, d in ers_sorted)

    raw_slots = [max(1, round(d / total_demand * n_fit)) for _, d in ers_sorted]
    diff = sum(raw_slots) - n_fit
    while diff != 0:
        if diff > 0:
            idx_fix = max(range(n_kinds), key=lambda i: raw_slots[i])
            raw_slots[idx_fix] -= 1
        else:
            idx_fix = max(range(n_kinds), key=lambda i: ers_sorted[i][1] / max(raw_slots[i], 1))
            raw_slots[idx_fix] += 1
        diff = sum(raw_slots) - n_fit

    n_sheets = max(1, math.ceil(max(
        ers_sorted[i][1] / max(raw_slots[i], 1) for i in range(n_kinds)
    )))
    total_printed = n_sheets * n_fit
    empty_slots = n_fit - sum(raw_slots)

    st.markdown(f"**Рекомендация: печатать {n_sheets:,} листов → {total_printed:,} оттисков всего**")

    _opt_rows: list[dict[str, Any]] = []
    for i, (er, dem) in enumerate(ers_sorted):
        actual = raw_slots[i] * n_sheets
        over = actual - dem
        over_pct = (over / max(dem, 1)) * 100
        itm = rows_by_er.get(er) or {}
        ann = annual_demand_by_er.get(er, 0)
        raw_dem = raw_demand_by_er.get(er, dem)
        stock = _stock_by_er.get(er, 0)
        gmp = _gmp_by_er.get(er, "")
        _opt_row: dict[str, Any] = {
            "er": er,
            "GMP": gmp or "—",
            "Название": (itm.get("name") or "")[:50],
            "Год. объём": int(ann) if ann else "—",
            f"Потр. ({_period_label})": int(raw_dem),
        }
        if stock > 0:
            _opt_row["Склад"] = int(stock)
            _opt_row["Чистая потр."] = int(dem)
        else:
            _opt_row["Склад"] = "—"
            _opt_row["Чистая потр."] = int(dem)
        _opt_row["Слотов"] = raw_slots[i]
        _opt_row["Напечатано"] = int(actual)
        _opt_row["Разница"] = int(over)
        _opt_row["Разница %"] = round(over_pct, 1)
        _opt_rows.append(_opt_row)
    if empty_slots > 0:
        _opt_rows.append({
            "er": "—",
            "GMP": "",
            "Название": f"Свободные ячейки ({empty_slots})",
            "Год. объём": "",
            f"Потр. ({_period_label})": 0,
            "Склад": "",
            "Чистая потр.": 0,
            "Слотов": empty_slots,
            "Напечатано": 0,
            "Разница": 0,
            "Разница %": 0.0,
        })
    st.dataframe(pd.DataFrame(_opt_rows), use_container_width=True, hide_index=True)

    # --- Расчёт стоимости (CG-тарифы или старые тарифы) ---
    _use_cg_pricing = bool(_sel_cutit and _sel_finish and _sel_cutit_prices)
    _cg_p1000_old_combined: float | None = None
    _total_cost_old: float = 0.0

    if _use_cg_pricing:
        _total_boxes_combined = total_printed
        _cg_p1000_combined = pkg_db.cg_price_for_qty(_sel_cutit_prices, _sel_finish, _total_boxes_combined)
        _cg_p1000_old_combined = pkg_db.cg_old_price_for_qty(_sel_cutit_prices, _sel_finish, _total_boxes_combined)
        if _cg_p1000_combined is not None:
            total_cost = _cg_p1000_combined * _total_boxes_combined / 1000.0
            price_per_sheet = total_cost / max(n_sheets, 1)
            if _cg_p1000_old_combined is not None:
                _total_cost_old = _cg_p1000_old_combined * _total_boxes_combined / 1000.0
        else:
            price_per_sheet = pkg_db.sheet_price(n_sheets, _tariffs_for_plan)
            total_cost = price_per_sheet * n_sheets
            _use_cg_pricing = False
    else:
        price_per_sheet = pkg_db.sheet_price(n_sheets, _tariffs_for_plan)
        total_cost = price_per_sheet * n_sheets

    _extra_per_sheet, _extra_breakdown = _pl_sum_enabled_print_extras(_extras_for_plan)
    _extra_total_combined = _extra_per_sheet * float(n_sheets)
    total_cost = total_cost + _extra_total_combined
    if _cg_p1000_old_combined is not None:
        _total_cost_old = _total_cost_old + _extra_total_combined
    price_per_sheet = total_cost / max(n_sheets, 1)

    # --- Расчёт раздельной печати (каждый вид отдельно) ---
    _sep_details: list[dict[str, Any]] = []
    sep_sheets = 0
    sep_cost = 0.0
    for i, (er, dem) in enumerate(ers_sorted):
        sh_sep = max(1, math.ceil(dem / n_fit))
        printed_sep = sh_sep * n_fit
        if _use_cg_pricing:
            _cg_p_sep = pkg_db.cg_price_for_qty(_sel_cutit_prices, _sel_finish, printed_sep)
            if _cg_p_sep is not None:
                cost_sep = _cg_p_sep * printed_sep / 1000.0
                price_sep = cost_sep / max(sh_sep, 1)
            else:
                price_sep = pkg_db.sheet_price(sh_sep, _tariffs_for_plan)
                cost_sep = price_sep * sh_sep
        else:
            price_sep = pkg_db.sheet_price(sh_sep, _tariffs_for_plan)
            cost_sep = price_sep * sh_sep
        _extra_sep = _extra_per_sheet * float(sh_sep)
        cost_sep = cost_sep + _extra_sep
        price_sep = cost_sep / max(sh_sep, 1)
        slots_combined = raw_slots[i]
        printed_combined = slots_combined * n_sheets
        cost_combined = (slots_combined / max(n_fit, 1)) * total_cost
        saving_item = cost_sep - cost_combined
        itm = rows_by_er.get(er) or {}
        _sep_details.append({
            "er": er,
            "name": (itm.get("name") or "")[:50],
            "demand": dem,
            "sep_sheets": sh_sep,
            "sep_price": price_sep,
            "sep_cost": cost_sep,
            "sep_printed": printed_sep,
            "comb_slots": slots_combined,
            "comb_printed": printed_combined,
            "comb_cost": cost_combined,
            "saving": saving_item,
        })
        sep_sheets += sh_sep
        sep_cost += cost_sep

    savings = sep_cost - total_cost
    savings_pct = (savings / max(sep_cost, 0.01)) * 100
    cost_per_box_combined = total_cost / max(total_printed, 1)
    cost_per_box_separate = sep_cost / max(sum(d["sep_printed"] for d in _sep_details), 1)

    _extras_sess_parts: list[str] = []
    for _b in _extra_breakdown:
        if _b.get("enabled") and float(_b.get("per_sheet") or 0) > 0:
            _suf = " €/лист" if _use_cg_pricing else "/лист"
            _extras_sess_parts.append(
                f"{_b.get('label', '')} (+{float(_b.get('per_sheet') or 0):.2f}{_suf})"
            )
    st.session_state["pl_last_print_economics"] = {
        "size_key": _sel_sk,
        "size_disp": _sel_sk_disp,
        "period": _period_label,
        "n_sheets": int(n_sheets),
        "sep_sheets": int(sep_sheets),
        "total_cost": float(total_cost),
        "sep_cost": float(sep_cost),
        "savings": float(savings),
        "savings_pct": float(savings_pct),
        "cost_per_imprint_c": float(cost_per_box_combined),
        "cost_per_imprint_s": float(cost_per_box_separate),
        "use_cg": bool(_use_cg_pricing),
        "extras_note": "; ".join(_extras_sess_parts),
        "finish_label": _FINISH_LABELS.get(_sel_finish, _sel_finish or "")
        if _use_cg_pricing
        else "",
    }

    # --- Метрики сборного тиража ---
    _pricing_label = (
        f"Сборный тираж (CG: {_FINISH_LABELS.get(_sel_finish, _sel_finish)})"
        if _use_cg_pricing else "Сборный тираж"
    )
    st.markdown(f"#### {_pricing_label}")
    _cost_unit = " €" if _use_cg_pricing else ""

    if _use_cg_pricing and _cg_p1000_old_combined and _cg_p1000_old_combined != _cg_p1000_combined:
        mc1, mc2, mc3, mc4, mc5 = st.columns(5)
        with mc1:
            st.metric("Листов", f"{n_sheets:,}")
        with mc2:
            _delta_p = _cg_p1000_combined - _cg_p1000_old_combined
            st.metric(
                "Новая цена/1000 шт.",
                f"{_cg_p1000_combined:.2f} €",
                delta=f"{_delta_p:+.2f} €" if abs(_delta_p) > 0.01 else None,
                delta_color="inverse",
            )
        with mc3:
            st.metric("Старая цена/1000 шт.", f"{_cg_p1000_old_combined:.2f} €")
        with mc4:
            st.metric("Общая стоимость (нов.)", f"{total_cost:,.2f} €")
        with mc5:
            st.metric(
                "Общая стоимость (стар.)",
                f"{_total_cost_old:,.2f} €",
                delta=f"{total_cost - _total_cost_old:+,.2f} €",
                delta_color="inverse",
            )
    else:
        mc1, mc2, mc3, mc4 = st.columns(4)
        with mc1:
            st.metric("Листов", f"{n_sheets:,}")
        with mc2:
            if _use_cg_pricing:
                st.metric("Цена за 1000 шт.", f"{_cg_p1000_combined:.2f} €")
            else:
                st.metric("Цена за лист", f"{price_per_sheet:.2f}")
        with mc3:
            st.metric("Общая стоимость", f"{total_cost:,.2f}{_cost_unit}")
        with mc4:
            st.metric("Цена за оттиск", f"{cost_per_box_combined:.4f}{_cost_unit}")

    with st.expander(
        "Анализ цен: доплаты за отделку (лак UV / WB, фольга, Pantone +1)",
        expanded=bool(_extra_total_combined > 0),
    ):
        _eab_rows: list[dict[str, Any]] = []
        for b in _extra_breakdown:
            _en = bool(b.get("enabled"))
            _ps = float(b.get("per_sheet") or 0.0)
            _row_tot = (_ps * float(n_sheets)) if _en else 0.0
            _eab_rows.append({
                "Опция": b.get("label") or b.get("code"),
                "За 1 лист": round(_ps, 2),
                "В расчёте": "да" if _en else "нет",
                f"Сумма ({n_sheets} л.)": round(_row_tot, 2),
            })
        st.dataframe(pd.DataFrame(_eab_rows), use_container_width=True, hide_index=True)
        st.caption(
            f"Дополнительно к базе (CG или ступени): **{_extra_per_sheet:.2f}** за лист · "
            f"всего по сборному тиражу **{_extra_total_combined:,.2f}**{_cost_unit}."
        )

    # --- Сравнение: сборный vs раздельный ---
    st.markdown("#### Экономия сборного тиража vs раздельная печать")
    sc1, sc2, sc3, sc4 = st.columns(4)
    with sc1:
        st.metric(
            "Раздельно: листов",
            f"{sep_sheets:,}",
            delta=f"{n_sheets - sep_sheets:,}",
            delta_color="inverse",
        )
    with sc2:
        st.metric(
            "Раздельно: стоимость",
            f"{sep_cost:,.2f}{_cost_unit}",
        )
    with sc3:
        st.metric(
            "Экономия",
            f"{savings:,.2f}{_cost_unit}",
            delta=f"{savings_pct:+.1f}%",
            delta_color="normal",
        )
    with sc4:
        st.metric(
            "Цена за оттиск (разд.)",
            f"{cost_per_box_separate:.4f}{_cost_unit}",
            delta=f"{cost_per_box_combined - cost_per_box_separate:.4f}",
            delta_color="inverse",
        )

    # --- Детализация по каждому виду ---
    with st.expander("Детализация экономии по каждому виду", expanded=False):
        _econ_rows: list[dict[str, Any]] = []
        for d in _sep_details:
            saving_pct_item = (d["saving"] / max(d["sep_cost"], 0.01)) * 100
            _econ_rows.append({
                "Название": d["name"],
                f"Потр. ({_period_label})": int(d["demand"]),
                "Раздельно: листов": d["sep_sheets"],
                "Раздельно: цена/лист": round(d["sep_price"], 2),
                "Раздельно: стоимость": round(d["sep_cost"], 2),
                "Сборный: слотов": d["comb_slots"],
                "Сборный: стоимость": round(d["comb_cost"], 2),
                "Экономия": round(d["saving"], 2),
                "Экономия %": round(saving_pct_item, 1),
            })
        st.dataframe(pd.DataFrame(_econ_rows), use_container_width=True, hide_index=True)

    # SVG-превью раскладки
    _auto_slot_er: list[int | None] = []
    for i, (er, _) in enumerate(ers_sorted):
        for _ in range(raw_slots[i]):
            _auto_slot_er.append(er)
    for _ in range(empty_slots):
        _auto_slot_er.append(None)

    _auto_labels: list[str] = []
    for _se in _auto_slot_er:
        if _se is not None:
            _itm = rows_by_er.get(int(_se))
            k = (_itm.get("kind") or "").strip() if _itm else ""
            nm = (_itm.get("name") or "").strip() if _itm else ""
            if k and nm:
                _auto_labels.append(f"{k} — {nm}"[:48])
            elif nm:
                _auto_labels.append(nm[:48])
            elif k:
                _auto_labels.append(k[:48])
            else:
                _auto_labels.append(str(_se))
        else:
            _auto_labels.append("")

    if _pl_preview:
        _svg_auto = pp.sheet_layout_svg(
            sheet_params,
            _pl_preview,
            title=f"{_sel_sk_disp} · {n_sheets:,} листов · авто-раскладка",
            slot_labels=_auto_labels or None,
        )
        _b64a = base64.b64encode(_svg_auto.encode("utf-8")).decode("ascii")
        st.markdown(
            '<img src="data:image/svg+xml;base64,'
            f'{_b64a}" style="width:35%;max-width:35%;height:auto;display:block;" alt="авто"/>',
            unsafe_allow_html=True,
        )

    # ── Ручная корректировка ──
    st.divider()
    _pl_editable_section_title("Ручная корректировка слотов на листе")
    _man_rows: list[dict[str, Any]] = []
    for i, (er, dem) in enumerate(ers_sorted):
        itn = rows_by_er.get(er) or {}
        _man_rows.append({
            "er": er,
            "Название": (itn.get("name") or "")[:50],
            "Потребность": int(dem),
            "Слотов": raw_slots[i],
        })
    _man_df = pd.DataFrame(_man_rows)
    _man_edited = st.data_editor(
        _man_df,
        use_container_width=True,
        hide_index=True,
        key=f"pl_man_edit_{_sel_sk}",
        column_config={
            "er": st.column_config.NumberColumn("er", disabled=True, format="%d"),
            "Название": st.column_config.TextColumn("Название", disabled=True),
            "Потребность": st.column_config.NumberColumn("Потребность", disabled=True, format="%d"),
            "Слотов": st.column_config.NumberColumn(
                "Слотов",
                min_value=0,
                step=1,
                format="%d",
                help="Редактируемо · число ячеек листа под этот вид",
            ),
        },
    )

    if st.button("Пересчитать", key=f"pl_man_recalc_{_sel_sk}"):
        _new_slots: list[int] = []
        for _, _mr in _man_edited.iterrows():
            _new_slots.append(max(0, int(_mr["Слотов"])))
        used = sum(_new_slots)
        if used > n_fit:
            st.error(f"Сумма слотов ({used}) > ячеек на листе ({n_fit}).")
        elif used == 0:
            st.warning("Все слоты = 0.")
        else:
            _n_sh = max(1, math.ceil(max(
                ers_sorted[j][1] / max(_new_slots[j], 1)
                for j in range(n_kinds)
                if _new_slots[j] > 0
            )))
            _recalc_rows: list[dict[str, Any]] = []
            for j, (er, dem) in enumerate(ers_sorted):
                actual = _new_slots[j] * _n_sh
                over = actual - dem
                over_p = (over / max(dem, 1)) * 100
                _recalc_rows.append({
                    "er": er,
                    "Потребность": int(dem),
                    "Слотов": _new_slots[j],
                    "Листов": _n_sh,
                    "Напечатано": int(actual),
                    "Разница": int(over),
                    "Разница %": round(over_p, 1),
                })
            _price_r = pkg_db.sheet_price(_n_sh, _tariffs_for_plan)
            _cost_r = _price_r * _n_sh + _extra_per_sheet * float(_n_sh)
            st.dataframe(pd.DataFrame(_recalc_rows), use_container_width=True, hide_index=True)
            st.caption(
                f"Листов: **{_n_sh:,}** · база по ступеням: **{_price_r * _n_sh:,.2f}** · "
                f"доплаты: **{_extra_per_sheet * float(_n_sh):,.2f}** · всего: **{_cost_r:,.2f}**"
            )

    if st.button(
        "Применить раскладку и перейти на «Печать и заявки»",
        key=f"pl_man_apply_{_sel_sk}",
        type="primary",
    ):
        sk_safe_a = _sel_sk.replace("|", "_").replace("/", "_")[:80]
        for idx_a in range(n_fit):
            wk = f"pp_slot_{sk_safe_a}_{idx_a}"
            er_a = _auto_slot_er[idx_a] if idx_a < len(_auto_slot_er) else None
            if er_a is not None:
                st.session_state[wk] = int(er_a)
            else:
                st.session_state[wk] = "— пусто —"
        st.session_state["pp_size_group_select"] = _sel_sk
        # Чтобы на «Печать и заявки» совпадало с рекомендацией планировщика (тот же ключ, что у number_input).
        st.session_state[f"pp_n_sheets_{sk_safe_a}"] = int(n_sheets)
        st.session_state["pp_sheet_w"] = sheet_params.width_mm
        st.session_state["pp_sheet_h"] = sheet_params.height_mm
        st.session_state["pp_margin"] = sheet_params.margin_mm
        st.session_state["pp_gap"] = sheet_params.gap_mm
        st.session_state["pp_gap_y"] = sheet_params.gap_y_mm
        st.session_state["_pl_forced_layout"] = {
            "size_key": _sel_sk,
            "n_slots": n_fit,
            "placements": [(p.x, p.y, p.w, p.h, p.rotated) for p in _pl_preview],
            "knife_w": knife_w,
            "knife_h": knife_h,
        }
        st.session_state["_pl_navigate_to_print"] = True
        st.rerun()

    # ── Экспорт ──
    st.divider()
    if st.button("Экспорт плана (PDF)", key="pl_export_plan_pdf"):
        _exp_lines: list[str] = [
            f"Планировщик печати — {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            f"Размер: {_sel_sk_disp}",
            f"Период: {_period_label}",
            f"Лист {sheet_params.width_mm:g}×{sheet_params.height_mm:g} мм; "
            f"поле {sheet_params.margin_mm:g} мм; зазор X {sheet_params.gap_mm:g}, Y {sheet_params.gap_y_mm:g}",
            f"Ячеек на листе: {n_fit}",
            "",
            f"Рекомендация: {n_sheets:,} листов · стоимость (с доплатами): {total_cost:,.2f}",
            f"Доплаты за отделку: {_extra_per_sheet:.2f} за лист · всего {_extra_total_combined:,.2f}",
            f"Если раздельно: {sep_sheets:,} листов · стоимость: {sep_cost:,.2f}",
            "",
        ]
        for i, (er, dem) in enumerate(ers_sorted):
            ein = rows_by_er.get(er) or {}
            actual = raw_slots[i] * n_sheets
            over = actual - dem
            ann = annual_demand_by_er.get(er, 0)
            _exp_lines.append(
                f"  er={er}: {(ein.get('name') or '')[:50]} | "
                f"год={int(ann)} потр({_period_label})={int(dem)} слотов={raw_slots[i]} напеч={int(actual)} "
                f"разница={int(over)} ({(over / max(dem, 1)) * 100:.1f}%)"
            )
        if empty_slots:
            _exp_lines.append(f"  Свободных ячеек: {empty_slots}")
        _exp_lines.append("")

        _pdf_plan_bytes = pse.sheet_layout_to_pdf_bytes(
            sheet_params, [], [], _exp_lines,
            title_line=f"План печати — {_sel_sk_disp}",
        )
        st.download_button(
            "Скачать план (PDF)",
            data=_pdf_plan_bytes,
            file_name=f"print_plan_{_sel_sk}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
            mime="application/pdf",
            key="pl_dl_plan_pdf",
        )

    if _planner_conn is not None:
        try:
            _planner_conn.close()
        except Exception:
            pass


def clear_packaging_row_widget_keys() -> None:
    """Сбрасывает поля строк таблицы «Макеты» в session_state после перечитки Excel или БД."""
    prefixes = (
        "name_",
        "use_custom_",
        "kind_sel_",
        "kind_cust_",
        "price_",
        "price_new_",
        "qty_sheet_",
        "qty_year_",
    )
    for k in list(st.session_state.keys()):
        if isinstance(k, str) and any(k.startswith(p) for p in prefixes):
            st.session_state.pop(k, None)


def clear_kind_widget_keys_for_excel_rows(rows: list[dict[str, Any]], excel_rows: set[int]) -> None:
    """Сбрасывает ключи «Вид» для указанных excel_row (чтобы не затёрть данные из БД устаревшим selectbox)."""
    if not excel_rows:
        return
    for item in rows:
        rk = int(item["excel_row"])
        if rk not in excel_rows:
            continue
        suffix = _widget_key_suffix(item.get("file") or str(rk))
        st.session_state.pop(f"kind_sel_{suffix}_{rk}", None)
        st.session_state.pop(f"kind_cust_{suffix}_{rk}", None)


def merge_kind_from_db(
    rows: list[dict[str, Any]],
    db_path: Path,
    excel_path: Path | None = None,
    *,
    overwrite_nonempty_excel: bool = False,
) -> None:
    """
    После чтения Excel подставляет «Вид» из SQLite для совпадающих excel_row.

    По умолчанию (overwrite_nonempty_excel=False) подставляет вид из БД только если в Excel
    ячейка «Вид» пустая — непустое значение из файла не трогаем, чтобы после «Сохранить в Excel и БД»
    не затирать актуальные виды старыми данными из SQLite при следующей загрузке.

    Если overwrite_nonempty_excel=True — прежнее поведение: при расхождении с БД строка
    перезаписывается видом из SQLite (и при необходимости файл Excel).

    Для строк, у которых подставили вид из БД, сбрасываются ключи selectbox «Вид» в session_state,
    иначе устаревшее значение виджета при следующем sync_widgets_to_rows затрёт данные из БД.

    Если вид изменился относительно только что прочитанного Excel и задан excel_path —
    перезаписывает файл, чтобы на диске совпадало с подставленными значениями (одна операция save).
    """
    any_kind_change, kind_fixed_rows = merge_kind_values_from_sqlite(
        rows,
        db_path,
        overwrite_nonempty_excel=overwrite_nonempty_excel,
    )
    if kind_fixed_rows:
        clear_kind_widget_keys_for_excel_rows(rows, kind_fixed_rows)
    if any_kind_change and excel_path is not None and excel_path.is_file():
        try:
            save_rows_to_excel(excel_path, rows, db_path)
        except Exception:
            pass


def _format_cg_qty_band(min_qty: int, max_qty: int | None) -> str:
    a = f"{int(min_qty):,}".replace(",", "\u202f")
    if max_qty is None:
        return f"{a}–∞"
    b = f"{int(max_qty):,}".replace(",", "\u202f")
    return f"{a}–{b}"


def render_makety_cg_supplier_prices_by_kind(
    rows: list[dict[str, Any]],
    db_path: Path,
) -> None:
    """
    Блок «Макеты»: цены поставщика CG (старая / новая за 1000 шт.) по градациям тиража,
    сгруппировано по виду упаковки (коробка / блистер / пакет / этикетка).
    """
    import pandas as pd

    with st.expander(
        "Цены поставщика (CG): старые и новые по градациям тиража, по видам упаковки",
        expanded=False,
    ):
        st.caption(
            "Данные из SQLite: прайс Cutting Group и сопоставление «строка Excel → нож CG» "
            "(задаётся на вкладке «Планировщик»). Для каждого вида коробок показаны ножи, "
            "которые встречаются у строк с этим видом, и таблица ступеней: тип отделки, диапазон шт., цены за 1000 шт."
        )
        if not db_path.is_file():
            st.caption("Файл базы не найден — укажите путь в боковой панели.")
            return
        try:
            conn = pkg_db.connect(db_path)
            try:
                pkg_db.init_db(conn)
                cg_prices = pkg_db.load_cg_prices(conn)
                cg_knives = pkg_db.load_cg_knives(conn)
                cg_map = pkg_db.load_cg_mapping(conn)
            finally:
                conn.close()
        except Exception as e:
            st.warning(f"Не удалось прочитать прайс CG: {e}")
            return

        if not cg_prices:
            st.info(
                "Прайс поставщика пуст. Загрузите файл CG на вкладке «Планировщик» "
                "(блок загрузки ножей и цен)."
            )
            return

        knife_names = {k["cutit_no"]: (k.get("name") or "").strip() for k in cg_knives}
        cutits_by_bucket: dict[str, set[str]] = defaultdict(set)
        for it in rows:
            er = int(it["excel_row"])
            mp = cg_map.get(er)
            cn = (mp.get("cutit_no") or "").strip() if mp else ""
            if not cn:
                continue
            cutits_by_bucket[kind_bucket(it)].add(cn)

        bucket_title = {
            "box": "Коробка",
            "blister": "Блистер",
            "pack": "Пакет",
            "label": "Этикетка и прочее",
        }
        any_shown = False
        for bkey in ("box", "blister", "pack", "label"):
            cset = cutits_by_bucket.get(bkey) or set()
            if not cset:
                continue
            any_shown = True
            st.markdown(f"##### {bucket_title[bkey]}")
            for cutit in sorted(cset):
                nm = knife_names.get(cutit) or ""
                nm_tail = (" — " + html.escape(nm[:140])) if nm else ""
                st.markdown(
                    f'<p style="font-size:0.82rem;margin:0.15rem 0 0.35rem 0;">'
                    f"<strong>Нож</strong> <code>{html.escape(cutit)}</code>"
                    f"{nm_tail}</p>",
                    unsafe_allow_html=True,
                )
                tiers = [p for p in cg_prices if p["cutit_no"] == cutit]
                tiers.sort(key=lambda x: (str(x["finish_type"]), int(x["min_qty"])))
                df_rows: list[dict[str, Any]] = []
                for p in tiers:
                    ft = str(p["finish_type"])
                    old_v = p.get("price_old_per_1000")
                    old_s = f"{float(old_v):.2f}" if old_v is not None else "—"
                    df_rows.append(
                        {
                            "Отделка": CG_FINISH_LABELS_MAKETY.get(ft, ft),
                            "Тираж, шт.": _format_cg_qty_band(
                                int(p["min_qty"]),
                                int(p["max_qty"]) if p.get("max_qty") is not None else None,
                            ),
                            "Старая за 1000 шт.": old_s,
                            "Новая за 1000 шт.": f"{float(p['price_per_1000']):.2f}",
                        }
                    )
                if df_rows:
                    st.dataframe(
                        pd.DataFrame(df_rows),
                        use_container_width=True,
                        hide_index=True,
                    )
                else:
                    st.caption("Нет строк прайса для этого ножа.")
            st.markdown("")

        if not any_shown:
            st.info(
                "Ни одна строка макетов не сопоставлена с ножом CG. "
                "Откройте вкладку «Планировщик» и привяжите ножи к позициям."
            )



def render_packaging_color_analytics(
    rows: list[dict[str, Any]],
    pdf_root: Path,
    *,
    bucket: str,
) -> None:
    """Анализ цветов PDF для одного вида: box | blister | pack | label."""
    import packaging_color_analysis as pca
    import pandas as pd

    sk = (bucket or "").strip().lower()
    _meta = {
        "box": ("Коробка", "коробок", "Всего коробок"),
        "blister": ("Блистер", "блистеров", "Всего блистеров"),
        "pack": ("Пакет", "пакетов", "Всего пакетов"),
        "label": ("Этикетка", "этикеток", "Всего этикеток"),
    }
    _kind_disp, _, _summary_total = _meta.get(
        sk,
        ("Позиция", "позиций", "Всего позиций"),
    )
    _light = "silver" if sk == "blister" else "keep"
    _preview_cap = "серебро" if sk == "blister" else "как в PDF"

    all_kind_rows = pca.collect_rows_for_color_bucket(rows, sk)
    st.caption(
        f"Тот же алгоритм, что для блистера: частоты цветов в PDF, базовая и доминирующая палитра, "
        f"рекомендации и PDF-превью перекраски. "
        f"Для **{_kind_disp.lower()}** "
        + (
            "в превью светлые области показаны **с серебристой основой** (как блистер)."
            if sk == "blister"
            else "фон превью **как в исходном PDF** (без имитации фольги)."
        )
    )
    if not all_kind_rows:
        st.info(f"В текущем наборе нет позиций вида «{_kind_disp}».")
        return

    _size_counts: dict[str, int] = {}
    for _kr in all_kind_rows:
        _sk_val = size_key_str(_kr)
        _size_counts[_sk_val] = _size_counts.get(_sk_val, 0) + 1
    _size_options = ["Все размеры"] + [
        f"{format_size_key_label(k)} ({v} шт.)"
        for k, v in sorted(_size_counts.items(), key=lambda x: -x[1])
    ]
    _size_keys_ordered = [None] + [
        k for k, _ in sorted(_size_counts.items(), key=lambda x: -x[1])
    ]

    _sz_sel_key = f"bl_color_size_filter_{sk}"
    _sz_choice = st.selectbox(
        f"Размер ({_kind_disp.lower()})",
        options=range(len(_size_options)),
        format_func=lambda i: _size_options[i],
        key=_sz_sel_key,
        help="Выберите конкретный размер для анализа или «Все размеры».",
    )
    _chosen_size_key = _size_keys_ordered[int(_sz_choice)] if _sz_choice else None

    if _chosen_size_key is not None:
        kind_rows = [r for r in all_kind_rows if size_key_str(r) == _chosen_size_key]
    else:
        kind_rows = all_kind_rows

    c1, c2, c3 = st.columns([1.2, 1.2, 1.2], gap="small")
    with c1:
        cluster_thr = st.slider(
            "Порог объединения оттенков (RGB delta)",
            min_value=10,
            max_value=80,
            value=int(st.session_state.get(f"bl_color_cluster_thr_{sk}", 28)),
            step=1,
            key=f"bl_color_cluster_thr_{sk}",
        )
    with c2:
        top_dom = st.slider(
            "Сколько доминирующих цветов",
            min_value=3,
            max_value=20,
            value=int(st.session_state.get(f"bl_color_top_dom_{sk}", 8)),
            step=1,
            key=f"bl_color_top_dom_{sk}",
        )
    with c3:
        _run = st.button(
            f"Запустить анализ ({len(kind_rows)} шт.)",
            type="primary",
            key=f"bl_color_run_{sk}",
            use_container_width=True,
        )

    if _run:
        with st.spinner("Сканирование PDF и анализ цветов…"):
            res = pca.build_color_stats(
                kind_rows,
                pdf_root=pdf_root,
                cluster_threshold=float(cluster_thr),
                top_n_dominant=int(top_dom),
                summary_total_label=_summary_total,
            )
        st.session_state[f"bl_color_result_{sk}"] = res

    result = st.session_state.get(f"bl_color_result_{sk}")
    if not result:
        st.info("Нажмите «Запустить анализ», чтобы построить сводку.")
        return

    summary_df = result["summary_df"]
    pos_df = result["positions_df"]
    global_df = result["global_colors_df"]
    palette_ref_df = result.get("palette_reference_df")
    dom_palette = result.get("dominant_palette") or []

    m1, m2, m3 = st.columns(3)
    m1.metric(_kind_disp, int(len(kind_rows)))
    m2.metric("С анализом PDF", int((pos_df["status"] == "ok").sum() if "status" in pos_df else 0))
    m3.metric("Доминирующая палитра", int(len(dom_palette)))

    if dom_palette:
        st.caption("Доминирующие цвета выборки: " + ", ".join(dom_palette))

    st.markdown("##### Сводка")
    st.dataframe(summary_df, use_container_width=True, hide_index=True)

    st.markdown("##### Топ цветов (глобально)")
    st.dataframe(global_df.head(30), use_container_width=True, hide_index=True)

    # --- Визуальная таблица цветов с чекбоксами и превью фильтрации ---
    if not global_df.empty and "hex" in global_df.columns:
        st.markdown("##### Визуальный обзор цветов")
        st.caption(
            "Включите/отключите цвета чекбоксами. Превью ниже покажет, "
            "как будет выглядеть макет только с выбранными цветами (остальные → белый)."
        )
        _n_colors_vis = min(20, len(global_df))
        _vis_cols_per_row = 5
        _active_colors: list[tuple[int, int, int]] = []

        for _vr_start in range(0, _n_colors_vis, _vis_cols_per_row):
            _vcols = st.columns(_vis_cols_per_row, gap="small")
            for _vj in range(_vis_cols_per_row):
                _vidx = _vr_start + _vj
                if _vidx >= _n_colors_vis:
                    break
                _row_g = global_df.iloc[_vidx]
                _hex_g = str(_row_g["hex"]).strip()
                _rgb_g = pca.hex_to_rgb_u8(_hex_g)
                _weight_g = int(_row_g.get("weight", 0))
                _cmyk_g = str(_row_g.get("cmyk", ""))
                _pms_g = str(_row_g.get("pantone_approx", ""))
                _cb_key = f"bl_vis_cb_{sk}_{_vidx}"
                if _cb_key not in st.session_state:
                    st.session_state[_cb_key] = True
                with _vcols[_vj]:
                    _text_col = "#fff" if sum(_rgb_g) < 380 else "#000"
                    st.markdown(
                        f'<div style="background:{_hex_g};color:{_text_col};'
                        f'padding:10px 6px;border-radius:6px;text-align:center;'
                        f'margin-bottom:4px;font-size:0.78rem;line-height:1.3;">'
                        f'<b>{_hex_g}</b><br>{_cmyk_g}<br>≈ {_pms_g}<br>вес {_weight_g}</div>',
                        unsafe_allow_html=True,
                    )
                    _is_on = st.checkbox(
                        "Вкл",
                        value=st.session_state[_cb_key],
                        key=_cb_key,
                        label_visibility="collapsed",
                    )
                    if _is_on:
                        _active_colors.append(_rgb_g)

        _n_active = len(_active_colors)
        st.caption(f"Выбрано **{_n_active}** из {_n_colors_vis} цветов.")

        st.markdown("###### Превью: только выбранные цвета")
        _fp_vis, _item_vis = pca.find_first_pdf_path_in_rows(kind_rows, pdf_root)
        if _fp_vis is not None and _item_vis is not None:
            if _n_active == 0:
                st.warning("Не выбрано ни одного цвета — включите хотя бы один чекбокс выше.")
            else:
                _vis_png = pca.render_pdf_with_selected_colors(
                    _fp_vis,
                    _active_colors,
                    dpi=96.0,
                    cluster_threshold=float(cluster_thr),
                    light_base=_light,
                )
                if _vis_png:
                    _vis_c1, _vis_c2 = st.columns(2, gap="small")
                    with _vis_c1:
                        st.markdown("**Исходник**")
                        _orig_pair = pca.blister_recolor_comparison_pngs(
                            _fp_vis,
                            dpi=96.0,
                            cluster_threshold=float(cluster_thr),
                            pixel_match_radius=30.0,
                            mode="core",
                            dominant_palette=[],
                            palette_items=list(pca.CORE_PALETTE.items()),
                            light_base=_light,
                        )
                        if _orig_pair:
                            st.image(io.BytesIO(_orig_pair[0]), use_container_width=True)
                    with _vis_c2:
                        st.markdown("**Только выбранные цвета**")
                        st.image(io.BytesIO(_vis_png), use_container_width=True)
                    st.caption(
                        f"Строка Excel {_item_vis.get('excel_row', '—')}. "
                        f"Пиксели, не принадлежащие выбранным {_n_active} кластерам, заменены белым."
                    )
                else:
                    st.info("Не удалось отрендерить превью для первого PDF.")
        else:
            st.info("Ни у одной позиции не найден файл PDF — превью недоступно.")

        st.divider()

    if palette_ref_df is not None and not palette_ref_df.empty:
        st.markdown("##### Палитра: CMYK и Pantone (приближение)")
        st.caption(
            "CMYK рассчитан из sRGB (без ICC-профиля), Pantone указан как ближайшее приближение. "
            "Используйте как ориентир и сверяйте с официальным каталогом перед печатью."
        )
        st.dataframe(palette_ref_df, use_container_width=True, hide_index=True, height=220)

    st.markdown(f"##### Рекомендации по позициям ({_kind_disp})")
    show_cols = [
        "excel_row",
        "name",
        "pdf",
        "status",
        "top_colors",
        "core_reco",
        "risk_core",
        "dominant_reco",
        "risk_dominant",
    ]
    show_cols = [c for c in show_cols if c in pos_df.columns]
    st.dataframe(pos_df[show_cols], use_container_width=True, hide_index=True, height=420)

    csv_sum = pca.export_color_report_csv(summary_df)
    csv_pos = pca.export_color_report_csv(pos_df)
    xlsx_sheets = {
        "summary": summary_df,
        "global_colors": global_df,
        "positions": pos_df,
    }
    if palette_ref_df is not None:
        xlsx_sheets["palette_ref"] = palette_ref_df
    xlsx_all = pca.export_color_report_xlsx(xlsx_sheets)
    e1, e2, e3 = st.columns(3, gap="small")
    with e1:
        st.download_button(
            "Скачать summary.csv",
            data=csv_sum,
            file_name=f"{sk}_color_summary.csv",
            mime="text/csv",
            use_container_width=True,
            key=f"bl_color_dl_summary_csv_{sk}",
        )
    with e2:
        st.download_button(
            "Скачать positions.csv",
            data=csv_pos,
            file_name=f"{sk}_color_positions.csv",
            mime="text/csv",
            use_container_width=True,
            key=f"bl_color_dl_positions_csv_{sk}",
        )
    with e3:
        st.download_button(
            "Скачать report.xlsx",
            data=xlsx_all,
            file_name=f"{sk}_color_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key=f"bl_color_dl_xlsx_{sk}",
        )

    st.markdown("##### PDF-превью перекраски")
    if sk == "blister":
        st.caption(
            f"Основа в превью и в PDF: **серебристая фольга** "
            f"(светлые/белые области → RGB {pca.BLISTER_PREVIEW_SILVER_RGB}). "
            "На каждом листе слева — исходник, справа — перекраска."
        )
    else:
        st.caption(
            "Светлые области **без замены на фольгу** (как в экспорте PDF). "
            "На каждом листе слева — исходник, справа — перекраска."
        )

    _n_pdf = len(kind_rows)
    _kf = f"bl_color_preview_from_{sk}"
    _kt = f"bl_color_preview_to_{sk}"
    for _k_cl, _lo_cl, _hi_cl in ((_kf, 1, _n_pdf), (_kt, 1, _n_pdf)):
        if _k_cl in st.session_state:
            try:
                _v_cl = int(st.session_state[_k_cl])
            except (TypeError, ValueError):
                _v_cl = _lo_cl
            st.session_state[_k_cl] = max(_lo_cl, min(_v_cl, _hi_cl))
    if _kf in st.session_state and _kt in st.session_state and int(st.session_state[_kt]) < int(
        st.session_state[_kf]
    ):
        st.session_state[_kt] = st.session_state[_kf]

    r1, r2 = st.columns([1.4, 1], gap="small")
    with r1:
        preview_mode = st.radio(
            "Режим перекраски",
            options=["core", "dominant"],
            format_func=lambda x: "Базовая палитра" if x == "core" else "Доминирующая палитра",
            horizontal=True,
            key=f"bl_color_preview_mode_{sk}",
        )
    with r2:
        st.caption("Диапазон в списке (включительно)")
        _rc_a, _rc_b = st.columns(2, gap="small")
        with _rc_a:
            st.number_input(
                "С №",
                min_value=1,
                max_value=max(1, _n_pdf),
                value=1,
                step=1,
                help="Первая позиция в списке для страниц PDF.",
                key=_kf,
            )
        with _rc_b:
            st.number_input(
                "По №",
                min_value=1,
                max_value=max(1, _n_pdf),
                value=max(1, _n_pdf),
                step=1,
                help="Последняя позиция включительно.",
                key=_kt,
            )
    st.caption(
        f"В списке **{_n_pdf}** позиций — интервал **1–25** или **1–{_n_pdf}** для всех. "
        "Большой PDF может быть тяжёлым для браузера."
    )

    _n_dom_preview = int(top_dom)
    for _i, (_nm, _rgb) in enumerate(pca.CORE_PALETTE.items()):
        _k = f"bl_cp_core_{sk}_{_i}"
        if _k not in st.session_state:
            st.session_state[_k] = pca._rgb_to_hex(_rgb)
    for _i in range(_n_dom_preview):
        _k = f"bl_cp_dom_{sk}_{_i}"
        if _k not in st.session_state:
            _h = "#888888"
            if _i < len(global_df) and "hex" in global_df.columns:
                _hx = str(global_df.iloc[_i]["hex"]).strip()
                _h = _hx if _hx.startswith("#") else f"#{_hx}" if len(_hx) == 6 else _h
            st.session_state[_k] = _h

    st.markdown("###### Цвета для превью (можно изменить перед генерацией)")
    st.caption(
        "Ниже те оттенки, к которым будут подгоняться кластеры PDF при сборке превью. "
        "Таблица обновляется по текущим значениям пикеров."
    )
    if preview_mode == "core":
        _core_list = list(pca.CORE_PALETTE.items())
        _nc = len(_core_list)
        for _row_start in range(0, _nc, 3):
            _cols = st.columns(3, gap="small")
            for _j in range(3):
                _idx = _row_start + _j
                if _idx >= _nc:
                    break
                _nm, _rgb = _core_list[_idx]
                with _cols[_j]:
                    st.color_picker(
                        _nm,
                        key=f"bl_cp_core_{sk}_{_idx}",
                        help=f"По умолчанию: {pca._rgb_to_hex(_rgb)}",
                    )
    else:
        for _row_start in range(0, _n_dom_preview, 4):
            _cols = st.columns(4, gap="small")
            for _j in range(4):
                _idx = _row_start + _j
                if _idx >= _n_dom_preview:
                    break
                with _cols[_j]:
                    st.color_picker(
                        f"Dominant {_idx + 1}",
                        key=f"bl_cp_dom_{sk}_{_idx}",
                        help=f"Слот {_idx + 1} (глобальный топ {_idx + 1})",
                    )

    _b1, _b2 = st.columns([1, 2], gap="small")
    with _b1:
        if st.button(
            "Сбросить цвета к анализу",
            key=f"bl_cp_reset_preview_{sk}",
            help="Базовая палитра — как CORE_PALETTE; доминирующая — как топ цветов из таблицы выше.",
        ):
            for _i, (_, _rgb) in enumerate(pca.CORE_PALETTE.items()):
                st.session_state[f"bl_cp_core_{sk}_{_i}"] = pca._rgb_to_hex(_rgb)
            for _i in range(_n_dom_preview):
                _h = "#888888"
                if _i < len(global_df) and "hex" in global_df.columns:
                    _hx = str(global_df.iloc[_i]["hex"]).strip()
                    _h = _hx if _hx.startswith("#") else f"#{_hx}" if len(_hx) == 6 else _h
                st.session_state[f"bl_cp_dom_{sk}_{_i}"] = _h
            st.rerun()

    _preview_rows: list[dict[str, Any]] = []
    if preview_mode == "core":
        for _i, _nm in enumerate(pca.CORE_PALETTE.keys()):
            _hx = str(st.session_state.get(f"bl_cp_core_{sk}_{_i}", "#808080"))
            _rgb = pca.hex_to_rgb_u8(_hx)
            _pms, _ = pca.nearest_pantone_approx(_rgb)
            _preview_rows.append(
                {
                    "Слот": _nm,
                    "Hex": _hx.upper(),
                    "RGB": f"{_rgb[0]}, {_rgb[1]}, {_rgb[2]}",
                    "CMYK": pca.cmyk_percent_str(_rgb),
                    "Pantone ≈": _pms,
                }
            )
    else:
        for _i in range(_n_dom_preview):
            _hx = str(st.session_state.get(f"bl_cp_dom_{sk}_{_i}", "#808080"))
            _rgb = pca.hex_to_rgb_u8(_hx)
            _pms, _ = pca.nearest_pantone_approx(_rgb)
            _preview_rows.append(
                {
                    "Слот": f"Dominant {_i + 1}",
                    "Hex": _hx.upper(),
                    "RGB": f"{_rgb[0]}, {_rgb[1]}, {_rgb[2]}",
                    "CMYK": pca.cmyk_percent_str(_rgb),
                    "Pantone ≈": _pms,
                }
            )

    st.dataframe(
        pd.DataFrame(_preview_rows),
        use_container_width=True,
        hide_index=True,
        height=min(320, 56 + 36 * len(_preview_rows)),
    )

    p3, p4 = st.columns(2, gap="small")
    with p3:
        preview_dpi = st.slider(
            "DPI превью",
            min_value=72,
            max_value=200,
            value=96,
            step=8,
            key=f"bl_color_preview_dpi_{sk}",
        )
    with p4:
        preview_radius = st.slider(
            "Радиус сопоставления",
            min_value=10,
            max_value=90,
            value=30,
            step=1,
            key=f"bl_color_preview_radius_{sk}",
        )

    if preview_mode == "core":
        _palette_for_pdf = [
            (_nm, pca.hex_to_rgb_u8(str(st.session_state[f"bl_cp_core_{sk}_{_i}"])))
            for _i, _nm in enumerate(pca.CORE_PALETTE.keys())
        ]
    else:
        _palette_for_pdf = [
            (f"Dominant {_i + 1}", pca.hex_to_rgb_u8(str(st.session_state[f"bl_cp_dom_{sk}_{_i}"])))
            for _i in range(_n_dom_preview)
        ]

    st.markdown(f"###### Превью на экране (первая позиция {_kind_disp} с PDF)")
    _fp0, _item0 = pca.find_first_pdf_path_in_rows(kind_rows, pdf_root)
    if _fp0 is not None and _item0 is not None:
        _pair_ui = pca.blister_recolor_comparison_pngs(
            _fp0,
            dpi=float(preview_dpi),
            cluster_threshold=float(cluster_thr),
            pixel_match_radius=float(preview_radius),
            mode=preview_mode,
            dominant_palette=[],
            palette_items=_palette_for_pdf,
            light_base=_light,
        )
        if _pair_ui:
            _cap_base = "серебро, как в PDF-блистере" if sk == "blister" else "как в исходном PDF"
            st.caption(
                f"Строка Excel {_item0.get('excel_row', '—')}: слева — исходный макет, справа — перекраска. "
                f"Светлая основа — {_cap_base}."
            )
            _pvc1, _pvc2 = st.columns(2, gap="small")
            with _pvc1:
                st.markdown("**Исходник**")
                st.image(io.BytesIO(_pair_ui[0]), use_container_width=True)
            with _pvc2:
                st.markdown("**После перекраски**")
                st.image(io.BytesIO(_pair_ui[1]), use_container_width=True)
        else:
            st.info("Для первого PDF не удалось снять цвета/кластеры — превью недоступно.")
    else:
        st.info("Ни у одной позиции не найден файл PDF по пути из таблицы.")

    run_preview = st.button(
        "Собрать PDF превью перекраски",
        key=f"bl_color_preview_run_{sk}",
        use_container_width=True,
    )
    if run_preview:
        _p_from = max(1, min(int(st.session_state.get(_kf, 1)), _n_pdf))
        _p_to = max(1, min(int(st.session_state.get(_kt, _n_pdf)), _n_pdf))
        if _p_to < _p_from:
            _p_to = _p_from
        with st.spinner("Генерация PDF-превью перекраски..."):
            preview_res = pca.build_recolor_preview_pdf_bytes(
                kind_rows,
                pdf_root=pdf_root,
                mode=preview_mode,
                cluster_threshold=float(cluster_thr),
                top_n_dominant=int(top_dom),
                item_from_1=_p_from,
                item_to_1=_p_to,
                dpi=int(preview_dpi),
                pixel_match_radius=float(preview_radius),
                palette_items=_palette_for_pdf,
                light_base=_light,
                preview_base_caption=_preview_cap,
            )
        preview_res["_range_from"] = _p_from
        preview_res["_range_to"] = _p_to
        st.session_state[f"bl_color_preview_result_{sk}"] = preview_res

    preview_result = st.session_state.get(f"bl_color_preview_result_{sk}")
    if preview_result:
        if preview_result.get("ok"):
            _rf = int(preview_result.get("_range_from") or 1)
            _rt = int(preview_result.get("_range_to") or _n_pdf)
            st.success(
                f"Готово: диапазон **{_rf}–{_rt}**, страниц {int(preview_result.get('generated') or 0)}, "
                f"пропущено {int(preview_result.get('skipped') or 0)}."
            )
            st.download_button(
                "Скачать PDF превью",
                data=preview_result.get("pdf_bytes") or b"",
                file_name=f"{sk}_recolor_preview_{preview_mode}_{_rf}-{_rt}.pdf",
                mime="application/pdf",
                use_container_width=True,
                key=f"bl_color_preview_download_{sk}",
            )
        else:
            st.warning(preview_result.get("error") or "Не удалось собрать PDF-превью.")


def render_product_card_export(rows: list[dict[str, Any]], pdf_root: Path, db_path: Path | None = None) -> None:
    """UI: выбор коробки, подбор связанных элементов, остатки/прогноз, генерация PDF-карточки."""
    from modules.packaging_catalog.application.product_group import (
        find_related_items,
        format_row_label,
    )
    from modules.packaging_catalog.application.product_card_pdf import build_product_card_pdf
    from modules.packaging_catalog.application.product_card_data import collect_product_card_data

    with st.expander("Карточка препарата (PDF)", expanded=False):
        st.caption(
            "Выберите коробку — система предложит связанные блистеры, этикетки и пакеты "
            "по GMP-коду и названию. Укажите остатки и скачайте PDF-карточку с прогнозом."
        )

        box_rows = [r for r in rows if kind_bucket(r) == "box"]
        if not box_rows:
            st.info("В каталоге нет позиций вида «Коробка».")
            return

        _search_col, _clear_col = st.columns([4, 1], gap="small")
        with _search_col:
            _pc_search = st.text_input(
                "Поиск коробки",
                value=st.session_state.get("_pc_search_text", ""),
                placeholder="Название, GMP, размер…",
                key="_pc_search_text",
            )
        with _clear_col:
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            if st.button("✕", key="pc_clear_search", help="Очистить поиск"):
                st.session_state["_pc_search_text"] = ""
                st.session_state.pop("pc_box_select", None)
                st.session_state.pop("_pc_pdf_bytes", None)
                for _ck in ("pc_rel_blister", "pc_rel_label", "pc_rel_pack"):
                    st.session_state.pop(_ck, None)
                st.rerun()

        _q = (_pc_search or "").strip().lower()
        if _q:
            _filtered_boxes = [
                r for r in box_rows
                if _q in format_row_label(r).lower()
            ]
        else:
            _filtered_boxes = box_rows

        if not _filtered_boxes:
            st.info(f"По запросу «{_pc_search}» коробок не найдено.")
            return

        box_options = [format_row_label(r) for r in _filtered_boxes]
        box_idx = st.selectbox(
            f"Коробка ({len(_filtered_boxes)} из {len(box_rows)})",
            options=range(len(box_options)),
            format_func=lambda i: box_options[i],
            key="pc_box_select",
        )
        box_row = _filtered_boxes[int(box_idx)] if box_idx is not None else _filtered_boxes[0]

        _gmp = (box_row.get("gmp_code") or "").strip()
        if not _gmp:
            _gmp = pkg_db.extract_gmp_code(box_row.get("name") or "", box_row.get("file") or "")

        related = find_related_items(box_row, rows)

        selected: dict[str, dict[str, Any] | None] = {}
        _kind_labels = {"blister": "Блистер", "label": "Этикетка", "pack": "Пакет"}
        cols = st.columns(3, gap="small")
        for i, (kind_key, kind_label) in enumerate(_kind_labels.items()):
            candidates = related.get(kind_key, [])
            with cols[i]:
                if not candidates:
                    st.caption(f"{kind_label}: не найдено")
                    selected[kind_key] = None
                else:
                    opts = ["— нет —"] + [format_row_label(c) for c in candidates]
                    choice = st.selectbox(
                        kind_label,
                        options=range(len(opts)),
                        format_func=lambda j, _o=opts: _o[j],
                        key=f"pc_rel_{kind_key}",
                    )
                    if choice and int(choice) > 0:
                        selected[kind_key] = candidates[int(choice) - 1]
                    else:
                        selected[kind_key] = None

        n_selected = sum(1 for v in selected.values() if v is not None)
        st.caption(f"Коробка + {n_selected} связанных элементов.")

        # ── Остатки упаковки ──
        st.markdown("##### Остатки упаковки на складах")
        _stock_kinds = {"box": "Коробка", "blister": "Блистер", "label": "Этикетка", "pack": "Пакет"}

        _existing_pkg_stock: dict[str, float] = {}
        if db_path and db_path.is_file() and _gmp:
            try:
                _pc_conn = pkg_db.connect(db_path)
                pkg_db.init_db(_pc_conn)
                _existing_pkg_stock = pkg_db.load_packaging_stock_for_gmp(_pc_conn, _gmp)
                _pc_conn.close()
            except Exception:
                pass

        _stock_cols = st.columns(4, gap="small")
        _stock_vals: dict[str, float] = {}
        for idx, (sk, sl) in enumerate(_stock_kinds.items()):
            with _stock_cols[idx]:
                _stock_vals[sk] = st.number_input(
                    f"{sl} (шт.)",
                    min_value=0.0,
                    value=float(_existing_pkg_stock.get(sk, 0.0)),
                    step=100.0,
                    key=f"pc_stock_{sk}",
                )

        # ── Остатки субстанции ──
        st.markdown("##### Остатки субстанции (препарата)")

        _existing_sub: dict[str, Any] = {}
        if db_path and db_path.is_file() and _gmp:
            try:
                _pc_conn2 = pkg_db.connect(db_path)
                pkg_db.init_db(_pc_conn2)
                _sub_all = pkg_db.load_substance_stock(_pc_conn2, _gmp)
                _existing_sub = _sub_all.get(_gmp.strip().upper(), {})
                _pc_conn2.close()
            except Exception:
                pass

        _sub_c1, _sub_c2 = st.columns([2, 1], gap="small")
        with _sub_c1:
            _sub_qty = st.number_input(
                "Количество",
                min_value=0.0,
                value=float(_existing_sub.get("qty", 0.0)),
                step=1.0,
                key="pc_sub_qty",
            )
        with _sub_c2:
            _sub_unit_options = list(pkg_db.SUBSTANCE_UNITS)
            _cur_unit = _existing_sub.get("unit", "кг")
            _sub_unit_idx = _sub_unit_options.index(_cur_unit) if _cur_unit in _sub_unit_options else 0
            _sub_unit = st.selectbox(
                "Единица",
                options=_sub_unit_options,
                index=_sub_unit_idx,
                key="pc_sub_unit",
            )

        # ── Кнопка сохранения остатков ──
        if st.button("Сохранить остатки", key="pc_save_stock", help="Записать остатки в БД"):
            if db_path and _gmp:
                try:
                    _sc = pkg_db.connect(db_path)
                    pkg_db.init_db(_sc)
                    for _sk, _sv in _stock_vals.items():
                        pkg_db.upsert_packaging_stock(_sc, _gmp, _sk, _sv, source="manual")
                    pkg_db.upsert_substance_stock(_sc, _gmp, _sub_qty, _sub_unit, source="manual")
                    _sc.close()
                    st.success("Остатки сохранены.")
                except Exception as _e:
                    st.error(f"Ошибка сохранения: {_e}")
            else:
                st.warning("Не удалось сохранить: нет GMP-кода или пути к БД.")

        st.divider()

        # ── Генерация PDF ──
        if st.button("Сгенерировать PDF-карточку", key="pc_gen_btn", type="primary"):
            with st.spinner("Сборка PDF…"):
                _db_str = str(db_path) if db_path and db_path.is_file() else None
                card_data = collect_product_card_data(
                    _db_str, _gmp, box_row, selected, rows,
                )
                for _sk2, _sv2 in _stock_vals.items():
                    for ps in card_data.packaging_stock:
                        if ps.kind == _sk2:
                            ps.qty = _sv2
                card_data.substance.qty = _sub_qty
                card_data.substance.unit = _sub_unit

                pdf_bytes = build_product_card_pdf(
                    box_row,
                    selected,
                    pdf_root,
                    card_data=card_data,
                )
            st.session_state["_pc_pdf_bytes"] = pdf_bytes
            _fn_parts: list[str] = []
            _fn_name = re.sub(r"[^\w\s\-]", "", (box_row.get("name") or "").strip())[:60].strip()
            if _fn_name:
                _fn_parts.append(_fn_name.replace(" ", "_"))
            _fn_box_sz = (box_row.get("size") or "").strip().replace(" ", "")
            if _fn_box_sz:
                _fn_parts.append(f"box_{_fn_box_sz}")
            _fn_lbl_row = selected.get("label")
            if _fn_lbl_row:
                _fn_lbl_sz = (_fn_lbl_row.get("size") or "").strip().replace(" ", "")
                if _fn_lbl_sz:
                    _fn_parts.append(f"label_{_fn_lbl_sz}")
            st.session_state["_pc_pdf_filename"] = (
                ("_".join(_fn_parts) if _fn_parts else "product_card") + ".pdf"
            )

        if st.session_state.get("_pc_pdf_bytes"):
            st.download_button(
                "Скачать PDF-карточку",
                data=st.session_state["_pc_pdf_bytes"],
                file_name=st.session_state.get("_pc_pdf_filename", "product_card.pdf"),
                mime="application/pdf",
                use_container_width=True,
                key="pc_download_btn",
            )


def render_packaging_color_analytics_tabs(rows: list[dict[str, Any]], pdf_root: Path) -> None:
    """Четыре вкладки: анализ цветов PDF для коробки, блистера, пакета, этикетки."""
    with st.expander("Анализ цветов PDF (коробка, блистер, пакет, этикетка)", expanded=False):
        st.caption(
            "Одинаковая логика для всех видов: частоты цветов, рекомендации по палитре, выгрузка CSV/XLSX "
            "и PDF-превью перекраски. Для блистера в превью светлые области имитируют серебряную основу; "
            "для коробки, пакета и этикетки фон как в файле PDF."
        )
        tab_box, tab_bl, tab_pk, tab_lb = st.tabs(["Коробка", "Блистер", "Пакет", "Этикетка"])
        with tab_box:
            render_packaging_color_analytics(rows, pdf_root, bucket="box")
        with tab_bl:
            render_packaging_color_analytics(rows, pdf_root, bucket="blister")
        with tab_pk:
            render_packaging_color_analytics(rows, pdf_root, bucket="pack")
        with tab_lb:
            render_packaging_color_analytics(rows, pdf_root, bucket="label")


def render_blister_color_analytics(rows: list[dict[str, Any]], pdf_root: Path) -> None:
    """Совместимость: открывает общий блок с вкладками (блистер — одна из вкладок)."""
    render_packaging_color_analytics_tabs(rows, pdf_root)


def apply_excel_file_reload_to_session() -> None:
    """Сбрасывает кэш строк; при следующем проходе данные снова читаются из файла Excel."""
    st.session_state.pop("packaging_rows", None)
    st.session_state.pop("_db_row_mirror", None)
    st.session_state.pop("pkg_size_filter_key", None)
    clear_packaging_row_widget_keys()


def main() -> None:
    st.set_page_config(page_title="Макеты упаковки", layout="wide")
    st.markdown(
        """
        <style>
        .block-container { padding-top: 0.75rem !important; padding-bottom: 1rem !important; max-width: 100% !important; }
        h1 { font-size: 1.25rem !important; font-weight: 600 !important; margin: 0 0 0.35rem 0 !important; }
        div[data-testid="stHorizontalBlock"] { gap: 0.3rem !important; align-items: center !important; }
        div[data-testid="column"] > div { gap: 0.15rem !important; align-items: center !important; }
        div[data-testid="stTextInput"] input,
        div[data-baseweb="select"] > div { min-height: 2rem !important; padding-top: 0.2rem !important; padding-bottom: 0.2rem !important; font-size: 0.8125rem !important; }
        div[data-testid="stCheckbox"] { min-height: 2rem !important; }
        div[data-testid="stCheckbox"] label { font-size: 0.75rem !important; }
        .pkg-row-hr { border: none; border-top: 1px solid rgba(128,128,128,0.25); margin: 0.1rem 0 0.2rem 0; }
        .pkg-fn { font-size: 0.72rem; line-height: 1.25; margin: 0; padding: 0; max-width: 100%; white-space: normal; overflow-wrap: anywhere; word-break: break-word; hyphens: auto; }
        div[data-testid="stImage"] img { object-fit: contain; border-radius: 2px; box-shadow: 0 0 0 1px rgba(0,0,0,0.06); }
        .pkg-nav-spacer { height: 0.65rem; }
        </style>
        """,
        unsafe_allow_html=True,
    )

    with st.sidebar:
        st.header("Пути")
        pdf_root = st.text_input("Папка с PDF", value=str(ROOT))
        excel_path_str = st.text_input("Файл Excel", value=str(DEFAULT_EXCEL))
        db_path_str = st.text_input("База SQLite", value=str(pkg_db.DEFAULT_DB_PATH))
        st.subheader("Загрузка Excel")
        st.checkbox(
            "Перезаписывать «Вид» из SQLite даже если в Excel уже заполнено",
            value=False,
            key="pkg_merge_kind_overwrite_excel",
            help="Выключено (рекомендуется): при открытии/перезагрузке Excel пустые «Вид» подтягиваются из БД; непустые в файле не меняются. Включите, если колонка «Вид» в файле устарела, а истина только в SQLite.",
        )
        st.checkbox(
            "Требовать эталонные заголовки Excel (все 13 имён листа «Макеты»)",
            value=True,
            key="pkg_excel_strict_ref",
            help="Включено: без полного набора заголовков в строке 1 загрузка не выполняется. Выключите для старых файлов или используйте «Привести Excel к эталону».",
        )
        st.subheader("Миниатюры")
        scale = st.slider(
            "Базовый масштаб рендера",
            0.10,
            0.30,
            0.18,
            0.01,
            help="Выше — больше деталей при рендере (дольше по CPU).",
        )
        sharp = st.slider(
            "Чёткость (супер-сэмплинг)",
            2.0,
            4.2,
            2.85,
            0.05,
            help="Растр рисуется с запасом и сжимается до ширины превью — выше значение, читаемее мелкий текст.",
        )
        thumb_w = st.slider(
            "Ширина превью на экране (px)",
            144,
            320,
            224,
            2,
            help="Больше пикселей по ширине + супер-сэмплинг = лучше видно подписи на макете.",
        )
        st.subheader("PDF в окне")
        max_modal_mb = st.slider(
            "Просмотр в модальном окне до (МБ)",
            6,
            50,
            22,
            1,
            help="Больше — в окне только «Скачать PDF»",
        )
        per_page = st.slider("Строк на странице", 10, 50, 24, 1)

    st.markdown('<div class="pkg-nav-spacer" aria-hidden="true"></div>', unsafe_allow_html=True)
    if st.session_state.pop("_pl_navigate_to_print", False):
        st.session_state["pkg_app_screen"] = "Печать и заявки"
    app_screen = st.radio(
        "Раздел",
        ["Макеты", "Cutii: cutii → коробки", "Печать и заявки", "Планировщик"],
        horizontal=True,
        label_visibility="collapsed",
        key="pkg_app_screen",
    )
    st.divider()

    pdf_dir = Path(pdf_root).expanduser().resolve()
    excel_path = Path(excel_path_str).expanduser().resolve()

    db_path = Path(db_path_str).expanduser().resolve()
    max_modal_bytes = int(max_modal_mb * 1024 * 1024)

    if not excel_path.is_file():
        st.error(f"Файл Excel не найден: {excel_path}")
        st.stop()

    with st.sidebar:
        st.divider()
        st.subheader("Эталон «Макеты»")
        _ref_excel, _ref_db = load_makety_paths_ref()
        _ref_bits: list[str] = []
        if _ref_excel is not None:
            _ref_bits.append(
                f"Excel: `{_ref_excel.name}`"
                + ("" if _ref_excel.is_file() else " (нет файла)")
            )
        if _ref_db is not None:
            _ref_bits.append(
                f"БД: `{_ref_db.name}`"
                + ("" if _ref_db.is_file() else " (нет файла)")
            )
        if _ref_bits:
            st.caption("Из **makety_paths_ref.json**: " + " · ".join(_ref_bits))
        else:
            st.caption(
                f"Необязательно: создайте **{MAKETY_PATHS_REF_PATH.name}** с ключами "
                "`excel_path` и `db_path` (относительно папки приложения)."
            )
        if st.button(
            "Привести Excel к эталону",
            key="pkg_normalize_excel_ref",
            help="Перечитать файл (включая старый формат), записать строку заголовков и 13 столбцов как в шаблоне, очистить хвост строк.",
        ):
            try:
                normalize_excel_file_to_makety_reference(
                    excel_path,
                    db_path if db_path.is_file() else None,
                    overwrite_nonempty_excel=bool(
                        st.session_state.get("pkg_merge_kind_overwrite_excel", False)
                    ),
                )
                apply_excel_file_reload_to_session()
                st.success(f"Файл приведён к эталону: {excel_path.name}")
                st.rerun()
            except Exception as e:
                st.error(f"Не удалось нормализовать Excel: {e}")

    if "packaging_rows" not in st.session_state:
        try:
            _strict_excel = bool(st.session_state.get("pkg_excel_strict_ref", True))
            loaded = load_rows_from_excel(
                excel_path,
                strict_reference_layout=_strict_excel,
            )
            merge_kind_from_db(
                loaded,
                db_path,
                excel_path,
                overwrite_nonempty_excel=bool(
                    st.session_state.get("pkg_merge_kind_overwrite_excel", False)
                ),
            )
            st.session_state.packaging_rows = loaded
        except Exception as e:
            st.error(f"Не удалось прочитать Excel: {e}")
            st.stop()

    if st.sidebar.button("Загрузить из SQLite"):
        if not db_path.is_file():
            st.sidebar.error("Файл БД не найден")
        else:
            conn = pkg_db.connect(db_path)
            try:
                pkg_db.init_db(conn)
                if pkg_db.row_count(conn) == 0:
                    st.sidebar.warning("В БД нет строк")
                else:
                    st.session_state.packaging_rows = pkg_db.load_all(conn)
                    st.session_state.pop("_db_row_mirror", None)
                    st.session_state.pop("pkg_size_filter_key", None)
                    clear_packaging_row_widget_keys()
                    st.sidebar.success("Загружено из БД")
                    st.rerun()
            finally:
                conn.close()

    if st.sidebar.button(
        "GMP из имён PDF → БД",
        help="По полю «Исходный PDF» извлекает код вида (ВУМ-169-01) или ВУМ-169-01 в имени файла и записывает в SQLite (gmp_code). "
        "Затем подставляет GMP в текущую сессию макетов.",
        key="pkg_sync_gmp_pdf_sidebar",
    ):
        if not db_path.is_file():
            st.sidebar.error("Файл БД не найден")
        elif "packaging_rows" not in st.session_state:
            st.sidebar.warning("Сначала загрузите Excel")
        else:
            conn_g = pkg_db.connect(db_path)
            try:
                pkg_db.init_db(conn_g)
                u_g, same_g, skip_g = pkg_db.sync_gmp_from_pdf_filenames(conn_g)
                fresh_g = pkg_db.load_all(conn_g)
            finally:
                conn_g.close()
            by_gmp = {int(r["excel_row"]): (r.get("gmp_code") or "").strip() for r in fresh_g}
            new_pack = []
            for row in st.session_state.packaging_rows:
                nr = dict(row)
                er = int(nr["excel_row"])
                if er in by_gmp:
                    nr["gmp_code"] = by_gmp[er]
                new_pack.append(nr)
            st.session_state.packaging_rows = new_pack
            st.session_state.pop("_db_row_mirror", None)
            st.sidebar.success(
                f"GMP: обновлено в БД **{u_g}**, без изменений **{same_g}**, нет кода в имени PDF **{skip_g}**."
            )
            st.rerun()

    st.sidebar.text_input(
        "Cutii → БД (помесячно)",
        value=str(ROOT.parent / "Balcan 2025 cutii.xlsx"),
        key="pkg_cutii_xlsx_quick",
        help="Файл с датами месяцев в шапке (например «Balcan 2025 cutii.xlsx»). Данные пишутся в SQLite; «Профиль Excel» и планировщик читают их из БД.",
    )
    if st.sidebar.button(
        "Импорт cutii в БД",
        key="pkg_cutii_import_sidebar",
        help="Сопоставление строк buc. с коробками (как вкладка Cutii): qty за год и помесячно в packaging_monthly_qty.",
    ):
        import import_cutii_forecast as ic

        if not db_path.is_file():
            st.sidebar.error("Файл БД не найден")
        elif "packaging_rows" not in st.session_state:
            st.sidebar.warning("Нет строк макетов в сессии")
        else:
            cp = Path(str(st.session_state.get("pkg_cutii_xlsx_quick", "")).strip()).expanduser().resolve()
            if not cp.is_file():
                st.sidebar.error(f"Файл cutii не найден: {cp}")
            else:
                try:
                    with st.spinner("Импорт cutii…"):
                        n_w, res_c = ic.import_cutii_matched_volumes_to_db(
                            cp,
                            st.session_state.packaging_rows,
                            db_path,
                            excel_path,
                        )
                except Exception as e:
                    st.sidebar.error(str(e))
                else:
                    clear_packaging_row_widget_keys()
                    st.session_state.pop("packaging_rows", None)
                    st.session_state.pop("_db_row_mirror", None)
                    n_pend = len(res_c.get("pending") or [])
                    st.sidebar.success(
                        f"Записано **{n_w}** поз. Помесячные данные в БД — профиль Excel их подхватит. "
                        f"Не импортировано (нужен разбор): **{n_pend}** строк cutii — вкладка «Cutii»."
                    )
                    st.rerun()

    excel_download_bytes: bytes | None = None
    _excel_download_err: str | None = None
    try:
        excel_download_bytes = excel_path.read_bytes()
    except OSError as e:
        _excel_download_err = str(e)

    _hx1, _hx2, _hx3, _hx4 = st.columns([4.6, 1.25, 1.25, 1.45], gap="small")
    with _hx1:
        if _excel_download_err:
            st.caption(f"Не удалось прочитать Excel для скачивания: {_excel_download_err}")
        else:
            st.caption(f"Excel: **{excel_path.name}**")
    with _hx2:
        if st.button(
            "Загрузить обновлённый Excel",
            use_container_width=True,
            type="secondary",
            help="Перечитать Excel с диска. Пустые «Вид» подставляются из SQLite; см. галочку «Загрузка Excel» в сайдбаре.",
            key="reload_excel_header",
        ):
            apply_excel_file_reload_to_session()
            st.rerun()
    with _hx3:
        if excel_download_bytes is not None:
            st.download_button(
                label="Скачать Excel",
                data=excel_download_bytes,
                file_name=excel_path.name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="pkg_download_xlsx_header",
                help="Текущий файл с диска (путь в сайдбаре «Файл Excel»)",
            )
    with _hx4:
        if db_path.is_file() and "packaging_rows" in st.session_state and st.session_state.packaging_rows:
            import importlib

            import packaging_print_planning as _pp_hdr
            import packaging_profile_excel as _pprof_hdr

            importlib.reload(_pprof_hdr)
            _rows_hdr = sorted(
                st.session_state.packaging_rows,
                key=lambda x: int(x["excel_row"]),
            )
            _by_er_hdr = {int(r["excel_row"]): r for r in _rows_hdr}
            _sp_hdr = _pp_hdr.SheetParams(
                width_mm=float(st.session_state.get("pl_sheet_w", 700.0)),
                height_mm=float(st.session_state.get("pl_sheet_h", 1000.0)),
                margin_mm=float(st.session_state.get("pl_margin", 5.0)),
                gap_mm=float(st.session_state.get("pl_gap_x", 2.0)),
                gap_y_mm=float(st.session_state.get("pl_gap_y", 2.0)),
            )
            _fc_hdr = str(st.session_state.get("pl_profile_finish_pick") or "lac_wb")
            _ry_hdr = int(st.session_state.get("pl_profile_report_year") or date.today().year)
            _dc_hdr = str(st.session_state.get("pl_profile_doc_code") or "OM/ПУМ-192-01-373")
            try:
                _prof_hdr_bytes = _pprof_hdr.build_profile_workbook_bytes(
                    db_path=db_path,
                    size_key="all",
                    size_key_display_override=f"Вся база ({len(_rows_hdr)} поз.)",
                    group_rows=_rows_hdr,
                    rows_by_er=_by_er_hdr,
                    sheet_params=_sp_hdr,
                    document_code=_dc_hdr,
                    finish_code=_fc_hdr,
                    logo_path=None,
                    report_year=_ry_hdr,
                )
                st.download_button(
                    label="Профиль Excel",
                    data=_prof_hdr_bytes,
                    file_name="packaging_profile_all.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="pkg_download_profile_header",
                    help="Тот же шаблон, что в планировщике: шапка Balkan, таблица A–Z, анализ и график по **всем** строкам сессии. "
                    "Год месяцев, отделка CG и Cod document — как в блоке экспорта на вкладке «Планировщик» (или значения по умолчанию). "
                    "Параметры листа — из планировщика, если открывали, иначе 700×1000 мм.",
                )
            except Exception as _e_hdr_prof:
                st.caption(f"Профиль: {_e_hdr_prof}")
        elif not db_path.is_file():
            st.caption("БД нет — профиль недоступен")

    rows: list[dict[str, Any]] = st.session_state.packaging_rows

    if app_screen == "Cutii: cutii → коробки":
        render_cutii_tab(excel_path, db_path, rows)
        st.stop()

    if app_screen == "Печать и заявки":
        render_print_orders_tab(rows, pdf_dir, db_path)
        st.stop()

    if app_screen == "Планировщик":
        render_planner_tab(rows, db_path, pdf_dir, excel_path=excel_path)
        st.stop()

    st.title("Макеты упаковки")

    search = st.text_input(
        "Поиск",
        placeholder="PDF, нож CG, категория, лаки, вид, размер, цена, количество или № строки…",
        key="pkg_search_makety",
        help="Фильтр по подстроке (без учёта регистра). Сочетается с кнопками вида и габаритов.",
    )
    q = search.strip().lower()

    sync_widgets_to_rows(rows)
    apply_makety_cg_derived_from_db(db_path, rows)
    reconcile_row_to_storage(db_path, excel_path, rows)
    kind_options = build_kind_options(rows)

    stc = kind_stats(rows)
    if "pkg_bucket_filter" not in st.session_state:
        st.session_state.pkg_bucket_filter = "all"
    if "pkg_size_filter_key" not in st.session_state:
        st.session_state.pkg_size_filter_key = None

    st.markdown(
        '<p style="font-size:0.8rem;margin:0 0 0.35rem 0;color:#666;">'
        "Статистика по виду — <strong>нажмите категорию</strong>, чтобы в таблице "
        "остались только такие позиции.</p>",
        unsafe_allow_html=True,
    )
    if st.session_state.get("pkg_bucket_filter") == "other":
        st.session_state.pkg_bucket_filter = "all"

    bcols = st.columns(5, gap="small")
    bucket_buttons: list[tuple[str, int, str]] = [
        ("Все", len(rows), "all"),
        ("Коробка", stc["Коробки"], "box"),
        ("Блистер", stc["Блистеры"], "blister"),
        ("Пакет", stc["Пакеты"], "pack"),
        ("Этикетка", stc["Этикетки"], "label"),
    ]
    cur_b = st.session_state.pkg_bucket_filter
    for i, (title, count, bkey) in enumerate(bucket_buttons):
        with bcols[i]:
            btn_kw: dict[str, Any] = {"use_container_width": True}
            if cur_b == bkey:
                btn_kw["type"] = "primary"
            if st.button(
                f"{title} · {count}",
                key=f"bucket_filter_{bkey}",
                **btn_kw,
            ):
                st.session_state.pkg_bucket_filter = bkey
                st.rerun()

    ref_total, ref_kind = load_makety_catalog_ref()
    with st.expander(f"Эталон каталога ({ref_total} позиций)", expanded=False):
        st.caption(
            "Сверка по текущей загрузке «Макеты». Расхождение — другие виды или число строк. "
            f"Эталон хранится в **{MAKETY_CATALOG_REF_PATH.name}** (если файла нет — встроенные 852 позиции). "
            "После полного импорта нажмите «Обновить эталон», чтобы дальше сравнивать с актуальной базой."
        )
        if st.button(
            "Обновить эталон",
            key="makety_catalog_ref_update",
            help="Записать в JSON текущее число строк и разбивку по видам как новый эталон сверки.",
        ):
            try:
                save_makety_catalog_ref(len(rows), stc)
                st.toast(f"Эталон обновлён: {len(rows)} строк.", icon="✅")
                st.rerun()
            except Exception as e:
                st.error(f"Не удалось сохранить эталон: {e}")
        total_ok = len(rows) == ref_total
        lines = [
            "| Показатель | Эталон | Сейчас |",
            "| :--- | ---: | ---: |",
            f"| Всего строк | {ref_total} | {len(rows)} |",
        ]
        all_ok = total_ok
        for lbl in ("Коробки", "Блистеры", "Пакеты", "Этикетки"):
            exp = ref_kind[lbl]
            cur = stc[lbl]
            ok = cur == exp
            all_ok = all_ok and ok
            lines.append(f"| {lbl} | {exp} | {cur} |")
        st.markdown("\n".join(lines))
        if all_ok:
            st.success("Совпадает с эталоном.")
        else:
            st.warning(
                "Есть отличия от эталона. Если после сохранения «Вид» снова «прыгает», "
                "проверьте, выключена ли в сайдбаре опция перезаписи вида из SQLite при загрузке Excel."
            )

    render_makety_cg_supplier_prices_by_kind(rows, db_path)
    render_product_card_export(rows, pdf_dir, db_path)
    render_packaging_color_analytics_tabs(rows, pdf_dir)

    if q:
        filtered = [r for r in rows if item_matches_text_query(r, q)]
    else:
        filtered = list(rows)

    bkt = st.session_state.pkg_bucket_filter
    if bkt != "all":
        filtered = [r for r in filtered if item_matches_bucket(r, bkt)]

    size_counts = Counter(size_key_str(r) for r in filtered)
    cur_sz = st.session_state.pkg_size_filter_key
    if cur_sz is not None and cur_sz not in size_counts:
        st.session_state.pkg_size_filter_key = None
        cur_sz = None

    def _sort_size_keys(keys: list[str]) -> list[str]:
        def _keyfn(k: str) -> tuple:
            if k == "__empty__":
                return (1, ())
            return (0, tuple(int(x) for x in k.split("|")))

        return sorted(keys, key=_keyfn)

    sorted_size_keys = _sort_size_keys(list(size_counts.keys()))

    st.markdown(
        '<p style="font-size:0.8rem;margin:0.5rem 0 0.35rem 0;color:#666;">'
        "<strong>Габариты (мм)</strong> — для текущего поиска и вида показаны размеры и "
        "сколько позиций с каждым; <strong>нажмите размер</strong>, чтобы оставить только его. "
        'Кнопка «Все размеры» сбрасывает фильтр.</p>',
        unsafe_allow_html=True,
    )
    all_cols = st.columns([1.0] + [1.0] * min(5, max(0, len(sorted_size_keys))), gap="small")
    with all_cols[0]:
        btn_all: dict[str, Any] = {"use_container_width": True}
        if cur_sz is None:
            btn_all["type"] = "primary"
        if st.button("Все размеры", key="pkg_size_all", **btn_all):
            st.session_state.pkg_size_filter_key = None
            st.rerun()
    n_rest = len(all_cols) - 1
    for idx, sk in enumerate(sorted_size_keys[:n_rest]):
        with all_cols[idx + 1]:
            lbl = format_size_key_label(sk)
            cnt = size_counts[sk]
            btn_kw: dict[str, Any] = {"use_container_width": True}
            if cur_sz == sk:
                btn_kw["type"] = "primary"
            safe = "e" if sk == "__empty__" else sk.replace("|", "x")
            if st.button(f"{lbl} · {cnt}", key=f"pkg_size_btn_{safe}", **btn_kw):
                st.session_state.pkg_size_filter_key = sk
                st.rerun()
    if len(sorted_size_keys) > n_rest:
        with st.expander(
            f"Ещё размеры ({len(sorted_size_keys) - n_rest})",
            expanded=False,
        ):
            ex_cols = st.columns(4, gap="small")
            for j, sk in enumerate(sorted_size_keys[n_rest:]):
                with ex_cols[j % 4]:
                    lbl = format_size_key_label(sk)
                    cnt = size_counts[sk]
                    btn_kw2: dict[str, Any] = {"use_container_width": True}
                    if cur_sz == sk:
                        btn_kw2["type"] = "primary"
                    safe = "e" if sk == "__empty__" else sk.replace("|", "x")
                    if st.button(
                        f"{lbl} · {cnt}",
                        key=f"pkg_size_btn_more_{j}_{safe}",
                        **btn_kw2,
                    ):
                        st.session_state.pkg_size_filter_key = sk
                        st.rerun()

    if cur_sz is not None:
        filtered = [r for r in filtered if item_matches_size_key(r, cur_sz)]

    ctl = st.columns([1.55, 1.0, 0.9, 0.9], gap="small")
    with ctl[0]:
        sort_by = st.selectbox(
            "Сортировка",
            (
                "По строке Excel",
                "По виду",
                "По размеру (габариты мм)",
                "По названию",
                "По PDF",
            ),
            key="pkg_sort_by",
            label_visibility="collapsed",
        )
    with ctl[1]:
        sort_rev = st.checkbox("Обратно", key="pkg_sort_rev")

    filtered = sort_rows(filtered, sort_by, sort_rev)

    total_pages = max(1, (len(filtered) + per_page - 1) // per_page)
    # Кнопки ± внизу списка не могут писать в pkg_page после number_input — только через прыжок:
    if "_pkg_page_jump" in st.session_state:
        jp = int(st.session_state.pop("_pkg_page_jump"))
        st.session_state.pkg_page = min(max(1, jp), total_pages)
    elif "pkg_page" not in st.session_state:
        st.session_state.pkg_page = 1
    else:
        st.session_state.pkg_page = min(
            max(1, int(st.session_state.pkg_page)),
            total_pages,
        )
    with ctl[2]:
        st.number_input(
            "Стр.",
            min_value=1,
            max_value=total_pages,
            step=1,
            label_visibility="visible",
            key="pkg_page",
        )
    page = int(st.session_state.pkg_page)
    with ctl[3]:
        st.caption(f"{len(filtered)}/{len(rows)}")
    start = (page - 1) * per_page
    chunk = filtered[start : start + per_page]

    _init_makety_col_width_session()
    with st.expander("Ширина столбцов таблицы", expanded=False):
        st.caption(
            "Относительные доли ширины (сумма не важна — важны пропорции). "
            "Действует на заголовок и все строки на странице. "
            f"**Сохранить умолчание** записывает текущие ползунки в **{MAKETY_COL_WIDTHS_USER_PATH.name}** "
            "— при следующем открытии приложения подставятся они. **По умолчанию** — встроенные заводские пропорции."
        )
        _bc1, _bc2, _bc3 = st.columns([2, 1, 1], gap="small")
        with _bc2:
            if st.button("По умолчанию", key="pkg_col_w_reset", use_container_width=True):
                for i, d in enumerate(MAKETY_COL_WIDTH_DEFAULTS):
                    st.session_state[f"pkg_col_w_{i}"] = float(d)
                st.rerun()
        with _bc3:
            if st.button("Сохранить умолчание", key="pkg_col_w_save_user", use_container_width=True):
                try:
                    save_user_makety_col_widths(_makety_col_weights())
                    st.success(f"Сохранено: {MAKETY_COL_WIDTHS_USER_PATH.name}")
                except Exception as e:
                    st.error(f"Не удалось сохранить: {e}")
        for row0 in range(0, len(MAKETY_COL_LABELS), 3):
            sc = st.columns(3)
            for j in range(3):
                idx = row0 + j
                if idx >= len(MAKETY_COL_LABELS):
                    break
                with sc[j]:
                    st.slider(
                        MAKETY_COL_LABELS[idx],
                        min_value=0.15,
                        max_value=5.0,
                        step=0.05,
                        key=f"pkg_col_w_{idx}",
                        help="Шире столбец — больше доля строки",
                    )
    colw = _makety_col_weights()

    # Заголовок: PDF → каталог CG → вид → превью → размеры → нож CG → цены → кол-ва
    hdr = st.columns(colw, gap="small")
    with hdr[0]:
        st.markdown('<p style="font-size:0.72rem;margin:0;font-weight:600;">PDF</p>', unsafe_allow_html=True)
    with hdr[1]:
        st.markdown(
            '<p style="font-size:0.7rem;margin:0;font-weight:600;line-height:1.1;">Название<br/>ножа CG</p>',
            unsafe_allow_html=True,
        )
    with hdr[2]:
        st.markdown(
            '<p style="font-size:0.7rem;margin:0;font-weight:600;line-height:1.1;">Катего-<br/>рия CG</p>',
            unsafe_allow_html=True,
        )
    with hdr[3]:
        st.markdown('<p style="font-size:0.7rem;margin:0;font-weight:600;line-height:1.1;">Лаки CG</p>', unsafe_allow_html=True)
    with hdr[4]:
        st.markdown('<p style="font-size:0.72rem;margin:0;font-weight:600;">Вид</p>', unsafe_allow_html=True)
    with hdr[5]:
        st.markdown('<p style="font-size:0.72rem;margin:0;font-weight:600;">Превью</p>', unsafe_allow_html=True)
    with hdr[6]:
        st.markdown('<p style="font-size:0.72rem;margin:0;font-weight:600;">Размер</p>', unsafe_allow_html=True)
    with hdr[7]:
        st.markdown(
            '<p style="font-size:0.68rem;margin:0;font-weight:600;line-height:1.05;">Размер<br/>ножа</p>',
            unsafe_allow_html=True,
        )
    with hdr[8]:
        st.markdown('<p style="font-size:0.72rem;margin:0;font-weight:600;">Нож CG</p>', unsafe_allow_html=True)
    with hdr[9]:
        st.markdown(
            '<p style="font-size:0.68rem;margin:0;font-weight:600;line-height:1.05;">Цена CG</p>'
            '<p style="font-size:0.58rem;margin:0;opacity:0.88;font-weight:500;">за 1000 шт.</p>',
            unsafe_allow_html=True,
        )
    with hdr[10]:
        st.markdown(
            '<p style="font-size:0.65rem;margin:0;font-weight:600;line-height:1.05;">Нов. цена</p>',
            unsafe_allow_html=True,
        )
    with hdr[11]:
        st.markdown(
            '<p style="font-size:0.68rem;margin:0;font-weight:600;line-height:1.1;">На листе</p>',
            unsafe_allow_html=True,
        )
    with hdr[12]:
        st.markdown(
            '<p style="font-size:0.68rem;margin:0;font-weight:600;line-height:1.1;">За год</p>',
            unsafe_allow_html=True,
        )
    st.markdown('<hr class="pkg-row-hr"/>', unsafe_allow_html=True)

    tw = int(thumb_w)
    # Растр заметно шире экранной ширины — при уменьшении до tw px текст остаётся читаемым
    max_raster = max(600, int(tw * 4.25))
    prefetch_seen: set[tuple[str, float]] = set()
    prefetch_jobs: list[tuple[str, float]] = []
    for item in chunk:
        pp = pdf_dir / item["file"] if item["file"] else None
        if pp and pp.is_file():
            mt = os.path.getmtime(pp)
            key = (str(pp), mt)
            if key not in prefetch_seen:
                prefetch_seen.add(key)
                prefetch_jobs.append(key)
    prefetch_thumbs_parallel(
        prefetch_jobs,
        scale,
        sharpness=sharp,
        max_raster_px=max_raster,
        max_workers=12,
    )

    for item in chunk:
        pdf_path = pdf_dir / item["file"] if item["file"] else None
        suffix = _widget_key_suffix(item["file"] or str(item["excel_row"]))
        rk = item["excel_row"]

        row = st.columns(colw, gap="small")

        with row[0]:
            fn = item["file"] or "—"
            st.markdown(
                f'<p class="pkg-fn" title="{_escape_attr(fn)}">{html.escape(fn)}</p>',
                unsafe_allow_html=True,
            )

        with row[1]:
            st.caption(_makety_short_txt(item.get("_cg_knife_name") or "", 110))
        with row[2]:
            st.caption(_makety_short_txt(item.get("_cg_category") or "", 80))
        with row[3]:
            st.caption(_makety_short_txt(item.get("_cg_lacquers") or "", 120))

        with row[4]:
            use_custom_key = f"use_custom_{suffix}_{rk}"
            sel_key = f"kind_sel_{suffix}_{rk}"
            cust_key = f"kind_cust_{suffix}_{rk}"
            lock_key = f"kind_lock_{suffix}_{rk}"

            opts = list(kind_options)
            if item["kind"] and item["kind"] not in opts:
                opts = sorted(set(opts) | {item["kind"]})

            if use_custom_key not in st.session_state:
                st.session_state[use_custom_key] = False

            _is_locked = bool(item.get("kind_locked"))
            if lock_key not in st.session_state:
                st.session_state[lock_key] = _is_locked

            ic1, ic2, ic3 = st.columns([0.15, 1, 0.18], gap="small")
            with ic1:
                st.checkbox(
                    "✎",
                    key=use_custom_key,
                    help="Свой вид",
                    on_change=_kind_use_custom_off_sync_selectbox,
                    args=(suffix, rk),
                )
            with ic2:
                if st.session_state[use_custom_key]:
                    if cust_key not in st.session_state:
                        st.session_state[cust_key] = item["kind"]
                    st.text_input(
                        "Вид",
                        key=cust_key,
                        label_visibility="collapsed",
                        placeholder="Вид",
                    )
                else:
                    if sel_key not in st.session_state:
                        st.session_state[sel_key] = (
                            item["kind"]
                            if item["kind"] in opts
                            else (opts[0] if opts else "")
                        )
                    st.selectbox(
                        "Вид",
                        options=opts,
                        key=sel_key,
                        label_visibility="collapsed",
                    )
            with ic3:
                st.checkbox(
                    "🔒" if st.session_state.get(lock_key) else "📌",
                    key=lock_key,
                    help=(
                        "Закрепить вид — сохранить во все хранилища. "
                        "Закреплённый вид не будет перезаписан при синхронизации из Excel."
                    ),
                )
                if st.session_state.get(lock_key) != _is_locked:
                    _new_lock = bool(st.session_state.get(lock_key))
                    _cur_kind = item["kind"]
                    item["kind_locked"] = 1 if _new_lock else 0
                    if db_path and db_path.is_file():
                        try:
                            _lk_conn = pkg_db.connect(db_path)
                            pkg_db.init_db(_lk_conn)
                            pkg_db.set_kind_locked(_lk_conn, rk, _cur_kind, locked=_new_lock)
                            _lk_conn.close()
                        except Exception:
                            pass
                    if excel_path and excel_path.is_file():
                        try:
                            save_one_row_to_excel(excel_path, item, db_path)
                        except Exception:
                            pass

        with row[5]:
            mtime: float | None = None
            if pdf_path and pdf_path.is_file():
                mtime = os.path.getmtime(pdf_path)
            if mtime is not None and pdf_path is not None:
                png = render_pdf_thumb(
                    str(pdf_path),
                    mtime,
                    scale,
                    sharpness=sharp,
                    max_raster_px=max_raster,
                )
                pv1, pv2 = st.columns([1, 0.4], gap="small")
                with pv1:
                    if png:
                        st.image(png, width=tw)
                    else:
                        st.caption("—")
                with pv2:
                    if st.button(
                        "PDF",
                        key=f"pdf_open_{rk}",
                        help="Открыть настоящий PDF в модальном окне",
                    ):
                        open_real_pdf_dialog(
                            str(pdf_path),
                            max_modal_bytes,
                            f"dlg_dl_{rk}",
                        )
            else:
                st.caption("—")

        with row[6]:
            with st.form(key=f"size_form_{rk}", border=False, clear_on_submit=False):
                sf1, sf2 = st.columns([1, 0.32], gap="small")
                with sf1:
                    sz = st.text_input(
                        "Размер",
                        value=item["size"],
                        label_visibility="collapsed",
                        placeholder="80 × 57 mm",
                    )
                with sf2:
                    submitted = st.form_submit_button(
                        "↵",
                        help="Enter в поле или клик — нормализация (пробелы, ×)",
                    )
            if submitted:
                item["size"] = canonicalize_size_mm(normalize_size(sz))
                ok = True
                try:
                    apply_makety_cg_derived_from_db(db_path, [item])
                    save_rows_to_db(db_path, [item])
                except Exception as e:
                    ok = False
                    st.toast(f"БД (размер): {e}", icon="⚠️")
                if ok:
                    try:
                        save_one_row_to_excel(excel_path, item, db_path=None)
                    except Exception as e:
                        st.toast(f"Excel (размер): {e}", icon="⚠️")
                    mirror = st.session_state.setdefault("_db_row_mirror", {})
                    mirror[int(rk)] = row_snapshot_for_mirror(item)
                st.rerun()

        with row[7]:
            st.caption(_makety_short_txt(item.get("_knife_size_mm") or "", 40))

        with row[8]:
            st.caption(_makety_short_txt(item.get("_cg_cutit_no") or "", 24))

        with row[9]:
            st.caption((item.get("price") or "").strip() or "—")

        with row[10]:
            pnew_key = f"price_new_{suffix}_{rk}"
            if pnew_key not in st.session_state:
                st.session_state[pnew_key] = item.get("price_new") or ""
            st.text_input(
                "Новая цена",
                key=pnew_key,
                label_visibility="collapsed",
                placeholder="—",
            )

        with row[11]:
            qps_key = f"qty_sheet_{suffix}_{rk}"
            if qps_key not in st.session_state:
                st.session_state[qps_key] = item.get("qty_per_sheet") or ""
            st.text_input(
                "На листе",
                key=qps_key,
                label_visibility="collapsed",
                placeholder="—",
            )

        with row[12]:
            qpy_key = f"qty_year_{suffix}_{rk}"
            if qpy_key not in st.session_state:
                st.session_state[qpy_key] = item.get("qty_per_year") or ""
            st.text_input(
                "За год",
                key=qpy_key,
                label_visibility="collapsed",
                placeholder="—",
            )

        st.markdown('<hr class="pkg-row-hr"/>', unsafe_allow_html=True)

    end_idx = min(start + per_page, len(filtered))
    shown_from = start + 1 if filtered else 0
    nav_b = st.columns([0.14, 0.14, 1], gap="small")
    with nav_b[0]:
        if st.button(
            "−",
            key="pkg_page_minus_end",
            disabled=page <= 1,
            help="Предыдущая страница",
        ):
            st.session_state["_pkg_page_jump"] = page - 1
            st.rerun()
    with nav_b[1]:
        if st.button(
            "+",
            key="pkg_page_plus_end",
            disabled=page >= total_pages,
            help="Следующая страница",
        ):
            st.session_state["_pkg_page_jump"] = page + 1
            st.rerun()
    with nav_b[2]:
        st.caption(
            f"Стр. {page} / {total_pages} · позиции {shown_from}–{end_idx} из {len(filtered)}"
        )

    st.divider()
    c1, c2 = st.columns([1, 3])
    with c1:
        if st.button("Сохранить в Excel и БД", type="primary", key="pkg_save_all_btn"):
            sync_widgets_to_rows(rows)
            for it in rows:
                it["size"] = canonicalize_size_mm(
                    normalize_size(it.get("size") or "")
                )
            try:
                apply_makety_cg_derived_from_db(db_path, rows)
                save_rows_to_db(db_path, rows)
                save_rows_to_excel(excel_path, rows, db_path=None)
                mirror = st.session_state.setdefault("_db_row_mirror", {})
                for r in rows:
                    mirror[int(r["excel_row"])] = row_snapshot_for_mirror(r)
                st.success(
                    f"Файл перезаписан (лишние строки ниже данных очищены), SQLite обновлён — "
                    f"{excel_path.name} · {db_path.name}"
                )
            except Exception as e:
                st.error(f"Ошибка сохранения: {e}")
    with c2:
        st.caption(
            "Вид, новая цена и количества при изменении сразу пишутся в SQLite и Excel; размер — после ↵. "
            "Колонка «Цена CG» и блок каталога CG считаются из SQLite (сопоставление ножа, прайс, контур в БД). "
            "Название для cutii хранится в **последнем столбце** Excel (столбец M), в таблице не показывается. "
            "Кнопка выше записывает все строки: временный .xlsx и атомарная замена; строки ниже последней excel_row очищаются."
        )


if __name__ == "__main__":
    main()
