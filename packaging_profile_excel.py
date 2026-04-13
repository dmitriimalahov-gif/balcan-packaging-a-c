#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Excel «профиль» размерной группы: шапка Balkan, таблица позиций, анализ, график."""

from __future__ import annotations

import io
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import packaging_db as pkg_db
from packaging_print_planning import SheetParams, size_key_display

ROOT = Path(__file__).resolve().parent
DEFAULT_LOGO_PATH = ROOT / "assets" / "balkan_header_logo.png"

# Широкая таблица: A–Z (26 столбцов)
_LAST_COL = 26
_LAST_LETTER = get_column_letter(_LAST_COL)

_FINISH_LABELS: dict[str, str] = {
    "lac_wb": "Lac WB (водный лак)",
    "uv_no_foil": "UV без фольги",
    "uv_foil": "UV с фольгой",
}

_MONTH_SHORT_RU = (
    "Янв",
    "Фев",
    "Мар",
    "Апр",
    "Май",
    "Июн",
    "Июл",
    "Авг",
    "Сен",
    "Окт",
    "Ноя",
    "Дек",
)

_THIN = Side(style="thin", color="000000")


def _parse_qty(val: str | None) -> float:
    if not val:
        return 0.0
    cleaned = str(val).replace(" ", "").replace("\u00a0", "").replace(",", ".")
    try:
        return float(cleaned)
    except (ValueError, TypeError):
        return 0.0


def _parse_qty_int(val: str | None) -> int:
    q = _parse_qty(val)
    if q <= 0:
        return 0
    return max(1, int(round(q)))


def _sheet_chars(sp: SheetParams) -> str:
    return (
        f"{sp.width_mm:g}×{sp.height_mm:g} мм; поле {sp.margin_mm:g} мм; "
        f"зазор X {sp.gap_mm:g} мм, Y {sp.gap_y_mm:g} мм"
    )


def _resolve_report_year(monthly: list[dict[str, Any]], report_year: int | None) -> int:
    if report_year is not None:
        return int(report_year)
    if monthly:
        return max(int(m["year"]) for m in monthly)
    return date.today().year


def _monthly_row_vector(
    er: int,
    year: int,
    monthly_by_er: dict[int, list[dict[str, Any]]],
) -> list[int]:
    """12 целых значений qty по месяцам 1..12 для заданного года."""
    by_month: dict[int, float] = {}
    for x in monthly_by_er.get(er, []):
        if int(x["year"]) != year:
            continue
        mo = int(x["month"])
        by_month[mo] = by_month.get(mo, 0.0) + float(x["qty"])
    return [int(round(by_month.get(m, 0.0))) for m in range(1, 13)]


def _enrich_group_rows(
    conn: Any,
    rows: list[dict[str, Any]],
    *,
    finish_code: str,
) -> list[dict[str, Any]]:
    """Добавляет поля _cg_* и _knife_size_mm (как в apply_makety), без записи price."""
    pkg_db.init_db(conn)
    cg_map = pkg_db.load_cg_mapping(conn)
    cg_knives = pkg_db.load_cg_knives(conn)
    knives_by = {k["cutit_no"]: k for k in cg_knives}
    cg_prices = pkg_db.load_cg_prices(conn)
    knife_meta = pkg_db.load_knives_meta(conn)
    finish_pref = ("lac_wb", "uv_no_foil", "uv_foil")
    out: list[dict[str, Any]] = []
    for item in rows:
        row = dict(item)
        er = int(row["excel_row"])
        cutit = ""
        m = cg_map.get(er)
        if m:
            cutit = (m.get("cutit_no") or "").strip()
        kinfo = knives_by.get(cutit) if cutit else None
        row["_cg_cutit_no"] = cutit
        row["_cg_knife_name"] = (kinfo.get("name") or "").strip() if kinfo else ""
        row["_cg_category"] = (kinfo.get("category") or "").strip() if kinfo else ""
        pr_c = [p for p in cg_prices if p["cutit_no"] == cutit]
        fts = sorted(set(str(p["finish_type"]) for p in pr_c))
        lac_labels = [_FINISH_LABELS.get(f, f) for f in fts]
        row["_cg_lacquers"] = ", ".join(lac_labels)
        km = knife_meta.get(er)
        w0 = float(km["width_mm"]) if km else 0.0
        h0 = float(km["height_mm"]) if km else 0.0
        if km and w0 > 0 and h0 > 0:
            row["_knife_size_mm"] = f"{w0:.1f} × {h0:.1f} мм"
        else:
            row["_knife_size_mm"] = ""
        qty = _parse_qty_int(row.get("qty_per_year") or "")
        if qty <= 0:
            qty = 1
        cg_eur: float | None = None
        if cutit and pr_c:
            ft_use = finish_code if any(p["finish_type"] == finish_code for p in pr_c) else None
            if ft_use is None:
                ft_use = next((f for f in finish_pref if any(p["finish_type"] == f for p in pr_c)), None)
            if ft_use is None and pr_c:
                ft_use = str(pr_c[0]["finish_type"])
            if ft_use:
                cg_eur = pkg_db.cg_price_for_qty(pr_c, ft_use, qty)
        row["_cg_eur_1000"] = cg_eur
        row["_cg_finish_used"] = finish_code
        out.append(row)
    return out


def _aggregate_monthly_group(
    monthly_rows: list[dict[str, Any]],
) -> list[tuple[str, float]]:
    """Сумма qty по (year, month) для всех строк группы."""
    buckets: dict[tuple[int, int], float] = {}
    for m in monthly_rows:
        key = (int(m["year"]), int(m["month"]))
        buckets[key] = buckets.get(key, 0.0) + float(m["qty"])
    ordered = sorted(buckets.keys())
    return [(f"{y:04d}-{mo:02d}", buckets[(y, mo)]) for y, mo in ordered]


def _format_qty_sheet_cell(raw: str | None) -> int | float | str:
    q = _parse_qty(raw)
    if q <= 0:
        return "—"
    if abs(q - round(q)) < 1e-6:
        return int(round(q))
    return round(q, 2)


def build_profile_workbook_bytes(
    *,
    db_path: Path,
    size_key: str,
    size_key_display_override: str | None,
    group_rows: list[dict[str, Any]],
    rows_by_er: dict[int, dict[str, Any]],
    sheet_params: SheetParams,
    document_code: str,
    finish_code: str = "lac_wb",
    logo_path: Path | None = None,
    report_year: int | None = None,
) -> bytes:
    """
    Собирает .xlsx в память.

    group_rows — элементы size_group['rows'] (excel_row, name, size, …).
    rows_by_er — полные строки из сессии для подстановки актуальных полей.
    report_year — год для столбцов Янв…Дек; None = max(year) в помесячных или текущий год.
    """
    if not group_rows:
        raise ValueError("group_rows пуст")

    merged: list[dict[str, Any]] = []
    for r in group_rows:
        er = int(r["excel_row"])
        full = dict(rows_by_er.get(er) or r)
        full.setdefault("excel_row", er)
        merged.append(full)

    conn = pkg_db.connect(db_path.expanduser().resolve())
    try:
        pkg_db.init_db(conn)
        enriched = _enrich_group_rows(conn, merged, finish_code=finish_code)
        ers = [int(r["excel_row"]) for r in enriched]
        monthly = pkg_db.load_monthly_for_rows(conn, ers)
        _st = pkg_db.load_stock(conn)
        stock_map = {str(k).strip().upper(): float(v) for k, v in _st.items() if k}
    finally:
        conn.close()

    eff_year = _resolve_report_year(monthly, report_year)

    sk_disp = size_key_display_override or size_key_display(size_key)
    sheet_line = _sheet_chars(sheet_params)

    wb = Workbook()
    ws = wb.active
    ws.title = "Профиль"

    # ── шапка (слияния на всю ширину таблицы A–Z) ──
    ws.merge_cells("A1:A3")
    ws.merge_cells(f"B1:O3")
    ws.merge_cells(f"P1:{_LAST_LETTER}3")
    ws.merge_cells(f"A4:{_LAST_LETTER}4")
    ws.merge_cells(f"A5:{_LAST_LETTER}5")

    _col_widths: dict[str, float] = {
        "A": 5.5,
        "B": 12.0,
        "C": 28.0,
        "D": 12.0,
        "E": 14.0,
        "F": 26.0,
        "G": 14.0,
        "H": 11.0,
        "I": 11.0,
        "J": 11.0,
        "K": 11.0,
        "L": 11.0,
        "M": 11.0,
        "N": 12.0,
    }
    for letter, w in _col_widths.items():
        ws.column_dimensions[letter].width = w
    for ci in range(15, _LAST_COL + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 7.5

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 16

    brand_blue = "1F4E79"
    gray_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    border_outer = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

    logo_p = logo_path if logo_path is not None else DEFAULT_LOGO_PATH
    if logo_p.is_file():
        try:
            img = XLImage(str(logo_p))
            img.width = min(140, img.width) if img.width else 120
            img.height = min(72, img.height) if img.height else 60
            ws.add_image(img, "A1")
        except Exception:
            pass

    cell_title = ws["B1"]
    cell_title.value = "SC Balkan Pharmaceuticals SRL"
    cell_title.font = Font(name="Calibri", size=18, bold=True, italic=True, color=brand_blue)
    cell_title.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for r in range(1, 4):
        for col_letter in (get_column_letter(c) for c in range(2, 16)):
            c = ws[f"{col_letter}{r}"]
            c.border = border_outer

    doc_code = (document_code or "").strip() or "—"
    ws["P1"].value = f"Codul documentului:\n{doc_code}"
    ws["P1"].font = Font(name="Calibri", size=11, bold=True)
    ws["P1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(1, 4):
        for ci in range(16, _LAST_COL + 1):
            col_letter = get_column_letter(ci)
            c = ws[f"{col_letter}{r}"]
            c.fill = gray_fill
            c.border = border_outer
            if col_letter == "P" and r == 1:
                continue
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cap = ws["A4"]
    cap.value = "Macheta originală a ambalajului primar"
    cap.font = Font(name="Calibri", size=9, italic=True, color="666666")
    cap.alignment = Alignment(horizontal="center", vertical="center")

    note_row = 5
    ncell = ws[f"A{note_row}"]
    ncell.value = (
        f"Размер группы: {sk_disp}. Лист: {sheet_line}. Помесячные столбцы — год {eff_year} "
        f"(из SQLite packaging_monthly_qty; заполняется импортом cutii, напр. «Balcan 2025 cutii.xlsx»). "
        f"€/1000 CG: отделка {_FINISH_LABELS.get(finish_code, finish_code)}, тираж из годового объёма строки. "
        "Кол-во на листе — из макетов (не из файла CG). Подрядчики 1–5 — вручную (например €/1000)."
    )
    ncell.font = Font(size=8, color="555555")
    ncell.alignment = Alignment(wrap_text=True, vertical="top")

    # ── таблица 1 (основная сетка) ──
    t1 = note_row + 2
    h1: list[str] = [
        "№",
        "GMP",
        "Наименование",
        "Вид",
        "Размер (мм)",
        "Нож CG",
        "Кол-во на листе (макеты)",
        "€/1000 CG",
        "Подрядчик 1",
        "Подрядчик 2",
        "Подрядчик 3",
        "Подрядчик 4",
        "Подрядчик 5",
        "Годовой объём (шт)",
    ]
    h1.extend(f"{m} {eff_year}" for m in _MONTH_SHORT_RU)

    for col, title in enumerate(h1, start=1):
        cell = ws.cell(row=t1, column=col, value=title)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="E2EFDA")
        cell.border = border_outer
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    monthly_by_er: dict[int, list[dict[str, Any]]] = {}
    for m in monthly:
        er = int(m["excel_row"])
        monthly_by_er.setdefault(er, []).append(m)

    for i, row in enumerate(enriched, start=1):
        r = t1 + i
        er = int(row["excel_row"])
        gmp = (row.get("gmp_code") or "").strip()
        if not gmp:
            gmp = pkg_db.extract_gmp_code(row.get("name") or "", row.get("file") or "")
        gmp_u = gmp.upper() if gmp else "—"

        cut = (row.get("_cg_cutit_no") or "").strip()
        nm = (row.get("_cg_knife_name") or "").strip()
        if cut and nm:
            knife_cell = f"{cut} — {nm[:100]}"
        elif cut:
            knife_cell = cut
        else:
            knife_cell = "—"

        cg_val = row.get("_cg_eur_1000")
        cg_str = f"{cg_val:.2f}" if cg_val is not None else "—"

        annual = _parse_qty(row.get("qty_per_year") or "")
        annual_cell = int(annual) if annual else "—"

        months_12 = _monthly_row_vector(er, eff_year, monthly_by_er)

        vals: list[Any] = [
            i,
            gmp_u,
            (row.get("name") or "")[:200],
            (row.get("kind") or "")[:80],
            (row.get("size") or "")[:80],
            knife_cell,
            _format_qty_sheet_cell(row.get("qty_per_sheet")),
            cg_str,
            "",
            "",
            "",
            "",
            "",
            annual_cell,
        ]
        vals.extend(months_12)

        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.border = border_outer
            c.alignment = Alignment(wrap_text=True, vertical="top")

    last_t1 = t1 + len(enriched)

    # ── таблица 2 «Анализ» ──
    t2 = last_t1 + 3
    ws.cell(row=t2, column=1, value="Анализ по позициям").font = Font(bold=True, size=12)
    t2 += 1
    h2 = [
        "GMP",
        "Наименование",
        "Годовой объём (шт)",
        "Сумма помесячных (БД)",
        "Среднее в месяц (по месяцам с данными)",
        "Остаток на складе (шт)",
        "Хватит на (мес.)",
        "Хватит на (дни)",
    ]
    for col, title in enumerate(h2, start=1):
        cell = ws.cell(row=t2, column=col, value=title)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="DDEBF7")
        cell.border = border_outer
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for i, row in enumerate(enriched, start=1):
        r = t2 + i
        er = int(row["excel_row"])
        gmp = (row.get("gmp_code") or "").strip()
        if not gmp:
            gmp = pkg_db.extract_gmp_code(row.get("name") or "", row.get("file") or "")
        gmp_u = gmp.upper() if gmp else ""
        stock = float(stock_map.get(gmp_u, 0.0)) if gmp_u else 0.0

        annual = _parse_qty(row.get("qty_per_year") or "")
        er_m = monthly_by_er.get(er, [])
        monthly_sum = sum(float(x["qty"]) for x in er_m)
        n_months = len({(x["year"], x["month"]) for x in er_m})
        avg_m = monthly_sum / n_months if n_months else 0.0

        monthly_rate = max(annual / 12.0, avg_m) if (annual > 0 or avg_m > 0) else 0.0
        months_cov = stock / monthly_rate if monthly_rate > 0 and stock > 0 else 0.0
        days_cov = months_cov * 30.4375 if months_cov > 0 else 0.0

        vals2 = [
            gmp_u or "—",
            (row.get("name") or "")[:200],
            int(annual) if annual else "—",
            int(monthly_sum) if monthly_sum else "—",
            round(avg_m, 1) if avg_m else "—",
            int(stock) if stock else "—",
            round(months_cov, 2) if months_cov > 0 else "—",
            int(round(days_cov)) if days_cov > 0 else "—",
        ]
        for col, v in enumerate(vals2, start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.border = border_outer
            c.alignment = Alignment(wrap_text=True, vertical="top")

    last_t2 = t2 + len(enriched)

    # ── скрытый лист + график ──
    agg = _aggregate_monthly_group(monthly)
    wsd = wb.create_sheet("__data", 1)
    wsd.sheet_state = "hidden"
    wsd["A1"] = "Период"
    wsd["B1"] = "Кол-во, шт"
    for idx, (label, qty) in enumerate(agg, start=2):
        wsd.cell(row=idx, column=1, value=label)
        wsd.cell(row=idx, column=2, value=qty)

    chart = LineChart()
    chart.title = "Динамика заказов (сумма по группе), шт/мес"
    chart.y_axis.title = "Шт."
    chart.x_axis.title = "Период"
    chart.height = 10
    chart.width = 18
    if len(agg) >= 1:
        max_r = 1 + len(agg)
        data_ref = Reference(wsd, min_col=2, min_row=1, max_row=max_r)
        cats_ref = Reference(wsd, min_col=1, min_row=2, max_row=max_r)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        anchor_row = last_t2 + 3
        ws.add_chart(chart, f"A{anchor_row}")
    else:
        ws.cell(row=last_t2 + 3, column=1, value="Нет помесячных данных для графика.")

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
