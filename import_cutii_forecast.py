#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Импорт cutii: колонка B ↔ name; при неудаче — PDF. Ручные подтверждения: cutii_confirmations.csv
# или --export-pending → заполнить confirmed_excel_row → повторный запуск.

from __future__ import annotations

import argparse
import csv
import sys
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from rapidfuzz import fuzz

import packaging_db as pkg_db

ROOT = Path(__file__).resolve().parent
DEFAULT_EXCEL = ROOT / "Упаковка_макеты.xlsx"
DEFAULT_CUTII = ROOT.parent / "Balcan 2025 cutii.xlsx"
DEFAULT_OVERRIDES = ROOT / "cutii_name_overrides.csv"
DEFAULT_CONFIRMATIONS = ROOT / "cutii_confirmations.csv"
DEFAULT_PENDING_EXPORT = ROOT / "cutii_pending_review.csv"


def normalize_for_match(s: str, *, strip_parens: bool = False) -> str:
    s = unicodedata.normalize("NFKC", (s or "").strip())
    s = s.lower()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("_", " ")
    s = s.replace("sol.inj", "sol inj")
    s = s.replace("anti-xa", "anti xa")
    s = re.sub(r"\s*/\s*", "/", s)
    if strip_parens:
        s = re.sub(r"\([^)]*\)", " ", s)
        s = re.sub(r"\s+", " ", s).strip()
    return s


def is_packaging_box(kind: str) -> bool:
    raw = (kind or "").strip()
    k = raw.lower()
    if "без вторичной" in k or "fara secundar" in k or "fara cutie" in k:
        return False
    if raw == "Коробка" or "короб" in k:
        return True
    return False


def is_printable_packaging(kind: str) -> bool:
    """Коробки, блистеры и пакеты — всё, что печатается на листах."""
    raw = (kind or "").strip()
    k = raw.lower()
    if "без вторичной" in k or "fara secundar" in k or "fara cutie" in k:
        return False
    if raw == "Коробка" or "короб" in k:
        return True
    if "блистер" in k or "blister" in k:
        return True
    if raw == "Пакет" or "пакет" in k:
        return True
    return False


def is_printable_label_kind(kind: str) -> bool:
    """Явная этикетка в поле «Вид» — может участвовать в раскладке на лист (вкладки «Печать» / «Планировщик»)."""
    raw = (kind or "").strip()
    k = raw.lower()
    if "без вторичной" in k or "fara secundar" in k or "fara cutie" in k:
        return False
    if raw == "Этикетка" or "этикет" in k:
        return True
    return False


def is_sheet_layout_candidate_kind(kind: str) -> bool:
    """Коробка, блистер, пакет или этикетка — кандидат на геометрию листа."""
    return is_printable_packaging(kind) or is_printable_label_kind(kind)


def _cell_float(ws: Any, row: int, col: int) -> float | None:
    v = ws.cell(row, col).value
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", ".").strip())
    except ValueError:
        return None


def _sum_months_2025(ws: Any, row: int) -> float:
    s = 0.0
    for c in range(4, 16):
        x = _cell_float(ws, row, c)
        if x is not None:
            s += x
    return s


def load_overrides(path: Path) -> dict[str, int]:
    if not path.is_file():
        return {}
    out: dict[str, int] = {}
    with path.open(encoding="utf-8-sig", newline="") as f:
        r = csv.DictReader(f)
        for row in r:
            key = (row.get("cutii_name_exact") or row.get("cutii") or "").strip()
            er = row.get("excel_row") or row.get("row")
            if not key or er is None or str(er).strip() == "":
                continue
            try:
                out[key] = int(er)
            except ValueError:
                continue
    return out


def load_confirmations(path: Path) -> tuple[dict[int, int], dict[int, str]]:
    """
    Ручные подтверждения: cutii_sheet_row → excel_row.
    Колонки: cutii_sheet_row, confirmed_excel_row (или excel_row);
    опционально cutii_name — если указано и не совпало с листом cutii, строка игнорируется.
    """
    mapping: dict[int, int] = {}
    name_expected: dict[int, str] = {}
    if not path.is_file():
        return mapping, name_expected
    with path.open(encoding="utf-8-sig", newline="") as f:
        r = csv.DictReader(f)
        for row in r:
            sr = row.get("cutii_sheet_row") or row.get("sheet_row") or row.get("row_cutii")
            er = row.get("confirmed_excel_row") or row.get("excel_row")
            if sr is None or er is None or str(er).strip() == "":
                continue
            try:
                sri = int(sr)
                eri = int(er)
            except ValueError:
                continue
            snap = (row.get("cutii_name") or "").strip()
            mapping[sri] = eri
            if snap:
                name_expected[sri] = snap
    return mapping, name_expected


def top_combined_candidates(
    cutii_raw: str,
    box_rows: list[dict[str, Any]],
    k: int = 5,
) -> list[tuple[int, int, int, dict[str, Any]]]:
    """До k кандидатов: (score_best, score_name, score_file, row)."""
    cutii_n = normalize_for_match(cutii_raw)
    cutii_alt = normalize_for_match(cutii_raw, strip_parens=True)
    scored: list[tuple[int, int, int, dict[str, Any]]] = []
    for r in box_rows:
        n = r.get("name") or ""
        f = r.get("file") or ""
        sn = max(score_cutii_to_name(cutii_n, n), score_cutii_to_name(cutii_alt, n))
        sf = 0
        if f.strip():
            sf = max(
                score_cutii_to_file(cutii_n, f),
                score_cutii_to_file(cutii_alt, f),
            )
        best = max(sn, sf)
        scored.append((best, sn, sf, r))
    scored.sort(key=lambda x: -x[0])
    return scored[:k]


def score_cutii_to_file(cutii_n: str, pdf_file: str) -> int:
    if not pdf_file.strip():
        return 0
    p = Path(pdf_file)
    stem = normalize_for_match(p.stem)
    base = normalize_for_match(p.name)
    return max(fuzz.token_set_ratio(cutii_n, stem), fuzz.token_set_ratio(cutii_n, base))


def score_cutii_to_name(cutii_n: str, name: str) -> int:
    return fuzz.token_set_ratio(cutii_n, normalize_for_match(name))


def _rank_by_name(
    cutii_n: str,
    cutii_alt: str,
    box_rows: list[dict[str, Any]],
) -> list[tuple[int, dict[str, Any]]]:
    scored: list[tuple[int, dict[str, Any]]] = []
    for r in box_rows:
        n = r.get("name") or ""
        s1 = score_cutii_to_name(cutii_n, n)
        s2 = score_cutii_to_name(cutii_alt, n)
        scored.append((max(s1, s2), r))
    scored.sort(key=lambda x: -x[0])
    return scored


def _rank_by_file(
    cutii_n: str,
    cutii_alt: str,
    box_rows: list[dict[str, Any]],
) -> list[tuple[int, dict[str, Any]]]:
    scored: list[tuple[int, dict[str, Any]]] = []
    for r in box_rows:
        f = r.get("file") or ""
        s1 = score_cutii_to_file(cutii_n, f)
        s2 = score_cutii_to_file(cutii_alt, f)
        scored.append((max(s1, s2), r))
    scored.sort(key=lambda x: -x[0])
    return scored


def _resolve_ranked(
    ranked: list[tuple[int, dict[str, Any]]],
    min_score: int,
    gap: int,
    via: str,
) -> tuple[str, dict[str, Any] | None, int, str]:
    if not ranked:
        return ("no_match", None, 0, via)
    top_s = ranked[0][0]
    if top_s < min_score:
        return ("no_match", None, top_s, via)
    best_s, best_r = ranked[0]
    second_s = ranked[1][0] if len(ranked) > 1 else -1
    if best_s - second_s < gap:
        return ("ambiguous", None, best_s, via)
    return ("ok", best_r, best_s, via)


def pick_match(
    cutii_raw: str,
    box_rows: list[dict[str, Any]],
    *,
    min_score: int,
    gap: int,
    fallback_pdf: bool,
) -> tuple[str, dict[str, Any] | None, int, str]:
    cutii_n = normalize_for_match(cutii_raw)
    cutii_alt = normalize_for_match(cutii_raw, strip_parens=True)

    ranked_n = _rank_by_name(cutii_n, cutii_alt, box_rows)
    st, row, sc, via = _resolve_ranked(ranked_n, min_score, gap, "name")
    if st in ("ok", "ambiguous"):
        return (st, row, sc, via)

    if not fallback_pdf:
        return (st, row, sc, "name")

    ranked_f = _rank_by_file(cutii_n, cutii_alt, box_rows)
    return _resolve_ranked(ranked_f, min_score, gap, "pdf")


def apply_cutii_import(
    to_apply: list[dict[str, Any]],
    db_path: Path,
    excel_path: Path,
    source_tag: str,
) -> None:
    """Обновляет qty_per_year и packaging_monthly_qty; колонка 8 в Excel упаковки."""
    now = datetime.now(timezone.utc).isoformat()
    conn = pkg_db.connect(db_path)
    try:
        pkg_db.init_db(conn)
        monthly_payload: list[dict[str, Any]] = []
        for item in to_apply:
            er = item["excel_row"]
            conn.execute(
                "UPDATE packaging_items SET qty_per_year = ?, updated_at = ? WHERE excel_row = ?",
                (item["qty_per_year"], now, er),
            )
            for m in item["monthly"]:
                monthly_payload.append(
                    {
                        "excel_row": er,
                        "year": m["year"],
                        "month": m["month"],
                        "qty": m["qty"],
                        "source": source_tag,
                    }
                )
        pkg_db.upsert_monthly_batch(conn, monthly_payload, default_source=source_tag)
        conn.commit()
    finally:
        conn.close()

    if excel_path.is_file():
        xwb = load_workbook(excel_path)
        xws = xwb.active
        for item in to_apply:
            xws.cell(row=item["excel_row"], column=8, value=item["qty_per_year"])
        xwb.save(excel_path)
        xwb.close()


def save_cutii_confirmations_csv(
    path: Path,
    entries: list[dict[str, Any]],
) -> None:
    """entries: cutii_sheet_row, confirmed_excel_row, опционально cutii_name."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(
            f,
            fieldnames=["cutii_sheet_row", "confirmed_excel_row", "cutii_name"],
            extrasaction="ignore",
        )
        w.writeheader()
        for e in entries:
            w.writerow(
                {
                    "cutii_sheet_row": e["cutii_sheet_row"],
                    "confirmed_excel_row": e["confirmed_excel_row"],
                    "cutii_name": e.get("cutii_name") or "",
                }
            )


def run_cutii_analysis(
    cutii_path: Path,
    all_rows: list[dict[str, Any]],
    overrides_path: Path,
    confirmations_path: Path,
    *,
    no_confirmations: bool = False,
    min_score: int = 50,
    ambiguous_gap: int = 5,
    fallback_pdf: bool = True,
    interactive: bool = False,
    confirmations_db_path: Path | None = None,
) -> dict[str, Any]:
    """
    Полный проход по листу cutii (buc.).
    Возвращает report_rows, to_apply, pending (с tops для UI), confirm_warnings, source_tag.
    Подтверждения: CSV, затем записи из БД (таблица cutii_confirmations) перекрывают совпадающие ключи.
    """
    box_rows = [r for r in all_rows if is_packaging_box(r.get("kind") or "")]
    overrides = load_overrides(overrides_path.expanduser().resolve())
    conf_path = confirmations_path.expanduser().resolve()
    if no_confirmations:
        confirmations: dict[int, int] = {}
        confirm_names: dict[int, str] = {}
    elif conf_path.is_file():
        confirmations, confirm_names = load_confirmations(conf_path)
    else:
        confirmations, confirm_names = {}, {}

    if not no_confirmations and confirmations_db_path is not None:
        dbp = confirmations_db_path.expanduser().resolve()
        if dbp.is_file():
            cdb = pkg_db.connect(dbp)
            try:
                pkg_db.init_db(cdb)
                db_m, db_n = pkg_db.load_cutii_confirmations(cdb)
                confirmations.update(db_m)
                confirm_names.update(db_n)
            finally:
                cdb.close()

    wb = load_workbook(cutii_path, read_only=True, data_only=True)
    ws = wb.active
    source_tag = cutii_path.name

    report_rows: list[dict[str, Any]] = []
    to_apply: list[dict[str, Any]] = []
    pending: list[dict[str, Any]] = []
    confirm_warnings = 0

    row_idx = 3
    while True:
        row_idx += 1
        name_cell = ws.cell(row_idx, 2).value
        if name_cell is None or str(name_cell).strip() == "":
            break
        unit = ws.cell(row_idx, 3).value
        if (unit or "").strip().lower() != "buc.":
            continue

        cutii_name = str(name_cell).strip()
        qty_r = _cell_float(ws, row_idx, 18)
        sum_m = _sum_months_2025(ws, row_idx)
        qty_year = int(round(qty_r)) if qty_r is not None else int(round(sum_m))

        monthly: list[dict[str, Any]] = []
        for m, c in enumerate(range(4, 16), start=1):
            q = _cell_float(ws, row_idx, c)
            monthly.append({"year": 2025, "month": m, "qty": float(q) if q is not None else 0.0})
        q_p = _cell_float(ws, row_idx, 16)
        if q_p is not None:
            monthly.append({"year": 2026, "month": 1, "qty": float(q_p)})

        matched: dict[str, Any] | None = None
        score = 0
        status = "no_match"
        detail = ""
        if cutii_name in overrides:
            er = overrides[cutii_name]
            matched = next((r for r in box_rows if int(r["excel_row"]) == er), None)
            if matched:
                status, score, detail = "ok", 100, "override"
            else:
                status = "no_match"
                detail = "override_bad_row"
        elif row_idx in confirmations:
            exp = confirm_names.get(row_idx)
            if exp and exp != cutii_name:
                status, matched, score, detail = "no_match", None, 0, "confirmed_name_mismatch"
                confirm_warnings += 1
                print(
                    f"Пропуск подтверждения строка листа {row_idx}: "
                    f"cutii_name в файле ≠ колонке B",
                    file=sys.stderr,
                )
            else:
                er = confirmations[row_idx]
                matched = next((r for r in box_rows if int(r["excel_row"]) == er), None)
                if matched:
                    status, score, detail = "ok", 100, "confirmed"
                else:
                    status, matched, score, detail = "no_match", None, 0, "confirmed_bad_row"
        else:
            status, matched, score, detail = pick_match(
                cutii_name,
                box_rows,
                min_score=min_score,
                gap=ambiguous_gap,
                fallback_pdf=fallback_pdf,
            )

        if (
            interactive
            and sys.stdin.isatty()
            and status in ("ambiguous", "no_match")
        ):
            tops_i = top_combined_candidates(cutii_name, box_rows, k=3)
            print(f"\n--- Лист cutii строка {row_idx} ---\n{cutii_name}\nСтатус: {status}")
            for i, (best, sn, sf, br) in enumerate(tops_i, start=1):
                print(
                    f"  {i}. excel_row={br['excel_row']} score={best} (name={sn} file={sf})\n"
                    f"     {br.get('name', '')[:100]}"
                )
            try:
                ans = input("Введите excel_row коробки или Enter (пропуск): ").strip()
            except EOFError:
                ans = ""
            if ans.isdigit():
                er_i = int(ans)
                m2 = next((r for r in box_rows if int(r["excel_row"]) == er_i), None)
                if m2:
                    matched = m2
                    status, score, detail = "ok", 99, "interactive"
                else:
                    print(f"  Нет коробки с excel_row={er_i}", file=sys.stderr)

        pdf_disp = (matched or {}).get("file") or ""
        report_rows.append(
            {
                "cutii_sheet_row": row_idx,
                "cutii_name": cutii_name,
                "excel_row": matched["excel_row"] if matched else "",
                "matched_pdf": pdf_disp,
                "score": score,
                "match_via": detail,
                "qty_per_year": qty_year if status == "ok" else "",
                "status": status,
            }
        )

        if status == "ok" and matched:
            to_apply.append(
                {
                    "excel_row": int(matched["excel_row"]),
                    "qty_per_year": str(qty_year),
                    "monthly": monthly,
                    "cutii_name": cutii_name,
                    "pdf": pdf_disp,
                }
            )
        elif status in ("ambiguous", "no_match"):
            tops_list: list[dict[str, Any]] = []
            for best, sn, sf, br in top_combined_candidates(cutii_name, box_rows, k=5):
                tops_list.append(
                    {
                        "excel_row": int(br["excel_row"]),
                        "score": best,
                        "score_name": sn,
                        "score_file": sf,
                        "name": (br.get("name") or "")[:220],
                        "pdf": (br.get("file") or "")[:160],
                    }
                )
            pending.append(
                {
                    "cutii_sheet_row": row_idx,
                    "cutii_name": cutii_name,
                    "qty_year": qty_year,
                    "monthly": monthly,
                    "status": status,
                    "score": score,
                    "match_via": detail,
                    "tops": tops_list,
                }
            )

    wb.close()
    return {
        "report_rows": report_rows,
        "to_apply": to_apply,
        "pending": pending,
        "confirm_warnings": confirm_warnings,
        "source_tag": source_tag,
        "box_rows": box_rows,
    }


def build_to_apply_with_ui_picks(
    auto_to_apply: list[dict[str, Any]],
    pending: list[dict[str, Any]],
    user_picks: dict[int, int],
    box_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Добавляет к auto_to_apply строки из pending, для которых в UI выбран excel_row (ключ — cutii_sheet_row)."""
    by_er = {int(r["excel_row"]): r for r in box_rows}
    out = list(auto_to_apply)
    seen: set[int] = {int(x["excel_row"]) for x in out}
    for p in pending:
        sr = int(p["cutii_sheet_row"])
        if sr not in user_picks:
            continue
        er = int(user_picks[sr])
        br = by_er.get(er)
        if not br:
            continue
        item = {
            "excel_row": er,
            "qty_per_year": str(p["qty_year"]),
            "monthly": p["monthly"],
            "cutii_name": p["cutii_name"],
            "pdf": (br.get("file") or ""),
        }
        if er in seen:
            out = [x for x in out if int(x["excel_row"]) != er]
        out.append(item)
        seen.add(er)
    return out


def flatten_pending_for_csv_export(pending: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Преобразует pending с tops в плоские строки для cutii_pending_review.csv."""
    out: list[dict[str, Any]] = []
    for p in pending:
        rec: dict[str, Any] = {
            "cutii_sheet_row": p["cutii_sheet_row"],
            "cutii_name": p["cutii_name"],
            "qty_per_year": p["qty_year"],
            "status": p["status"],
            "confirmed_excel_row": "",
        }
        for i in range(1, 4):
            rec[f"top{i}_excel_row"] = ""
            rec[f"top{i}_score"] = ""
            rec[f"top{i}_score_name"] = ""
            rec[f"top{i}_score_pdf"] = ""
            rec[f"top{i}_name"] = ""
            rec[f"top{i}_pdf"] = ""
        for i, t in enumerate(p.get("tops") or [], start=1):
            if i > 3:
                break
            rec[f"top{i}_excel_row"] = t["excel_row"]
            rec[f"top{i}_score"] = t["score"]
            rec[f"top{i}_score_name"] = t["score_name"]
            rec[f"top{i}_score_pdf"] = t["score_file"]
            rec[f"top{i}_name"] = t["name"][:200]
            rec[f"top{i}_pdf"] = t["pdf"][:120]
        out.append(rec)
    return out


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Импорт cutii → БД и Упаковка_макеты.xlsx",
        epilog=(
            "Ручное подтверждение: 1) python3 import_cutii_forecast.py --export-pending\n"
            "  2) В cutii_pending_review.csv в колонке confirmed_excel_row укажите excel_row.\n"
            "  3) Сохраните как cutii_confirmations.csv (или --confirmations ПУТЬ) и снова запустите импорт.\n"
            "Порядок приоритета: overrides → confirmations → автоматика → (--interactive)."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("--cutii", type=Path, default=DEFAULT_CUTII)
    ap.add_argument("--db", type=Path, default=pkg_db.DEFAULT_DB_PATH)
    ap.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    ap.add_argument("--overrides", type=Path, default=DEFAULT_OVERRIDES)
    ap.add_argument(
        "--confirmations",
        type=Path,
        default=DEFAULT_CONFIRMATIONS,
        help="CSV: cutii_sheet_row + confirmed_excel_row (+ опц. cutii_name)",
    )
    ap.add_argument(
        "--no-confirmations",
        action="store_true",
        help="Не загружать файл подтверждений",
    )
    ap.add_argument(
        "--export-pending",
        type=Path,
        nargs="?",
        const=DEFAULT_PENDING_EXPORT,
        default=None,
        metavar="CSV",
        help="Выгрузить ambiguous/no_match с топ-3 кандидатами и пустым confirmed_excel_row; без записи в БД",
    )
    ap.add_argument(
        "--interactive",
        action="store_true",
        help="В консоли запросить excel_row для строк ambiguous/no_match (только TTY)",
    )
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--min-score", type=int, default=50)
    ap.add_argument("--ambiguous-gap", type=int, default=5)
    ap.add_argument(
        "--no-fallback-pdf",
        action="store_true",
        help="Не сопоставлять по имени PDF, если по полю name не нашлось",
    )
    ap.add_argument("--report-dir", type=Path, default=ROOT)
    args = ap.parse_args()

    cutii_path = args.cutii.expanduser().resolve()
    db_path = args.db.expanduser().resolve()
    excel_path = args.excel.expanduser().resolve()

    if not cutii_path.is_file():
        print(f"Файл cutii не найден: {cutii_path}")
        return

    conn = pkg_db.connect(db_path)
    try:
        pkg_db.init_db(conn)
        all_rows = pkg_db.load_all(conn)
    finally:
        conn.close()

    conf_path = args.confirmations.expanduser().resolve()
    if not args.no_confirmations and conf_path.is_file():
        cm, _ = load_confirmations(conf_path)
        if cm:
            print(f"Подтверждения: {len(cm)} строк из {conf_path.name}")

    result = run_cutii_analysis(
        cutii_path,
        all_rows,
        args.overrides.expanduser().resolve(),
        args.confirmations.expanduser().resolve(),
        no_confirmations=args.no_confirmations,
        min_score=args.min_score,
        ambiguous_gap=args.ambiguous_gap,
        fallback_pdf=not args.no_fallback_pdf,
        interactive=args.interactive,
        confirmations_db_path=db_path,
    )
    report_rows = result["report_rows"]
    to_apply = result["to_apply"]
    confirm_warnings = result["confirm_warnings"]
    source_tag = result["source_tag"]
    pending_export = flatten_pending_for_csv_export(result["pending"])

    rep_csv = args.report_dir.expanduser().resolve() / "cutii_import_report.csv"
    with rep_csv.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "cutii_sheet_row",
                "cutii_name",
                "excel_row",
                "matched_pdf",
                "score",
                "match_via",
                "qty_per_year",
                "status",
            ],
        )
        w.writeheader()
        w.writerows(report_rows)

    n_ok = sum(1 for r in report_rows if r["status"] == "ok")
    n_amb = sum(1 for r in report_rows if r["status"] == "ambiguous")
    n_miss = sum(1 for r in report_rows if r["status"] == "no_match")
    print(f"Строк cutii (buc.): {len(report_rows)}")
    print(f"  сопоставлено: {n_ok}, неоднозначно: {n_amb}, нет пары: {n_miss}")
    print(f"Отчёт: {rep_csv}")
    if confirm_warnings:
        print(f"Предупреждений по подтверждениям (несовпадение cutii_name): {confirm_warnings}")

    if args.export_pending is not None:
        out_p = args.export_pending.expanduser().resolve()
        p_fields = [
            "cutii_sheet_row",
            "cutii_name",
            "qty_per_year",
            "status",
            "confirmed_excel_row",
        ]
        for i in range(1, 4):
            p_fields.extend(
                [
                    f"top{i}_excel_row",
                    f"top{i}_score",
                    f"top{i}_score_name",
                    f"top{i}_score_pdf",
                    f"top{i}_name",
                    f"top{i}_pdf",
                ]
            )
        with out_p.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=p_fields, extrasaction="ignore")
            w.writeheader()
            w.writerows(pending_export)
        print(f"На ручную проверку: {len(pending_export)} строк → {out_p}")
        print(
            "Заполните confirmed_excel_row (и при необходимости cutii_name для контроля), "
            "сохраните как cutii_confirmations.csv и снова запустите импорт."
        )
        return

    if args.dry_run:
        print("--dry-run: БД и Excel не менялись.")
        return

    if not to_apply:
        print("Нечего записывать.")
        return

    apply_cutii_import(to_apply, db_path, excel_path, source_tag)
    print(f"SQLite: qty_per_year и помесячно для {len(to_apply)} позиций.")
    if excel_path.is_file():
        print(f"Excel: колонка «Кол-во за год» для {len(to_apply)} строк.")
    else:
        print(f"Excel не найден ({excel_path}), только БД.")


if __name__ == "__main__":
    main()
