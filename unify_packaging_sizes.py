#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Унифицирует габариты в SQLite и Excel: перестановки одних и тех же мм (например 84×15×62 и 15×62×84)
приводятся к одному виду — по убыванию, например «84 × 62 × 15 mm».

  python3 unify_packaging_sizes.py              # применить изменения
  python3 unify_packaging_sizes.py --dry-run    # только отчёт

Перед первым запуском сделайте копию Excel и packaging_data.db.
"""

from __future__ import annotations

import argparse
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

import packaging_db as pkg_db
from packaging_sizes import canonicalize_size_mm, normalize_size

ROOT = Path(__file__).resolve().parent
DEFAULT_EXCEL = ROOT / "Упаковка_макеты.xlsx"


def main() -> None:
    ap = argparse.ArgumentParser(description="Канонизация размеров в БД и Excel")
    ap.add_argument("--db", type=Path, default=pkg_db.DEFAULT_DB_PATH)
    ap.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    ap.add_argument("--dry-run", action="store_true", help="Не записывать, только отчёт")
    args = ap.parse_args()

    db_path: Path = args.db
    excel_path: Path = args.excel

    changes: list[tuple[int, str, str]] = []

    if excel_path.is_file():
        wb = load_workbook(excel_path)
        ws = wb.active
        for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None or all(v is None for v in row):
                continue
            old = "" if len(row) < 2 or row[1] is None else str(row[1]).strip()
            if not old:
                continue
            new = canonicalize_size_mm(normalize_size(old))
            if new != old:
                changes.append((excel_row, old, new))

    db_changes: list[tuple[int, str, str]] = []
    if db_path.is_file():
        conn = pkg_db.connect(db_path)
        try:
            pkg_db.init_db(conn)
            rows = pkg_db.load_all(conn)
            for r in rows:
                old = (r.get("size") or "").strip()
                if not old:
                    continue
                new = canonicalize_size_mm(normalize_size(old))
                if new != old:
                    db_changes.append((int(r["excel_row"]), old, new))
        finally:
            conn.close()

    # объединить по excel_row для отчёта
    by_row: dict[int, tuple[str, str]] = {}
    for er, old, new in changes:
        by_row[er] = (old, new)
    for er, old, new in db_changes:
        if er not in by_row:
            by_row[er] = (old, new)

    print(f"Excel: {excel_path} ({'найден' if excel_path.is_file() else 'нет файла'})")
    print(f"БД:    {db_path} ({'найден' if db_path.is_file() else 'нет файла'})")
    print(f"Строк с изменением размера (Excel): {len(changes)}")
    print(f"Строк с изменением размера (БД):    {len(db_changes)}")
    if by_row:
        print("\nПримеры (до → после):")
        for er, (old, new) in sorted(by_row.items())[:25]:
            print(f"  row {er}: {old!r} → {new!r}")
        if len(by_row) > 25:
            print(f"  … и ещё {len(by_row) - 25} строк")

    if args.dry_run:
        print("\n--dry-run: запись отключена.")
        return

    if changes and excel_path.is_file():
        wb = load_workbook(excel_path)
        ws = wb.active
        for excel_row, old, new in changes:
            ws.cell(row=excel_row, column=2, value=new or None)
        wb.save(excel_path)
        print(f"\nЗаписано в Excel: {len(changes)} ячеек.")

    if db_changes and db_path.is_file():
        conn = pkg_db.connect(db_path)
        try:
            pkg_db.init_db(conn)
            now = datetime.now(timezone.utc).isoformat()
            for er, _old, new in db_changes:
                conn.execute(
                    "UPDATE packaging_items SET size = ?, updated_at = ? WHERE excel_row = ?",
                    (new, now, er),
                )
            conn.commit()
            print(f"Записано в БД: {len(db_changes)} строк.")
        finally:
            conn.close()

    if not changes and not db_changes:
        print("\nИзменений не требуется.")


if __name__ == "__main__":
    main()
