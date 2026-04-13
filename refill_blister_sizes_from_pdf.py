#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Только позиции «Блистер»: заново извлекает размер (мм) из текста внутри PDF
и обновляет ячейку «Размер (мм)» в Excel (макет v3, 13 столбцов) и SQLite.

  python3 refill_blister_sizes_from_pdf.py --dry-run
  python3 refill_blister_sizes_from_pdf.py
  python3 refill_blister_sizes_from_pdf.py --pages 10
"""

from __future__ import annotations

import argparse
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

import packaging_db as pkg_db
from packaging_pdf_sizes import (
    canonicalize_extracted_size_text,
    extract_blister_flat_size_from_text,
    extract_text_from_pdf,
)
from packaging_sizes import (
    canonicalize_size_mm,
    extract_gabarit_mm_values,
    normalize_size,
)

ROOT = Path(__file__).resolve().parent

# Как в packaging_viewer.HEADERS
MAKETY_HEADERS = (
    "Название (нож CG)",
    "Категория (CG)",
    "Лаки (CG)",
    "Размер (мм)",
    "Вид",
    "Исходный PDF",
    "Размер ножа (мм)",
    "Нож CG",
    "Цена (текущая)",
    "Новая цена",
    "Кол-во на листе",
    "Кол-во за год",
    "Название (cutii)",
)


def _normalize_header_title(val: Any) -> str:
    if val is None:
        return ""
    s = str(val).replace("\u00a0", " ").strip()
    s = re.sub(r"\s+", " ", s)
    return s.casefold()


def _build_makety_column_index_map(hdr_t: tuple[Any, ...]) -> list[int] | None:
    norm_expected = [_normalize_header_title(h) for h in MAKETY_HEADERS]
    norm_actual = [_normalize_header_title(x) for x in hdr_t]
    used: set[int] = set()
    out: list[int] = []
    for exp in norm_expected:
        found: int | None = None
        for j, act in enumerate(norm_actual):
            if j in used or act != exp:
                continue
            found = j
            break
        if found is None:
            return None
        used.add(found)
        out.append(found)
    return out


def _cell_str(row_t: tuple[Any, ...], idx: int) -> str:
    if row_t is None or len(row_t) <= idx:
        return ""
    v = row_t[idx]
    return "" if v is None else str(v).strip()


def _is_blister(kind: str) -> bool:
    k = (kind or "").strip().lower()
    return "блистер" in k or "blister" in k


def _canonical_old_size(s: str) -> str:
    if not (s or "").strip():
        return ""
    try:
        return canonicalize_size_mm(normalize_size(s.strip()))
    except Exception:
        return s.strip()


def _size_dim_count(sz: str) -> int:
    """Число габаритов в канонической строке (разделитель ×)."""
    if not (sz or "").strip():
        return 0
    part = (sz or "").lower().replace("mm", "").strip()
    return part.count("×") + 1


def _sanitize_blister_size_string(sz: str) -> str:
    """
    Убирает из канонической строки числа в полосе 29–33.5 мм (типичный шум ±2 в PDF),
    затем снова канонизирует. Пара «80 × 30.8 mm» → «80 mm»; «80 × 57 × 30.8 mm» → «80 × 57 mm».
    """
    if not (sz or "").strip():
        return ""
    vals = extract_gabarit_mm_values(sz)
    if not vals:
        return sz.strip()
    lo, hi = 29.0, 33.5
    cleaned = [v for v in vals if not (lo <= v <= hi)]
    if len(cleaned) == len(vals):
        return sz.strip()
    if len(cleaned) >= 2:
        inner = "×".join(
            str(int(v)) if abs(v - round(v)) < 0.05 else str(v) for v in cleaned
        )
        return canonicalize_size_mm(normalize_size(inner + " mm"))
    if len(cleaned) == 1:
        v = cleaned[0]
        s = str(int(v)) if abs(v - round(v)) < 0.05 else str(v)
        return canonicalize_size_mm(normalize_size(s + " mm"))
    return ""


def main() -> None:
    ap = argparse.ArgumentParser(description="Размеры блистеров из PDF → Excel + БД")
    ap.add_argument("--excel", type=Path, default=ROOT / "Упаковка_макеты.xlsx")
    ap.add_argument("--pdf-dir", type=Path, default=ROOT)
    ap.add_argument("--db", type=Path, default=pkg_db.DEFAULT_DB_PATH)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument(
        "--pages",
        type=int,
        default=8,
        help="Сколько первых страниц PDF читать при извлечении текста",
    )
    ap.add_argument(
        "--force",
        action="store_true",
        help="Записывать даже если размер в таблице совпадает с извлечённым",
    )
    ap.add_argument(
        "--max-dims",
        type=int,
        default=5,
        help="Не записывать, если из PDF получилось больше стольких чисел габарита (защита от шума; капсулы часто 4–5 чисел)",
    )
    args = ap.parse_args()

    excel_path = args.excel.expanduser().resolve()
    pdf_dir = args.pdf_dir.expanduser().resolve()
    db_path = args.db.expanduser().resolve()

    if not excel_path.is_file():
        print(f"Нет файла Excel: {excel_path}")
        return

    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    hdr_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    hdr_t = tuple(hdr_row) if hdr_row else tuple()
    cmap = _build_makety_column_index_map(hdr_t)
    if cmap is None:
        print(
            "Ошибка: в первой строке нет полного набора заголовков эталона «Макеты» (13 столбцов). "
            "Приведите файл к эталону или проверьте имена столбцов."
        )
        wb.close()
        return

    size_col_1based = cmap[3] + 1
    kind_idx = cmap[4]
    file_idx = cmap[5]

    updates: list[tuple[int, str, str, str]] = []
    skipped_not_blister = 0
    skipped_no_file: list[tuple[int, str]] = []
    skipped_no_size: list[tuple[int, str]] = []
    skipped_too_many_dims: list[tuple[int, str, str]] = []
    unchanged = 0

    max_row = int(ws.max_row or 1)
    n_col = max(int(ws.max_column or 0), max(cmap) + 1)
    for excel_row in range(2, max_row + 1):
        row_cells = tuple(ws.cell(row=excel_row, column=j + 1).value for j in range(n_col))
        if all(v is None for v in row_cells):
            continue
        kind = _cell_str(row_cells, kind_idx)
        file_val = _cell_str(row_cells, file_idx)
        if not _is_blister(kind):
            skipped_not_blister += 1
            continue
        if not file_val:
            continue
        pdf_path = pdf_dir / file_val
        if not pdf_path.is_file():
            skipped_no_file.append((excel_row, file_val))
            continue
        pdf_text = extract_text_from_pdf(pdf_path, max_pages=int(args.pages))
        new_sz = extract_blister_flat_size_from_text(pdf_text)
        if not new_sz:
            new_sz = canonicalize_extracted_size_text(pdf_text)
        new_sz = _sanitize_blister_size_string(new_sz)
        if not new_sz:
            skipped_no_size.append((excel_row, file_val))
            continue
        if _size_dim_count(new_sz) > int(args.max_dims):
            skipped_too_many_dims.append((excel_row, file_val, new_sz))
            continue
        old_raw = _cell_str(row_cells, cmap[3])
        old_c = _canonical_old_size(old_raw)
        if not args.force and old_c == new_sz:
            unchanged += 1
            continue
        updates.append((excel_row, file_val, old_raw, new_sz))

    print(f"Строк не-блистер (пропуск): {skipped_not_blister}")
    print(f"Блистеров без PDF на диске: {len(skipped_no_file)}")
    print(f"Блистеров, размер из PDF не распознан: {len(skipped_no_size)}")
    print(
        f"Пропуск (слишком много габаритов в разборе, >{int(args.max_dims)}): "
        f"{len(skipped_too_many_dims)}"
    )
    print(f"Без изменений (уже совпадает с PDF): {unchanged}")
    print(f"К обновлению: {len(updates)}")
    for er, fn, old_s, new_s in updates[:45]:
        o = (old_s or "—")[:40]
        fn_short = fn[:50] + "…" if len(fn) > 50 else fn
        print(f"  row {er} | {fn_short}")
        print(f"         {o!r} → {new_s!r}")
    if len(updates) > 45:
        print(f"  … ещё {len(updates) - 45}")

    if skipped_no_file[:12]:
        print("Примеры: PDF не найден:")
        for er, fn in skipped_no_file[:12]:
            print(f"  row {er} {fn[:70]}")

    if skipped_too_many_dims[:8]:
        print("Примеры: слишком много чисел в кандидате из PDF:")
        for er, fn, sz in skipped_too_many_dims[:8]:
            print(f"  row {er} {fn[:50]}… → {sz[:80]}")

    if args.dry_run:
        print("--dry-run: запись отключена.")
        wb.close()
        return

    now = datetime.now(timezone.utc).isoformat()
    for excel_row, _fn, _old, new_sz in updates:
        ws.cell(row=excel_row, column=size_col_1based, value=new_sz)
    wb.save(excel_path)
    wb.close()
    print(f"Excel обновлён: {len(updates)} ячеек «Размер (мм)».")

    if updates and db_path.is_file():
        conn = pkg_db.connect(db_path)
        try:
            pkg_db.init_db(conn)
            for excel_row, _fn, _old, new_sz in updates:
                conn.execute(
                    "UPDATE packaging_items SET size = ?, updated_at = ? WHERE excel_row = ?",
                    (new_sz, now, excel_row),
                )
            conn.commit()
            print(f"SQLite обновлена: {len(updates)} строк.")
        finally:
            conn.close()
    elif updates:
        print("БД не найдена — только Excel.")


if __name__ == "__main__":
    main()
