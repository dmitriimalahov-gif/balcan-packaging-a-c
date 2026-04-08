#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Строки с пустым size в SQLite → повторное чтение PDF → UPDATE БД (и ячейка в Excel).

  python3 refill_db_missing_sizes_from_pdf.py --dry-run
  python3 refill_db_missing_sizes_from_pdf.py
"""

from __future__ import annotations

import argparse
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

import packaging_db as pkg_db
from packaging_pdf_sizes import canonicalize_extracted_size_text, extract_text_from_pdf

ROOT = Path(__file__).resolve().parent
DEFAULT_EXCEL = ROOT / "Упаковка_макеты.xlsx"


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", type=Path, default=pkg_db.DEFAULT_DB_PATH)
    ap.add_argument("--pdf-dir", type=Path, default=ROOT)
    ap.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument(
        "--pages",
        type=int,
        default=6,
        help="Сколько первых страниц PDF читать (по умолчанию 6)",
    )
    args = ap.parse_args()

    db_path = args.db.expanduser().resolve()
    pdf_dir = args.pdf_dir.expanduser().resolve()
    excel_path = args.excel.expanduser().resolve()

    if not db_path.is_file():
        print(f"БД не найдена: {db_path}")
        return

    conn = pkg_db.connect(db_path)
    try:
        pkg_db.init_db(conn)
        all_rows = pkg_db.load_all(conn)
    finally:
        conn.close()

    missing = [
        r
        for r in all_rows
        if not (r.get("size") or "").strip() and (r.get("file") or "").strip()
    ]

    updates: list[tuple[int, str, str]] = []
    skipped_no_file: list[str] = []
    skipped_no_text: list[tuple[int, str]] = []

    for r in missing:
        fn = (r.get("file") or "").strip()
        er = int(r["excel_row"])
        pdf_path = pdf_dir / fn
        if not pdf_path.is_file():
            skipped_no_file.append(fn)
            continue
        text = extract_text_from_pdf(pdf_path, max_pages=args.pages)
        new_size = canonicalize_extracted_size_text(text)
        if new_size:
            updates.append((er, fn, new_size))
        else:
            skipped_no_text.append((er, fn))

    print(f"В БД без размера (с именем PDF): {len(missing)}")
    print(f"Удалось извлечь размер: {len(updates)}")
    for er, fn, sz in updates[:50]:
        tail = "…" if len(fn) > 58 else ""
        print(f"  row {er} {fn[:58]}{tail} → {sz}")
    if len(updates) > 50:
        print(f"  … ещё {len(updates) - 50}")

    if skipped_no_file:
        print(f"PDF нет на диске ({len(skipped_no_file)}): показано до 15")
        for fn in skipped_no_file[:15]:
            print(f"  — {fn[:70]}")

    if skipped_no_text and not updates:
        print(f"Размер не распознан ({len(skipped_no_text)} файлов открыты, текста/шаблонов мало)")

    if args.dry_run:
        print("--dry-run: запись отключена.")
        return

    now = datetime.now(timezone.utc).isoformat()

    if updates and excel_path.is_file():
        wb = load_workbook(excel_path)
        ws = wb.active
        for excel_row, _fn, new_size in updates:
            ws.cell(row=excel_row, column=2, value=new_size)
        wb.save(excel_path)
        wb.close()
        print(f"Excel обновлён: {len(updates)} ячеек.")
    elif updates:
        print("Excel не найден — обновляю только БД.")

    if updates:
        conn = pkg_db.connect(db_path)
        try:
            pkg_db.init_db(conn)
            for excel_row, _fn, new_size in updates:
                conn.execute(
                    "UPDATE packaging_items SET size = ?, updated_at = ? WHERE excel_row = ?",
                    (new_size, now, excel_row),
                )
            conn.commit()
            print(f"SQLite обновлена: {len(updates)} строк.")
        finally:
            conn.close()


if __name__ == "__main__":
    main()
