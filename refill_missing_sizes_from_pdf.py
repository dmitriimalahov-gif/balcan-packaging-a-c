#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Повторно извлекает размеры из PDF для строк Excel, где столбец «Размер» пустой.
Использует расширенные шаблоны (packaging_pdf_sizes) и обновляет Excel + SQLite.

  python3 refill_missing_sizes_from_pdf.py --dry-run
  python3 refill_missing_sizes_from_pdf.py
"""

from __future__ import annotations

import argparse
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

import packaging_db as pkg_db
from packaging_pdf_sizes import canonicalize_extracted_size_text, extract_text_from_pdf

ROOT = Path(__file__).resolve().parent
DEFAULT_EXCEL = ROOT / "Упаковка_макеты.xlsx"


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    ap.add_argument("--pdf-dir", type=Path, default=ROOT)
    ap.add_argument("--db", type=Path, default=pkg_db.DEFAULT_DB_PATH)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--pages", type=int, default=4, help="Сколько первых страниц PDF читать")
    args = ap.parse_args()

    pdf_dir: Path = args.pdf_dir.expanduser().resolve()
    excel_path: Path = args.excel.expanduser().resolve()
    db_path: Path = args.db.expanduser().resolve()

    if not excel_path.is_file():
        print(f"Нет файла Excel: {excel_path}")
        return

    wb = load_workbook(excel_path)
    ws = wb.active
    updates: list[tuple[int, str, str]] = []

    for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None for v in row):
            continue
        size_cell = row[1] if len(row) > 1 else None
        size_s = "" if size_cell is None else str(size_cell).strip()
        if size_s:
            continue
        pdf_name = row[3] if len(row) > 3 else None
        if not pdf_name or not str(pdf_name).strip():
            continue
        fn = str(pdf_name).strip()
        pdf_path = pdf_dir / fn
        if not pdf_path.is_file():
            continue
        text = extract_text_from_pdf(pdf_path, max_pages=args.pages)
        new_size = canonicalize_extracted_size_text(text)
        if new_size:
            updates.append((excel_row, fn, new_size))

    print(f"Строк без размера, удалось извлечь размер: {len(updates)}")
    for er, fn, sz in updates[:40]:
        print(f"  row {er} {fn[:56]}… → {sz}" if len(fn) > 56 else f"  row {er} {fn} → {sz}")
    if len(updates) > 40:
        print(f"  … ещё {len(updates) - 40}")

    if args.dry_run:
        print("--dry-run: запись отключена.")
        wb.close()
        return

    now = datetime.now(timezone.utc).isoformat()
    for excel_row, _fn, new_size in updates:
        ws.cell(row=excel_row, column=2, value=new_size)
    wb.save(excel_path)
    wb.close()
    print(f"Excel обновлён: {len(updates)} ячеек.")

    if updates and db_path.is_file():
        conn = pkg_db.connect(db_path)
        try:
            pkg_db.init_db(conn)
            for excel_row, _fn, new_size in updates:
                conn.execute(
                    "UPDATE packaging_items SET size = ?, updated_at = ? WHERE excel_row = ?",
                    (new_size, now, excel_row),
                )
            conn.commit()
            print(f"SQLite обновлена: {len(updates)} строк.")
        finally:
            conn.close()
    elif updates and not db_path.is_file():
        print("БД не найдена — только Excel.")


if __name__ == "__main__":
    main()
